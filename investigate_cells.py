
import json
import openpyxl
from pathlib import Path

ranges = [
    'AA36:AA140', 'AB36:AB140', 'AC36:AC140', 'AD36:AD140', 'AE36:AE140', 'AF37:AF141',
    'BD27:BD131', 'D9:D140', 'H36:H140', 'I36:I140', 'J36:J140', 'K36:K140', 'L36:L140',
    'M36:M140', 'N36:N140', 'O36:O140', 'P36:P140', 'Q36:Q140', 'R36:R140', 'S36:S140',
    'T36:T140', 'U36:U140', 'V36:V140', 'W36:W140', 'X36:X140', 'Y36:Y140', 'Z36:Z140'
]

def get_representative_cells(rng):
    cells = openpyxl.utils.cell.range_boundaries(rng)
    # Check start, middle, end
    start_row = cells[1]
    end_row = cells[3]
    col = openpyxl.utils.cell.get_column_letter(cells[0])
    return [f"{col}{start_row}", f"{col}{(start_row + end_row) // 2}", f"{col}{end_row}"]

sheet_name = 'PV_stress_com'

audit_path = Path('src/configs/2025-08-12/enrichment_audit.json')
template_path = Path('workbooks/lic-dsf-template-2025-08-12.xlsm')
uga_path = Path('workbooks/dsf-uga.xlsm')

with open(audit_path, 'r') as f:
    audit_data = json.load(f)

def find_label(sheet, cell):
    sheet_data = audit_data.get('by_sheet', {}).get(sheet, {})
    cells = sheet_data.get('cells', [])
    for entry in cells:
        if entry.get('address') == cell:
            return {
                'row': entry.get('row_labels'),
                'col': entry.get('column_labels')
            }
    return None

results = {}

wb_template = openpyxl.load_workbook(template_path, data_only=True)
ws_template = wb_template[sheet_name]

wb_uga = openpyxl.load_workbook(uga_path, data_only=True)
ws_uga = wb_uga[sheet_name]

for rng in ranges:
    rep_cells = get_representative_cells(rng)
    range_results = []
    for cell in rep_cells:
        label = find_label(sheet_name, cell)
        template_val = ws_template[cell].value
        uga_val = ws_uga[cell].value
        range_results.append({
            'cell': cell,
            'label': label,
            'template': template_val,
            'uga': uga_val
        })
    results[rng] = range_results

print(json.dumps(results, indent=2))
