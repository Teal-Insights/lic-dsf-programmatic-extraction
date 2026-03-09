#!/usr/bin/env python3
"""
Central configuration for LIC-DSF programmatic extraction.

This module owns:
- Workbook paths / download URL
- Explicit export/annotation target ranges
- Helpers for expanding configured ranges into dependency-graph targets

Dynamic refs (OFFSET/INDIRECT) are resolved via a constraint-based config.
Iterative workflow: run the export script; if DynamicRefError is raised, the
message includes the formula cell that needs a constraint. Inspect that cell
and the row/column headers in the workbook to decide plausible input domains,
add the address to LicDsfConstraints (with Annotated[int, Between(lo, hi)] or
Literal[...]), then re-run until the graph
builds.
"""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path
from typing import Annotated, Literal, TypedDict
from urllib.request import urlopen

import openpyxl.utils.cell
from excel_grapher import format_cell_key
from excel_grapher.grapher import DynamicRefConfig
from excel_grapher.grapher.dynamic_refs import FromWorkbook


class ExportRangeConfig(TypedDict):
    """
    Explicit range specification for export/annotation targets.

    Attributes:
        label: Human-readable label for the range (used for reporting only).
        range_spec: Sheet-qualified A1 range, e.g. "'Chart Data'!D10:D17".
        entrypoint_mode: Controls how export entrypoints are grouped for this
            range: "row_group" (one entrypoint per row) or "per_cell" (one
            entrypoint per cell, no row grouping).
    """

    label: str
    range_spec: str
    entrypoint_mode: Literal["row_group", "per_cell"]


# Default workbook used for dependency mapping / enrichment.
WORKBOOK_PATH = Path("workbooks/lic-dsf-template-2026-01-31.xlsm")
WORKBOOK_TEMPLATE_URL = (
    "https://thedocs.worldbank.org/en/doc/f0ade6bcf85b6f98dbeb2c39a2b7770c-0360012025/new-lic-dsf-template"
)


def ensure_workbook_available(
    path: Path = WORKBOOK_PATH, url: str | None = None
) -> bool:
    """
    Ensure the default LIC-DSF template workbook exists locally.

    If the workbook is missing, downloads it from `url` (or `WORKBOOK_TEMPLATE_URL`) into `path`.
    """
    if path.exists() and path.stat().st_size > 0:
        return True

    path.parent.mkdir(parents=True, exist_ok=True)

    try:
        source_url = url or WORKBOOK_TEMPLATE_URL
        with urlopen(source_url, timeout=60) as resp:
            with tempfile.NamedTemporaryFile(
                prefix=f".{path.name}.", suffix=".download", dir=str(path.parent), delete=False
            ) as tmp:
                shutil.copyfileobj(resp, tmp)
                tmp_path = Path(tmp.name)

        if tmp_path.stat().st_size == 0:
            tmp_path.unlink(missing_ok=True)
            return False

        tmp_path.replace(path)
        return True
    except Exception:
        return False


# Row labels for the multi-row stress-test blocks (same row layout in each block).
# Blank string means that row is skipped when splitting by row.
STRESS_TEST_ROW_LABELS: list[str] = [
    "Baseline",
    "A1. Key variables at their historical averages in 2024-2034 2/",
    "B1. Real GDP growth",
    "B2. Primary balance",
    "B3. Exports",
    "B4. Other flows 3/",
    "B5. Depreciation",
    "B6. Combination of B1-B5",
    "",
    "C1. Combined contingent liabilities",
    "C2. Natural disaster",
    "C3. Commodity price",
    "C4. Market Financing",
    "A2. Alternative Scenario :[Customize, enter title]",
]


STRESS_TEST_BLOCKS: list[tuple[str, int]] = [
    ("PV of Debt-to-GDP Ratio", 239),
    ("PV of Debt-to-Revenue Ratio", 281),
    ("Debt Service-to-Revenue Ratio", 318),
    ("Debt Service-to-GDP Ratio", 351),
]


EXPORT_FIXED_RANGES: list[ExportRangeConfig] = [
    {
        "label": "External DSA risk rating signals",
        "range_spec": "'Chart Data'!D10:D17",
        "entrypoint_mode": "per_cell",
    },
    {
        "label": "Fiscal (Total Public Debt) risk rating signals",
        "range_spec": "'Chart Data'!I10:I14",
        "entrypoint_mode": "per_cell",
    },
    {
        "label": "Applicable tailored stress test signals",
        "range_spec": "'Chart Data'!I17:I19",
        "entrypoint_mode": "row_group",
    },
    {
        "label": "Fiscal space for moderate risk category",
        "range_spec": "'Chart Data'!E25:E27",
        "entrypoint_mode": "row_group",
    },
    {
        "label": "Overall rating",
        "range_spec": "'Chart Data'!L10:L11",
        "entrypoint_mode": "row_group",
    },
]


def _export_chart_data_ranges() -> list[ExportRangeConfig]:
    out: list[ExportRangeConfig] = list(EXPORT_FIXED_RANGES)
    for metric_label, start_row in STRESS_TEST_BLOCKS:
        for i, row_label in enumerate(STRESS_TEST_ROW_LABELS):
            if not row_label:
                continue
            row = start_row + i
            out.append(
                {
                    "label": f"{metric_label} - {row_label}",
                    "range_spec": f"'Chart Data'!D{row}:X{row}",
                    "entrypoint_mode": "row_group",
                }
            )
    return out


EXPORT_RANGES: list[ExportRangeConfig] = _export_chart_data_ranges()


def parse_range_spec(spec: str) -> tuple[str, str]:
    """
    Parse a sheet-qualified range spec into (sheet_name, range_a1).

    Accepts specs like "'Chart Data'!D10:D17" or "Sheet1!A1:B2".
    """
    if "!" not in spec:
        raise ValueError(f"Range spec must contain '!': {spec!r}")
    sheet_part, range_part = spec.split("!", 1)
    sheet_part = sheet_part.strip()
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1].replace("''", "'")
    return sheet_part, range_part.strip()


def cells_in_range(sheet: str, range_a1: str) -> list[str]:
    """
    Expand an A1 range to a list of sheet-qualified cell keys.

    The returned keys use `excel_grapher.format_cell_key` so they match the
    dependency-graph expectations (including sheet-name quoting when needed).
    """
    if ":" in range_a1:
        start_a1, end_a1 = range_a1.split(":", 1)
        start_a1 = start_a1.strip()
        end_a1 = end_a1.strip()
    else:
        start_a1 = end_a1 = range_a1.strip()

    c1, r1 = openpyxl.utils.cell.coordinate_from_string(start_a1)
    c2, r2 = openpyxl.utils.cell.coordinate_from_string(end_a1)
    start_col_idx = openpyxl.utils.cell.column_index_from_string(c1)
    end_col_idx = openpyxl.utils.cell.column_index_from_string(c2)
    rlo, rhi = (r1, r2) if r1 <= r2 else (r2, r1)
    clo, chi = (start_col_idx, end_col_idx) if start_col_idx <= end_col_idx else (
        end_col_idx,
        start_col_idx,
    )

    out: list[str] = []
    for row in range(rlo, rhi + 1):
        for col_idx in range(clo, chi + 1):
            col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
            out.append(format_cell_key(sheet, col_letter, row))
    return out


def discover_targets_from_ranges(_workbook: Path) -> list[str]:
    """
    Discover export/annotation targets from explicit range specs.

    The workbook path is accepted for API compatibility but is not needed
    for discovery; ranges are expanded purely from configuration.
    """
    targets: list[str] = []
    for cfg in EXPORT_RANGES:
        sheet_name, range_a1 = parse_range_spec(cfg["range_spec"])
        targets.extend(cells_in_range(sheet_name, range_a1))
    # Preserve order while de-duplicating.
    return list(dict.fromkeys(targets))


# Constraint-based config for dynamic references (OFFSET/INDIRECT).

class LicDsfConstraints(TypedDict, total=False):
    """
    Constraint types for cells that feed OFFSET/INDIRECT.

    Keys are address-style (e.g. "Sheet1!B1"). Add entries when the graph
    builder raises DynamicRefError and reports missing constraints.
    """

    pass

# PV_Base!B9xx = CONCAT("$", A9xx, "$", $A$<row>) → INDIRECT($B9xx). Row-index cells A917, A941, A965.
# These are fixed constants derived from their current workbook values.
LicDsfConstraints.__annotations__["PV_Base!A917"] = Annotated[int, FromWorkbook()]
LicDsfConstraints.__annotations__["PV_Base!A941"] = Annotated[int, FromWorkbook()]
LicDsfConstraints.__annotations__["PV_Base!A965"] = Annotated[int, FromWorkbook()]

# A918:A938, A942:A962, A966:A986 each has a single cached letter D, E, …, X.
# Treat these as constants derived from their current workbook values.
for _start, _end in [(918, 939), (942, 963), (966, 987)]:
    for _row in range(_start, _end):
        LicDsfConstraints.__annotations__[f"PV_Base!A{_row}"] = Annotated[str, FromWorkbook()]

# Country-name lookup table cells are treated as constants resolved from the workbook.
for _row in range(4, 74):  # lookup!C4:C73
    LicDsfConstraints.__annotations__[f"lookup!C{_row}"] = Annotated[str, FromWorkbook()]

# Language selector and lookup table (feed INDIRECT/VLOOKUP for language-dependent refs).
# START!L10 = VLOOKUP(K10, lookup!BB4:BC7, 2); evaluator does not support VLOOKUP, so L10 is constrained.
_LANG = Literal["English", "French", "Portuguese", "Spanish"]
_LANG_LOOKUP = Literal[
    "English", "French", "Portuguese", "Spanish", "Français", "Portugues", "Español"
]
LicDsfConstraints.__annotations__["START!L10"] = _LANG
LicDsfConstraints.__annotations__["START!K10"] = _LANG
for _r in range(4, 8):
    for _c in ("BB", "BC"):
        LicDsfConstraints.__annotations__[f"lookup!{_c}{_r}"] = _LANG_LOOKUP


def get_dynamic_ref_config() -> DynamicRefConfig:
    """Return a DynamicRefConfig for constraint-based resolution of OFFSET/INDIRECT."""
    return DynamicRefConfig.from_constraints_and_workbook(
        LicDsfConstraints, WORKBOOK_PATH
    )


