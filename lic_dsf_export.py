#!/usr/bin/env python3
"""
Export LIC-DSF workbook formulas as standalone Python code.

This script discovers formula targets from `map_lic_dsf_indicators.INDICATOR_CONFIG`,
builds a dependency graph, and uses `excel-formula-expander`'s `CodeGenerator` to
emit a small Python package under `export/`.
"""

from __future__ import annotations

from pathlib import Path
import ast
import json
import re

import openpyxl
import openpyxl.utils.cell
from openpyxl.worksheet.worksheet import Worksheet
from excel_grapher import create_dependency_graph
from excel_formula_expander.codegen import CodeGenerator

from map_lic_dsf_indicators import (
    INDICATOR_CONFIG,
    WORKBOOK_PATH,
    discover_formula_cells_in_rows,
    ensure_workbook_available,
    find_region_config,
    get_column_labels,
    get_row_labels,
    get_labels_from_region_config,
)


EXPORT_DIR = Path("export")
ENRICHMENT_AUDIT_PATH = Path("enrichment_audit.json")
INPUT_GROUPS_PATH = Path("input_groups_export.json")
MAX_DEPTH = 50


def discover_targets_by_indicator_row(workbook: Path) -> dict[tuple[str, int], list[str]]:
    targets_by_row: dict[tuple[str, int], list[str]] = {}
    for config in INDICATOR_CONFIG:
        sheet = config["sheet"]
        rows = config["indicator_rows"]
        for row in rows:
            targets = discover_formula_cells_in_rows(workbook, sheet, [row])
            if targets:
                targets_by_row[(sheet, row)] = list(dict.fromkeys(targets))
    return targets_by_row


def normalize_entrypoint_name(name: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z]+", "_", name.strip()).strip("_").lower()
    if not cleaned:
        return "sheet"
    if not cleaned[0].isalpha():
        return f"sheet_{cleaned}"
    return cleaned


def load_enrichment_audit_labels(path: Path) -> dict[tuple[str, int], list[str]]:
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}

    labels_by_row: dict[tuple[str, int], list[str]] = {}
    by_sheet = payload.get("by_sheet", {})
    for sheet, sheet_payload in by_sheet.items():
        for cell in sheet_payload.get("cells", []):
            address = cell.get("address")
            if not isinstance(address, str):
                continue
            match = re.match(r"^[A-Z]+(\d+)$", address)
            if not match:
                continue
            row = int(match.group(1))
            row_labels = cell.get("row_labels") or []
            if not isinstance(row_labels, list):
                continue
            if row_labels:
                labels_by_row.setdefault((sheet, row), []).extend(
                    label for label in row_labels if isinstance(label, str) and label.strip()
                )
    return labels_by_row


def build_entrypoints(
    targets_by_row: dict[tuple[str, int], list[str]]
) -> dict[str, list[str]]:
    labels_by_row = load_enrichment_audit_labels(ENRICHMENT_AUDIT_PATH)
    entrypoints: dict[str, list[str]] = {}
    for (sheet, row), targets in targets_by_row.items():
        prefix_match = re.match(r"^([A-Za-z]\d+)", sheet.strip())
        sheet_prefix = normalize_entrypoint_name(prefix_match.group(1) if prefix_match else sheet)
        label = next(iter(labels_by_row.get((sheet, row), [])), "")
        base = normalize_entrypoint_name(label or f"{sheet} {row}")
        if base.startswith(f"{sheet_prefix}_") or base == sheet_prefix:
            name = base
        else:
            name = f"{sheet_prefix}_{base}"
        suffix = 2
        while name in entrypoints:
            name = f"{base}_{suffix}"
            suffix += 1
        entrypoints[name] = targets
    return entrypoints


def populate_leaf_values(graph, workbook: Path) -> None:
    """
    Populate values for leaf (non-formula) nodes from cached workbook values.

    Code generation only needs `node.value` for leaf cells; formulas are emitted
    from `node.formula`.
    """
    wb = openpyxl.load_workbook(workbook, data_only=True, keep_vba=True)
    try:
        worksheets: dict[str, Worksheet] = {}
        for key in graph:
            node = graph.get_node(key)
            if node is None or node.formula is not None:
                continue
            if node.sheet not in worksheets:
                if node.sheet not in wb.sheetnames:
                    continue
                worksheets[node.sheet] = wb[node.sheet]
            ws = worksheets[node.sheet]
            col_idx = openpyxl.utils.cell.column_index_from_string(node.column)
            node.value = ws.cell(row=node.row, column=col_idx).value
    finally:
        wb.close()

_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z_][0-9A-Za-z_]*$")


def format_sheet_name(sheet: str) -> str:
    """
    Format a sheet name for sheet-qualified Excel addresses.

    Only quote sheet names when needed (e.g., spaces or punctuation), and escape
    embedded single quotes by doubling them.
    """
    if _SAFE_SHEET_NAME_RE.match(sheet):
        return sheet
    escaped = sheet.replace("'", "''")
    return f"'{escaped}'"


def format_address(sheet: str, a1: str) -> str:
    return f"{format_sheet_name(sheet)}!{a1}"


def parse_year_label(label: str) -> int | None:
    try:
        y = int(str(label).strip())
    except ValueError:
        return None
    return y if 1900 <= y <= 2100 else None


def year_for_column(ws: Worksheet, row: int, col: int) -> int | None:
    """
    Determine the year label for a cell's column header context.

    Uses the same region-config-aware label extraction as enrichment.
    Returns a single year if exactly one year label is found; otherwise None.
    """
    cfg = find_region_config(ws.title, row, col)
    if cfg is not None:
        _row_labels, col_labels = get_labels_from_region_config(ws, row, col, cfg)
    else:
        col_labels = get_column_labels(ws, row, col)

    years: list[int] = []
    for lab in col_labels:
        y = parse_year_label(lab)
        if y is not None:
            years.append(y)

    years = list(dict.fromkeys(years))
    if len(years) != 1:
        return None
    return years[0]


def year_for_row(ws: Worksheet, row: int, col: int) -> int | None:
    """
    Determine the year label for a cell's row label context.

    Uses the same region-config-aware label extraction as enrichment.
    Returns a single year if exactly one year label is found; otherwise None.
    """
    cfg = find_region_config(ws.title, row, col)
    if cfg is not None:
        row_labels, _col_labels = get_labels_from_region_config(ws, row, col, cfg)
    else:
        row_labels = get_row_labels(ws, row, col)

    years: list[int] = []
    for lab in row_labels:
        y = parse_year_label(lab)
        if y is not None:
            years.append(y)

    years = list(dict.fromkeys(years))
    if len(years) != 1:
        return None
    return years[0]


def load_input_groups(path: Path = INPUT_GROUPS_PATH) -> list[dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    groups = payload.get("groups", [])
    if not isinstance(groups, list):
        return []
    return [g for g in groups if isinstance(g, dict)]


def iter_wide_year_series_groups(groups: list[dict]) -> list[dict]:
    out: list[dict] = []
    for g in groups:
        if g.get("mode") != "ignore_column_axis_years":
            continue
        shape = g.get("shape") or {}
        if not isinstance(shape, dict):
            continue
        if shape.get("rows") != 1:
            continue
        if not g.get("bounding_box"):
            continue
        if not g.get("range_a1"):
            continue
        out.append(g)
    return out


def generate_setter_method_name(sheet: str, row_labels_key: list[str], group_id: str) -> str:
    prefix = normalize_entrypoint_name(sheet)
    label = next((s for s in row_labels_key if isinstance(s, str) and s.strip()), "")
    base = normalize_entrypoint_name(label) if label else group_id.lower()
    return f"set_{prefix}_{base}"


def split_columns_by_year_presence(
    *,
    start_col: int,
    end_col: int,
    year_by_col: dict[int, int | None],
) -> list[tuple[str, int, int]]:
    """
    Split a contiguous column span into segments of year vs non-year columns.

    Returns segments as tuples: (kind, seg_start_col, seg_end_col), where kind is
    either "year" or "meta".
    """
    segments: list[tuple[str, int, int]] = []
    current_kind: str | None = None
    seg_start: int | None = None

    for col in range(start_col, end_col + 1):
        kind = "year" if year_by_col.get(col) is not None else "meta"
        if current_kind is None:
            current_kind = kind
            seg_start = col
            continue
        if kind == current_kind:
            continue
        # close previous segment
        segments.append((current_kind, int(seg_start), col - 1))
        current_kind = kind
        seg_start = col

    if current_kind is not None and seg_start is not None:
        segments.append((current_kind, int(seg_start), end_col))

    return segments


def split_rows_by_year_presence(
    *,
    start_row: int,
    end_row: int,
    year_by_row: dict[int, int | None],
) -> list[tuple[str, int, int]]:
    """
    Split a contiguous row span into segments of year vs non-year rows.

    Returns segments as tuples: (kind, seg_start_row, seg_end_row), where kind is
    either "year" or "meta".
    """
    segments: list[tuple[str, int, int]] = []
    current_kind: str | None = None
    seg_start: int | None = None

    for row in range(start_row, end_row + 1):
        kind = "year" if year_by_row.get(row) is not None else "meta"
        if current_kind is None:
            current_kind = kind
            seg_start = row
            continue
        if kind == current_kind:
            continue
        segments.append((current_kind, int(seg_start), row - 1))
        current_kind = kind
        seg_start = row

    if current_kind is not None and seg_start is not None:
        segments.append((current_kind, int(seg_start), end_row))

    return segments


def generate_setters_module(
    *,
    workbook: Path,
    groups: list[dict],
) -> str:
    """
    Generate a `setters.py` module for the exported package.

    Generates:
    - Wide-format year-series setters for eligible groups (mode=ignore_column_axis_years, 1-row range)
    - Tall-format year-series setters for eligible groups (mode=ignore_row_axis_years, sparse cell sets)
    - Scalar / 1D / 2D setters for non-year rectangular groups (mode=no_year_axis)

    For year-series groups that contain a mix of year and non-year columns,
    split the group at codegen time into year-series setters and metadata setters.
    """
    if not groups:
        return "from __future__ import annotations\n\n# No setters were generated.\n"

    wb = openpyxl.load_workbook(workbook, data_only=True, keep_vba=True)
    try:
        year_specs: list[dict] = []
        tall_year_specs: list[dict] = []
        range_specs: list[dict] = []
        used_names: set[str] = set()

        for g in groups:
            sheet = g.get("sheet")
            group_id = str(g.get("group_id", "group"))
            row_labels_key = g.get("row_labels_key") or []
            mode = g.get("mode")

            if not isinstance(sheet, str) or sheet not in wb.sheetnames:
                continue

            base_name = generate_setter_method_name(sheet, row_labels_key, group_id)

            def unique(name: str) -> str:
                n = name
                suffix = 2
                while n in used_names:
                    n = f"{name}_{suffix}"
                    suffix += 1
                used_names.add(n)
                return n

            ws = wb[sheet]

            # Tall-format year series (year axis is the row axis).
            # These groups are often sparse (not rectangular) because only leaf inputs appear in `cells`.
            if mode == "ignore_row_axis_years":
                cells = g.get("cells") or []
                if not isinstance(cells, list) or not cells:
                    continue

                year_to_addresses: dict[int, list[str]] = {}
                meta_addresses: list[str] = []

                for addr in cells:
                    if not isinstance(addr, str) or "!" not in addr:
                        continue
                    _sheet, a1 = addr.split("!", 1)
                    m = re.match(r"^([A-Z]+)(\d+)$", a1)
                    if not m:
                        continue
                    col_letter, row_str = m.group(1), m.group(2)
                    r = int(row_str)
                    c = openpyxl.utils.cell.column_index_from_string(col_letter)

                    y = year_for_row(ws, r, c)
                    if y is None:
                        meta_addresses.append(addr)
                        continue
                    year_to_addresses.setdefault(int(y), []).append(addr)

                if year_to_addresses:
                    years_sorted = sorted(year_to_addresses.keys())
                    year_to_addresses_sorted = {
                        y: tuple(sorted(year_to_addresses[y])) for y in years_sorted
                    }
                    tall_year_specs.append(
                        {
                            "name": unique(f"{base_name}_by_year"),
                            "years": years_sorted,
                            "year_to_addresses": year_to_addresses_sorted,
                        }
                    )

                if meta_addresses:
                    range_specs.append(
                        {
                            "name": unique(f"{base_name}_meta"),
                            "shape": (1, len(meta_addresses)),
                            "addresses": sorted(meta_addresses),
                        }
                    )

                continue

            # Everything below requires rectangular metadata from the grouping step.
            bbox = g.get("bounding_box") or {}
            shape = g.get("shape") or {}
            if not isinstance(bbox, dict) or not isinstance(shape, dict):
                continue
            if not g.get("range_a1"):
                continue

            row = bbox.get("start_row")
            start_col = bbox.get("start_col")
            end_col = bbox.get("end_col")
            end_row = bbox.get("end_row")
            if not all(isinstance(x, int) for x in (row, start_col, end_col, end_row)):
                continue

            rows = shape.get("rows")
            cols = shape.get("cols")
            if not (isinstance(rows, int) and isinstance(cols, int)):
                continue

            # Wide-format year series
            if mode == "ignore_column_axis_years" and rows == 1 and cols >= 1 and row == end_row:
                year_by_col: dict[int, int | None] = {}
                for col in range(start_col, end_col + 1):
                    year_by_col[col] = year_for_column(ws, row, col)

                segments = split_columns_by_year_presence(
                    start_col=start_col, end_col=end_col, year_by_col=year_by_col
                )

                # Create a year-series setter per contiguous year segment.
                year_seg_index = 0
                meta_seg_index = 0
                for kind, seg_start, seg_end in segments:
                    if kind == "year":
                        year_seg_index += 1
                        year_to_address: dict[int, str] = {}
                        years: list[int] = []
                        ok = True
                        for col in range(seg_start, seg_end + 1):
                            y = year_by_col.get(col)
                            if y is None:
                                ok = False
                                break
                            if y in year_to_address:
                                ok = False
                                break
                            col_letter = openpyxl.utils.cell.get_column_letter(col)
                            addr = format_address(sheet, f"{col_letter}{row}")
                            year_to_address[int(y)] = addr
                            years.append(int(y))
                        if not ok or not year_to_address:
                            continue

                        name = base_name if len(segments) == 1 else f"{base_name}_years_{year_seg_index}"
                        year_specs.append(
                            {
                                "name": unique(name),
                                "years": years,
                                "year_to_address": year_to_address,
                            }
                        )
                    else:
                        # Metadata segment: generate a simple 1D range setter.
                        meta_seg_index += 1
                        addresses: list[str] = []
                        for col in range(seg_start, seg_end + 1):
                            col_letter = openpyxl.utils.cell.get_column_letter(col)
                            addresses.append(format_address(sheet, f"{col_letter}{row}"))
                        if not addresses:
                            continue
                        meta_name = f"{base_name}_meta_{meta_seg_index}"
                        range_specs.append(
                            {
                                "name": unique(meta_name),
                                "shape": (1, len(addresses)),
                                "addresses": addresses,
                            }
                        )

                continue

            # Non-year rectangular setters (scalar/1D/2D)
            if mode == "no_year_axis":
                addresses: list[str] = []
                for r in range(row, row + rows):
                    for c in range(start_col, start_col + cols):
                        col_letter = openpyxl.utils.cell.get_column_letter(c)
                        addresses.append(format_address(sheet, f"{col_letter}{r}"))

                if not addresses:
                    continue

                range_specs.append(
                    {
                        "name": unique(base_name),
                        "shape": (rows, cols),
                        "addresses": addresses,
                    }
                )

        lines: list[str] = []
        lines.append("from __future__ import annotations")
        lines.append("")
        lines.append("from dataclasses import dataclass")
        lines.append("from typing import Mapping, Sequence")
        lines.append("")
        lines.append("from .internals import CellValue, EvalContext")
        lines.append("")
        lines.append("")
        lines.append("@dataclass(frozen=True, slots=True)")
        lines.append("class YearSeriesAssignment:")
        lines.append("    years: tuple[int, ...]")
        lines.append("    applied: dict[int, str]")
        lines.append("    ignored: dict[int, CellValue]")
        lines.append("")
        lines.append("")
        lines.append("@dataclass(frozen=True, slots=True)")
        lines.append("class RangeAssignment:")
        lines.append("    shape: tuple[int, int]")
        lines.append("    addresses: tuple[str, ...]")
        lines.append("")
        lines.append("")
        lines.append("@dataclass(frozen=True, slots=True)")
        lines.append("class YearRowAssignment:")
        lines.append("    years: tuple[int, ...]")
        lines.append("    applied: dict[int, tuple[str, ...]]")
        lines.append("    ignored: dict[int, CellValue]")
        lines.append("")
        lines.append("")
        lines.append("def _apply_range(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    shape: tuple[int, int],")
        lines.append("    addresses: Sequence[str],")
        lines.append("    values: object,")
        lines.append(") -> RangeAssignment:")
        lines.append("    rows, cols = shape")
        lines.append("    updates: dict[str, CellValue] = {}")
        lines.append("    flat: list[CellValue] = []")
        lines.append("    if rows == 1 and cols == 1:")
        lines.append("        flat = [values]  # scalar")
        lines.append("    elif rows == 1 or cols == 1:")
        lines.append("        if not isinstance(values, Sequence):")
        lines.append("            raise TypeError('Expected a sequence for 1D range')") 
        lines.append("        flat = list(values)")
        lines.append("        if len(flat) != len(addresses):")
        lines.append("            raise ValueError(f'Expected {len(addresses)} values, got {len(flat)}')")
        lines.append("    else:")
        lines.append("        if not isinstance(values, Sequence):")
        lines.append("            raise TypeError('Expected a sequence of sequences for 2D range')")
        lines.append("        rows_values = list(values)")
        lines.append("        if len(rows_values) != rows:")
        lines.append("            raise ValueError(f'Expected {rows} rows, got {len(rows_values)}')")
        lines.append("        for rv in rows_values:")
        lines.append("            if not isinstance(rv, Sequence):")
        lines.append("                raise TypeError('Expected a sequence of sequences for 2D range')")
        lines.append("            row_list = list(rv)")
        lines.append("            if len(row_list) != cols:")
        lines.append("                raise ValueError(f'Expected {cols} columns, got {len(row_list)}')")
        lines.append("            flat.extend(row_list)")
        lines.append("    if len(flat) != len(addresses):")
        lines.append("        raise ValueError(f'Expected {len(addresses)} values, got {len(flat)}')")
        lines.append("    for addr, value in zip(addresses, flat):")
        lines.append("        v = 0 if value is None else value")
        lines.append("        updates[str(addr)] = v")
        lines.append("    if updates:")
        lines.append("        ctx.set_inputs(updates)")
        lines.append("    return RangeAssignment(shape=shape, addresses=tuple(addresses))")
        lines.append("")
        lines.append("")
        lines.append("def _apply_year_row_mapping(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    years: tuple[int, ...],")
        lines.append("    year_to_addresses: dict[int, tuple[str, ...]],")
        lines.append("    values_by_year: Mapping[int, CellValue],")
        lines.append("    strict: bool = True,")
        lines.append(") -> YearRowAssignment:")
        lines.append("    applied: dict[int, tuple[str, ...]] = {}")
        lines.append("    ignored: dict[int, CellValue] = {}")
        lines.append("    updates: dict[str, CellValue] = {}")
        lines.append("    for year, value in values_by_year.items():")
        lines.append("        addrs = year_to_addresses.get(int(year))")
        lines.append("        if addrs is None:")
        lines.append("            if strict:")
        lines.append("                raise KeyError(f\"Year {year} is not in this table: {years}\")")
        lines.append("            ignored[int(year)] = value")
        lines.append("            continue")
        lines.append("        v = 0 if value is None else value")
        lines.append("        for addr in addrs:")
        lines.append("            updates[str(addr)] = v")
        lines.append("        applied[int(year)] = tuple(addrs)")
        lines.append("    if updates:")
        lines.append("        ctx.set_inputs(updates)")
        lines.append("    return YearRowAssignment(years=years, applied=applied, ignored=ignored)")
        lines.append("")
        lines.append("")
        lines.append("def _apply_year_row_array(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    years: tuple[int, ...],")
        lines.append("    year_to_addresses: dict[int, tuple[str, ...]],")
        lines.append("    values: Sequence[CellValue],")
        lines.append("    start_year: int,")
        lines.append("    strict: bool = True,")
        lines.append(") -> YearRowAssignment:")
        lines.append("    if start_year not in year_to_addresses:")
        lines.append("        raise KeyError(f\"start_year {start_year} is not in this table: {years}\")")
        lines.append("    years_list = list(years)")
        lines.append("    start_idx = years_list.index(start_year)")
        lines.append("    remaining_years = years_list[start_idx:]")
        lines.append("    if len(values) > len(remaining_years):")
        lines.append("        raise ValueError(")
        lines.append("            f\"Too many values ({len(values)}) for table from {start_year}; \"")
        lines.append("            f\"only {len(remaining_years)} years available\"")
        lines.append("        )")
        lines.append("    expected = list(range(start_year, start_year + len(remaining_years)))")
        lines.append("    if remaining_years != expected:")
        lines.append("        raise ValueError(")
        lines.append("            \"Non-contiguous years; array mapping is disallowed for this table. \"")
        lines.append("            \"Use dict-based mapping instead.\"")
        lines.append("        )")
        lines.append("    values_by_year = {start_year + i: values[i] for i in range(len(values))}")
        lines.append("    return _apply_year_row_mapping(")
        lines.append("        ctx, years=years, year_to_addresses=year_to_addresses, values_by_year=values_by_year, strict=strict")
        lines.append("    )")
        lines.append("")
        lines.append("")
        lines.append("def _apply_year_series_mapping(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    years: tuple[int, ...],")
        lines.append("    year_to_address: dict[int, str],")
        lines.append("    values_by_year: Mapping[int, CellValue],")
        lines.append("    strict: bool = True,")
        lines.append(") -> YearSeriesAssignment:")
        lines.append("    applied: dict[int, str] = {}")
        lines.append("    ignored: dict[int, CellValue] = {}")
        lines.append("    updates: dict[str, CellValue] = {}")
        lines.append("    for year, value in values_by_year.items():")
        lines.append("        addr = year_to_address.get(int(year))")
        lines.append("        if addr is None:")
        lines.append("            if strict:")
        lines.append("                raise KeyError(f\"Year {year} is not in this series: {years}\")")
        lines.append("            ignored[int(year)] = value")
        lines.append("            continue")
        lines.append("        v = 0 if value is None else value")
        lines.append("        updates[addr] = v")
        lines.append("        applied[int(year)] = addr")
        lines.append("    if updates:")
        lines.append("        ctx.set_inputs(updates)")
        lines.append("    return YearSeriesAssignment(years=years, applied=applied, ignored=ignored)")
        lines.append("")
        lines.append("")
        lines.append("def _apply_year_series_array(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    years: tuple[int, ...],")
        lines.append("    year_to_address: dict[int, str],")
        lines.append("    values: Sequence[CellValue],")
        lines.append("    start_year: int,")
        lines.append("    strict: bool = True,")
        lines.append(") -> YearSeriesAssignment:")
        lines.append("    if start_year not in year_to_address:")
        lines.append("        raise KeyError(f\"start_year {start_year} is not in this series: {years}\")")
        lines.append("    years_list = list(years)")
        lines.append("    start_idx = years_list.index(start_year)")
        lines.append("    remaining_years = years_list[start_idx:]")
        lines.append("    if len(values) > len(remaining_years):")
        lines.append("        raise ValueError(")
        lines.append("            f\"Too many values ({len(values)}) for series from {start_year}; \"")
        lines.append("            f\"only {len(remaining_years)} years available\"")
        lines.append("        )")
        lines.append("    expected = list(range(start_year, start_year + len(remaining_years)))")
        lines.append("    if remaining_years != expected:")
        lines.append("        raise ValueError(")
        lines.append("            \"Non-contiguous years; array mapping is disallowed for this series. \"")
        lines.append("            \"Use dict-based mapping instead.\"")
        lines.append("        )")
        lines.append("    values_by_year = {start_year + i: values[i] for i in range(len(values))}")
        lines.append("    return _apply_year_series_mapping(")
        lines.append("        ctx, years=years, year_to_address=year_to_address, values_by_year=values_by_year, strict=strict")
        lines.append("    )")
        lines.append("")
        lines.append("")
        lines.append("class LicDsfContext(EvalContext):")
        lines.append("    __slots__ = ()")
        lines.append("")

        for spec in year_specs:
            name = spec["name"]
            years = spec["years"]
            year_to_address = spec["year_to_address"]

            lines.append(f"    def {name}(")
            lines.append("        self,")
            lines.append("        values_by_year: Mapping[int, CellValue],")
            lines.append("        *,")
            lines.append("        strict: bool = True,")
            lines.append("    ) -> YearSeriesAssignment:")
            lines.append("        return _apply_year_series_mapping(")
            lines.append(f"            self, years={tuple(years)!r}, year_to_address={year_to_address!r},")
            lines.append("            values_by_year=values_by_year, strict=strict,")
            lines.append("        )")
            lines.append("")
            lines.append(f"    def {name}_from_array(")
            lines.append("        self,")
            lines.append("        values: Sequence[CellValue],")
            lines.append("        *,")
            lines.append("        start_year: int,")
            lines.append("        strict: bool = True,")
            lines.append("    ) -> YearSeriesAssignment:")
            lines.append("        return _apply_year_series_array(")
            lines.append(f"            self, years={tuple(years)!r}, year_to_address={year_to_address!r},")
            lines.append("            values=values, start_year=start_year, strict=strict,")
            lines.append("        )")
            lines.append("")

        for spec in range_specs:
            name = spec["name"]
            shape = spec["shape"]
            addresses = spec["addresses"]
            lines.append(f"    def {name}(")
            lines.append("        self,")
            lines.append("        values: object,")
            lines.append("    ) -> RangeAssignment:")
            lines.append("        return _apply_range(")
            lines.append(f"            self, shape={tuple(shape)!r}, addresses={list(addresses)!r}, values=values")
            lines.append("        )")
            lines.append("")

        for spec in tall_year_specs:
            name = spec["name"]
            years = spec["years"]
            year_to_addresses = spec["year_to_addresses"]
            lines.append(f"    def {name}(")
            lines.append("        self,")
            lines.append("        values_by_year: Mapping[int, CellValue],")
            lines.append("        *,")
            lines.append("        strict: bool = True,")
            lines.append("    ) -> YearRowAssignment:")
            lines.append("        return _apply_year_row_mapping(")
            lines.append(f"            self, years={tuple(years)!r}, year_to_addresses={year_to_addresses!r},")
            lines.append("            values_by_year=values_by_year, strict=strict,")
            lines.append("        )")
            lines.append("")
            lines.append(f"    def {name}_from_array(")
            lines.append("        self,")
            lines.append("        values: Sequence[CellValue],")
            lines.append("        *,")
            lines.append("        start_year: int,")
            lines.append("        strict: bool = True,")
            lines.append("    ) -> YearRowAssignment:")
            lines.append("        return _apply_year_row_array(")
            lines.append(f"            self, years={tuple(years)!r}, year_to_addresses={year_to_addresses!r},")
            lines.append("            values=values, start_year=start_year, strict=strict,")
            lines.append("        )")
            lines.append("")

        return "\n".join(lines).rstrip() + "\n"
    finally:
        wb.close()


def patch_entrypoint_for_setters(entrypoint_py: str) -> str:
    if "from .setters import LicDsfContext" not in entrypoint_py:
        entrypoint_py = entrypoint_py.replace(
            "from .internals import EvalContext, xl_cell, xl_range, _resolve_formula\n",
            "from .internals import EvalContext, xl_cell, xl_range, _resolve_formula\n"
            "from .setters import LicDsfContext\n",
        )
    entrypoint_py = entrypoint_py.replace(
        "return EvalContext(inputs=merged, resolver=_resolve_formula)",
        "return LicDsfContext(inputs=merged, resolver=_resolve_formula)",
    )
    return entrypoint_py


def patch_init_for_setters(init_py: str) -> str:
    # Ensure the public symbols are imported.
    if "from .setters import LicDsfContext" not in init_py:
        init_py = init_py.replace(
            "from .inputs import DEFAULT_INPUTS  # noqa: F401\n",
            "from .inputs import DEFAULT_INPUTS  # noqa: F401\n"
            "from .setters import LicDsfContext, YearSeriesAssignment, RangeAssignment, YearRowAssignment  # noqa: F401\n",
        )
    else:
        # Expand existing import line to include missing symbols.
        if "YearRowAssignment" not in init_py:
            init_py = init_py.replace(
                "from .setters import LicDsfContext, YearSeriesAssignment, RangeAssignment  # noqa: F401\n",
                "from .setters import LicDsfContext, YearSeriesAssignment, RangeAssignment, YearRowAssignment  # noqa: F401\n",
            )
        if "RangeAssignment" not in init_py:
            init_py = init_py.replace(
                "from .setters import LicDsfContext, YearSeriesAssignment  # noqa: F401\n",
                "from .setters import LicDsfContext, YearSeriesAssignment, RangeAssignment, YearRowAssignment  # noqa: F401\n",
            )

    # Ensure __all__ includes these names even if they are already imported.
    names_to_add = ["LicDsfContext", "YearSeriesAssignment", "RangeAssignment", "YearRowAssignment"]
    for line in init_py.splitlines():
        if line.strip().startswith("__all__"):
            try:
                _, rhs = line.split("=", 1)
                current = ast.literal_eval(rhs.strip())
            except Exception:
                current = None
            if isinstance(current, list):
                updated = list(current)
                for n in names_to_add:
                    if n not in updated:
                        updated.append(n)
                new_line = f"__all__ = {updated!r}"
                init_py = init_py.replace(line, new_line)
            break

    return init_py


def main() -> None:
    if not ensure_workbook_available(WORKBOOK_PATH):
        print(f"Error: Workbook not available at {WORKBOOK_PATH}")
        return

    targets_by_row = discover_targets_by_indicator_row(WORKBOOK_PATH)
    targets: list[str] = []
    for row_targets in targets_by_row.values():
        targets.extend(row_targets)
    targets = list(dict.fromkeys(targets))
    if not targets:
        print("No targets found. Nothing to export.")
        return

    print("=" * 70)
    print("LIC-DSF Workbook Export (standalone Python code)")
    print("=" * 70)
    print(f"Workbook: {WORKBOOK_PATH}")
    entrypoints = build_entrypoints(targets_by_row)
    print(f"Targets:  {len(targets)}")
    print(f"Entrypoints: {len(entrypoints)}")

    print("\nBuilding dependency graph...")
    graph = create_dependency_graph(
        WORKBOOK_PATH,
        targets,
        load_values=False,
        max_depth=MAX_DEPTH,
    )

    print("Populating leaf values from cached workbook values...")
    populate_leaf_values(graph, WORKBOOK_PATH)

    print("Generating Python package...")
    generator = CodeGenerator(graph)
    modules = generator.generate_modules(
        targets,
        package_name="lic_dsf",
        entrypoints=entrypoints if entrypoints else None,
    )

    # Generate year-aware input setters from canonical input groups.
    if INPUT_GROUPS_PATH.exists():
        groups = load_input_groups(INPUT_GROUPS_PATH)
        setters_py = generate_setters_module(workbook=WORKBOOK_PATH, groups=groups)
        modules["lic_dsf/setters.py"] = setters_py
        if "lic_dsf/entrypoint.py" in modules:
            modules["lic_dsf/entrypoint.py"] = patch_entrypoint_for_setters(
                modules["lic_dsf/entrypoint.py"]
            )
        if "lic_dsf/__init__.py" in modules:
            modules["lic_dsf/__init__.py"] = patch_init_for_setters(
                modules["lic_dsf/__init__.py"]
            )
    else:
        print(f"Warning: {INPUT_GROUPS_PATH} not found; skipping input setter generation.")

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    for rel, content in modules.items():
        dst = EXPORT_DIR / rel
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_text(content, encoding="utf-8")

    pkg_prefix = Path(next(iter(modules.keys()))).parts[0] if modules else "(unknown)"
    print(f"\nWrote {len(modules)} files under {EXPORT_DIR / pkg_prefix}")


if __name__ == "__main__":
    main()

