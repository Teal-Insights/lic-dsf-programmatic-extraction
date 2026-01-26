from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Mapping, Protocol, Sequence, TypeAlias

import openpyxl
import openpyxl.utils.cell

from map_lic_dsf_indicators import (
    find_region_config,
    get_column_labels,
    get_labels_from_region_config,
)


CellValue: TypeAlias = Any


class InputsContext(Protocol):
    def set_inputs(self, inputs: dict[str, CellValue]) -> None: ...


@dataclass(frozen=True, slots=True)
class YearSeriesAssignment:
    """
    Transparent record of how a year-series setter mapped values to addresses.

    - `years` is the canonical year ordering (left-to-right for wide series).
    - `applied` maps each year to the address that was updated.
    - `ignored` records user-provided years that were dropped under strict=False.
    """

    years: tuple[int, ...]
    applied: dict[int, str]
    ignored: dict[int, CellValue]


def _parse_year(label: str) -> int | None:
    try:
        y = int(str(label).strip())
    except ValueError:
        return None
    return y if 1900 <= y <= 2100 else None


def _dedupe_preserve_order(values: Sequence[int]) -> list[int]:
    seen: set[int] = set()
    out: list[int] = []
    for v in values:
        if v in seen:
            continue
        seen.add(v)
        out.append(v)
    return out


def _year_labels_for_cell(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int) -> list[int]:
    cfg = find_region_config(ws.title, row, col)
    if cfg is not None:
        _row_labels, col_labels = get_labels_from_region_config(ws, row, col, cfg)
    else:
        col_labels = get_column_labels(ws, row, col)

    years: list[int] = []
    for lab in col_labels:
        y = _parse_year(lab)
        if y is not None:
            years.append(y)

    return _dedupe_preserve_order(years)


@dataclass(frozen=True, slots=True)
class WideYearSeriesSpec:
    """
    Specification for a 1-row, multi-column year series (wide format).

    `year_to_address` must have unique year keys.
    """

    years: tuple[int, ...]
    year_to_address: dict[int, str]


def build_wide_year_series_spec(
    *,
    workbook_path: str,
    sheet: str,
    row: int,
    start_col: int,
    end_col: int,
) -> WideYearSeriesSpec:
    """
    Build a year->address mapping for a 1-row wide range.

    This uses the same label enrichment logic (region config or heuristic scan)
    to determine which year is associated with each column.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True)
    try:
        ws = wb[sheet]

        year_to_address: dict[int, str] = {}
        ordered_years: list[int] = []

        for col in range(start_col, end_col + 1):
            years = _year_labels_for_cell(ws, row, col)
            if len(years) != 1:
                col_letter = openpyxl.utils.cell.get_column_letter(col)
                raise ValueError(
                    f"Ambiguous year label for {sheet}!{col_letter}{row}: {years}"
                )

            y = years[0]
            if y in year_to_address:
                col_letter = openpyxl.utils.cell.get_column_letter(col)
                raise ValueError(
                    f"Duplicate year {y} in series (also mapped earlier); at {sheet}!{col_letter}{row}"
                )

            addr = f"{sheet}!{openpyxl.utils.cell.get_column_letter(col)}{row}"
            year_to_address[y] = addr
            ordered_years.append(y)

        return WideYearSeriesSpec(
            years=tuple(ordered_years),
            year_to_address=year_to_address,
        )
    finally:
        wb.close()


def apply_year_series_mapping(
    ctx: InputsContext,
    spec: WideYearSeriesSpec,
    values_by_year: Mapping[int, CellValue],
    *,
    strict: bool = True,
) -> YearSeriesAssignment:
    """
    Apply a year->value mapping to the provided context.

    Semantics:
    - Missing years are simply not touched.
    - Unknown years raise if strict=True; otherwise they're ignored and reported.
    - None values are stored as 0 (Excel-like numeric semantics).
    """
    applied: dict[int, str] = {}
    ignored: dict[int, CellValue] = {}
    updates: dict[str, CellValue] = {}

    for year, value in values_by_year.items():
        addr = spec.year_to_address.get(int(year))
        if addr is None:
            if strict:
                raise KeyError(f"Year {year} is not in this series: {spec.years}")
            ignored[int(year)] = value
            continue

        v = 0 if value is None else value
        updates[addr] = v
        applied[int(year)] = addr

    if updates:
        ctx.set_inputs(updates)

    return YearSeriesAssignment(years=spec.years, applied=applied, ignored=ignored)


def apply_year_series_array(
    ctx: InputsContext,
    spec: WideYearSeriesSpec,
    values: Sequence[CellValue],
    *,
    start_year: int,
    strict: bool = True,
) -> YearSeriesAssignment:
    """
    Apply an array to the series, aligned by `start_year`.

    Mapping is contiguous: values[0] -> start_year, values[1] -> start_year+1, ...
    """
    if start_year not in spec.year_to_address:
        raise KeyError(f"start_year {start_year} is not in this series: {spec.years}")

    # Enforce contiguity for Option B to avoid silent misalignment.
    years = list(spec.years)
    start_idx = years.index(start_year)
    remaining_years = years[start_idx:]

    if len(values) > len(remaining_years):
        raise ValueError(
            f"Too many values ({len(values)}) for series from {start_year}; "
            f"only {len(remaining_years)} years available"
        )

    expected = list(range(start_year, start_year + len(remaining_years)))
    if remaining_years != expected:
        raise ValueError(
            "Non-contiguous years; array mapping is disallowed for this series. "
            "Use dict-based mapping instead."
        )

    values_by_year = {start_year + i: values[i] for i in range(len(values))}
    return apply_year_series_mapping(ctx, spec, values_by_year, strict=strict)

