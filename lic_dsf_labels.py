#!/usr/bin/env python3
"""
Label extraction and workbook configuration helpers for LIC-DSF workflows.
"""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
import shutil
import tempfile
from typing import Any, Literal, TypedDict
from urllib.request import urlopen

import openpyxl
import openpyxl.utils.cell
from excel_grapher import DependencyGraph
from openpyxl.worksheet.worksheet import Worksheet


# Configuration: sheets and indicator rows to trace
class IndicatorConfig(TypedDict):
    sheet: str
    indicator_rows: list[int]


INDICATOR_CONFIG: list[IndicatorConfig] = [
    {"sheet": "B1_GDP_ext", "indicator_rows": [35, 36, 39, 40]},
    {"sheet": "B3_Exports_ext", "indicator_rows": [35, 36, 39, 40]},
    {"sheet": "B4_other flows_ext", "indicator_rows": [35, 36, 39, 40]},
]

# Default workbook used for dependency mapping / enrichment.
WORKBOOK_PATH = Path("workbooks/lic-dsf-template-2026-01-31.xlsm")
WORKBOOK_TEMPLATE_URL = (
    "https://thedocs.worldbank.org/en/doc/f0ade6bcf85b6f98dbeb2c39a2b7770c-0360012025/new-lic-dsf-template"
)


def ensure_workbook_available(
    path: Path = WORKBOOK_PATH, url: str | None = None
) -> bool:
    """
    Ensure the default LIC-DSF template workbook exists locally.

    If the workbook is missing, downloads it from `url` (or `WORKBOOK_TEMPLATE_URL`) into `path`.
    """
    if path.exists() and path.stat().st_size > 0:
        return True

    path.parent.mkdir(parents=True, exist_ok=True)

    try:
        source_url = url or WORKBOOK_TEMPLATE_URL
        with urlopen(source_url, timeout=60) as resp:
            with tempfile.NamedTemporaryFile(
                prefix=f".{path.name}.", suffix=".download", dir=str(path.parent), delete=False
            ) as tmp:
                shutil.copyfileobj(resp, tmp)
                tmp_path = Path(tmp.name)

        if tmp_path.stat().st_size == 0:
            tmp_path.unlink(missing_ok=True)
            return False

        tmp_path.replace(path)
        return True
    except Exception:
        return False


# Region-based label configuration
class RegionConfig(TypedDict, total=False):
    """
    Configuration for label extraction in a specific region of a sheet.

    Attributes:
        sheet: The sheet name this config applies to
        min_row: Minimum row of the region (inclusive, 1-indexed). None = no min.
        max_row: Maximum row of the region (inclusive, 1-indexed). None = no max.
        min_col: Minimum column of the region (inclusive, e.g., "A"). None = no min.
        max_col: Maximum column of the region (inclusive, e.g., "Z"). None = no max.
        header_rows: List of row numbers that contain column headers (1-indexed)
        label_columns: List of column letters that contain row labels
        annotation_axis: Axis for deduplicating annotations - "row" for wide-format
            time series (one annotation per row), "column" for columnar time series,
            or "cell" for individual cell annotations. Default: auto-detect.
    """

    sheet: str
    min_row: int | None
    max_row: int | None
    min_col: str | None
    max_col: str | None
    header_rows: list[int]
    label_columns: list[str]
    annotation_axis: Literal["row", "column", "cell"]


# Region configurations for custom label extraction
# More specific regions should come before less specific ones for the same sheet
REGION_CONFIG: list[RegionConfig] = [
    # Input 5 - Local-debt Financing: Years in row 5 (cols H+), labels in cols A & B
    {
        "sheet": "Input 5 - Local-debt Financing",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [5],
        "label_columns": ["A", "B"],
    },
    # Ext_Debt_Data: Years in rows 1 and 9, labels in column A
    {
        "sheet": "Ext_Debt_Data",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [1, 9],
        "label_columns": ["A"],
    },
    # PV_Base: Years in row 7, labels in columns A and C
    {
        "sheet": "PV_Base",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [7],
        "label_columns": ["A", "C"],
    },
    # PV_LC_NR1: Years in row 3, labels in columns A and C
    {
        "sheet": "PV_LC_NR1",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [3],
        "label_columns": ["A", "C"],
    },
    # PV_LC_NR2: Same structure as NR1
    {
        "sheet": "PV_LC_NR2",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [3],
        "label_columns": ["A", "C"],
    },
    # PV_LC_NR3: Same structure as NR1
    {
        "sheet": "PV_LC_NR3",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [3],
        "label_columns": ["A", "C"],
    },
    # Input 3 - Macro-Debt data(DMX): Years in row 7, labels in columns A-C
    {
        "sheet": "Input 3 - Macro-Debt data(DMX)",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [7],
        "label_columns": ["A", "B", "C"],
    },
    # Input 4 - External Financing: Years in row 6, labels in column B
    {
        "sheet": "Input 4 - External Financing",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [6],
        "label_columns": ["B"],
    },
    # Baseline - external: Years in row 8, labels in column B
    {
        "sheet": "Baseline - external",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [8],
        "label_columns": ["B"],
    },
    # Baseline - public: Similar structure to external
    {
        "sheet": "Baseline - public",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [8],
        "label_columns": ["B"],
    },
    # Input 8 - SDR: Years in row 9, labels in column A
    {
        "sheet": "Input 8 - SDR",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [9],
        "label_columns": ["A"],
    },
    # B1_GDP_ext: Years in row 8, labels in columns B and Z (for far-right area)
    {
        "sheet": "B1_GDP_ext",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [8],
        "label_columns": ["B", "Z"],
    },
    # B3_Exports_ext: Same structure as B1
    {
        "sheet": "B3_Exports_ext",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [8],
        "label_columns": ["B", "Z"],
    },
    # B4_other flows_ext: Same structure as B1
    {
        "sheet": "B4_other flows_ext",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [8],
        "label_columns": ["B", "Z"],
    },
    # Macro-Debt_Data: Years in rows 1 and 5, labels in columns B and E
    {
        "sheet": "Macro-Debt_Data",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [1, 5],
        "label_columns": ["B", "E"],
    },
    # PV Stress: Years in row 3, labels in columns A and C
    {
        "sheet": "PV Stress",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [3],
        "label_columns": ["A", "C"],
    },
    # Input 6(optional)-Standard Test: Labels in columns A, B, C
    {
        "sheet": "Input 6(optional)-Standard Test",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["A", "B", "C"],
    },
    # Input 7 - Residual Financing: Labels in column B
    {
        "sheet": "Input 7 - Residual Financing",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["B"],
    },
    # START: Labels in columns A and B
    {
        "sheet": "START",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["A", "B"],
    },
    # lookup: Labels in columns A and B
    {
        "sheet": "lookup",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["A", "B"],
    },
    # translation: Labels in columns A and B
    {
        "sheet": "translation",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["A", "B"],
    },
]

# Excel error values to filter out
EXCEL_ERRORS = frozenset(
    {
        "#DIV/0!",
        "#REF!",
        "#NAME?",
        "#VALUE!",
        "#N/A",
        "#NULL!",
        "#NUM!",
        "#GETTING_DATA",
        "#SPILL!",
        "#CALC!",
    }
)

# Placeholder patterns to filter out (exact matches after stripping)
PLACEHOLDER_PATTERNS = frozenset(
    {
        "...",
        "…",  # Unicode ellipsis
        "---",
        "--",
        "-",
        "n/a",
        "N/A",
        "n.a.",
        "N.A.",
        "TBD",
        "tbd",
    }
)


def is_valid_label(text: str) -> bool:
    """
    Check if a text string is a valid label (not an error or placeholder).
    """
    stripped = text.strip()

    # Filter out empty strings
    if not stripped:
        return False

    # Filter out Excel errors
    if stripped in EXCEL_ERRORS:
        return False

    # Filter out placeholder patterns
    if stripped in PLACEHOLDER_PATTERNS:
        return False

    # Filter out strings that are just repeated punctuation (like "....." or "----")
    if stripped and all(c in ".-…" for c in stripped):
        return False

    return True


import re as _re

# Patterns that identify a header cell as the ProjectionYear anchor (offset 0).
_ANCHOR_PATTERNS: list[_re.Pattern[str]] = [
    _re.compile(r"^=\+?ProjectionYear$", _re.IGNORECASE),
    _re.compile(r"^=\+?'?Macro-Debt_Data'?!U[45]$"),
    _re.compile(r"^=\+?U4$"),
]

# =<COL><ROW>+1  or  =+<COL><ROW>+1
_PLUS_ONE_RE = _re.compile(r"^=\+?([A-Z]{1,3})(\d+)\+1$")
# =<COL><ROW>-1  or  =+<COL><ROW>-1
_MINUS_ONE_RE = _re.compile(r"^=\+?([A-Z]{1,3})(\d+)-1$")
# =+<COL><OTHER_ROW>  (same column, different row — row-copy pattern)
_ROW_COPY_RE = _re.compile(r"^=\+?([A-Z]{1,3})(\d+)$")


def _is_anchor_formula(formula: str) -> bool:
    return any(p.match(formula) for p in _ANCHOR_PATTERNS)


def detect_year_offset_headers(
    ws_formulas: Worksheet,
    ws_values: Worksheet,
    sheet: str,
    header_row: int,
) -> dict[int, int]:
    """
    Detect year-offset header cells by formula pattern analysis.

    Scans a header row for cells whose formulas reference ProjectionYear
    (directly or via +1/-1 chains) and returns a mapping of column index
    to integer offset (0 = projection year, positive = future, negative = past).
    """
    max_col = ws_formulas.max_column or 1

    formulas: dict[int, str] = {}
    values: dict[int, int] = {}
    for col in range(1, max_col + 1):
        f = ws_formulas.cell(row=header_row, column=col).value
        v = ws_values.cell(row=header_row, column=col).value
        if isinstance(f, str) and f.startswith("="):
            formulas[col] = f
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            values[col] = int(v)

    offsets: dict[int, int] = {}

    # Pass 1: find anchor cells (offset 0)
    for col, f in formulas.items():
        if _is_anchor_formula(f):
            offsets[col] = values.get(col, 0)

    # Pass 2: walk +1/-1 chains outward from known offsets.
    # Repeat until no new cells are resolved.
    changed = True
    while changed:
        changed = False
        for col, f in formulas.items():
            if col in offsets:
                continue
            m = _PLUS_ONE_RE.match(f)
            if m:
                ref_col_letter, ref_row_str = m.group(1), m.group(2)
                if int(ref_row_str) == header_row:
                    ref_col = openpyxl.utils.cell.column_index_from_string(ref_col_letter)
                    if ref_col in offsets:
                        offsets[col] = offsets[ref_col] + 1
                        changed = True
                        continue
            m = _MINUS_ONE_RE.match(f)
            if m:
                ref_col_letter, ref_row_str = m.group(1), m.group(2)
                if int(ref_row_str) == header_row:
                    ref_col = openpyxl.utils.cell.column_index_from_string(ref_col_letter)
                    if ref_col in offsets:
                        offsets[col] = offsets[ref_col] - 1
                        changed = True
                        continue

    # Pass 3: neighbor-consistency propagation. For any unresolved formula
    # cell with an integer cached value, if a resolved neighbor (col ± 1)
    # has an offset differing by exactly 1, adopt the cached value.
    # This handles cross-sheet references (e.g. ='Macro-Debt_Data'!T5)
    # and row-copy formulas (e.g. =+E50) without needing to pattern-match
    # every formula variant.
    if offsets:
        changed = True
        while changed:
            changed = False
            for col in sorted(formulas.keys()):
                if col in offsets or col not in values:
                    continue
                v = values[col]
                left = offsets.get(col - 1)
                right = offsets.get(col + 1)
                if left is not None and v == left + 1:
                    offsets[col] = v
                    changed = True
                elif right is not None and v == right - 1:
                    offsets[col] = v
                    changed = True
    else:
        # No anchors found — check if ALL formula cells are row-copy
        # references with integer cached values forming a contiguous sequence.
        row_copy_candidates: dict[int, int] = {}
        all_row_copy = True
        for col, f in formulas.items():
            m = _ROW_COPY_RE.match(f)
            if m:
                ref_col_letter, ref_row_str = m.group(1), m.group(2)
                ref_col = openpyxl.utils.cell.column_index_from_string(ref_col_letter)
                ref_row = int(ref_row_str)
                if ref_row != header_row and ref_col == col and col in values:
                    row_copy_candidates[col] = values[col]
                    continue
            all_row_copy = False

        if all_row_copy and row_copy_candidates:
            sorted_cols = sorted(row_copy_candidates.keys())
            vals = [row_copy_candidates[c] for c in sorted_cols]
            if all(vals[i + 1] - vals[i] == 1 for i in range(len(vals) - 1)):
                offsets = row_copy_candidates

    return offsets


_OFFSET_PREFIX = "offset:"


def is_offset_label(label: str) -> bool:
    """Return True if ``label`` is an offset label (e.g. ``"offset:0"``)."""
    if not label.startswith(_OFFSET_PREFIX):
        return False
    rest = label[len(_OFFSET_PREFIX) :]
    return rest.lstrip("-").isdigit()


def parse_offset_label(label: str) -> int:
    """Parse ``"offset:N"`` → ``N``.  Raises ValueError if not valid."""
    if not label.startswith(_OFFSET_PREFIX):
        raise ValueError(f"Not an offset label: {label!r}")
    return int(label[len(_OFFSET_PREFIX) :])


def is_year_like(value: int | float) -> bool:
    """
    Check if a numeric value looks like a year (1900-2100 range).
    """
    # Only integers can be years
    if not isinstance(value, int) or isinstance(value, bool):
        return False

    return 1900 <= value <= 2100


def dedupe_labels(labels: list[str]) -> list[str]:
    """
    De-duplicate labels while preserving their original order.
    """
    seen: set[str] = set()
    out: list[str] = []
    for label in labels:
        if label in seen:
            continue
        seen.add(label)
        out.append(label)
    return out


def find_region_config(sheet: str, row: int, col: int) -> RegionConfig | None:
    """
    Find the first matching region configuration for a cell.
    """
    for config in REGION_CONFIG:
        # Check sheet name
        if config.get("sheet") != sheet:
            continue

        # Check row bounds
        min_row = config.get("min_row")
        max_row = config.get("max_row")
        if min_row is not None and row < min_row:
            continue
        if max_row is not None and row > max_row:
            continue

        # Check column bounds
        min_col = config.get("min_col")
        max_col = config.get("max_col")
        if min_col is not None:
            min_col_idx = openpyxl.utils.cell.column_index_from_string(min_col)
            if col < min_col_idx:
                continue
        if max_col is not None:
            max_col_idx = openpyxl.utils.cell.column_index_from_string(max_col)
            if col > max_col_idx:
                continue

        return config

    return None


def get_labels_from_region_config(
    ws: Worksheet,
    row: int,
    col: int,
    config: RegionConfig,
    offset_maps: dict[int, dict[int, int]] | None = None,
) -> tuple[list[str], list[str]]:
    """
    Extract row and column labels using explicit region configuration.

    ``offset_maps`` is an optional mapping of ``header_row → {col_idx → offset}``
    produced by :func:`detect_year_offset_headers`.  When present, header cells
    that are year-offset formulas emit ``"offset:<N>"`` labels.
    """
    row_labels: list[str] = []
    col_labels: list[str] = []
    offset_maps = offset_maps or {}

    # Get row labels from specified label columns
    label_columns = config.get("label_columns", [])
    for col_letter in label_columns:
        label_col_idx = openpyxl.utils.cell.column_index_from_string(col_letter)
        cell_value = ws.cell(row=row, column=label_col_idx).value

        if cell_value is not None:
            if isinstance(cell_value, str):
                text = cell_value.strip()
                if text and is_valid_label(text):
                    row_labels.append(text)
            elif is_year_like(cell_value):
                row_labels.append(str(cell_value))

    # Get column labels from specified header rows
    header_rows = config.get("header_rows", [])
    for header_row in header_rows:
        # Check precomputed offset map first
        hr_offsets = offset_maps.get(header_row, {})
        if col in hr_offsets:
            col_labels.append(f"{_OFFSET_PREFIX}{hr_offsets[col]}")
            continue

        cell_value = ws.cell(row=header_row, column=col).value

        if cell_value is not None:
            if isinstance(cell_value, str):
                text = cell_value.strip()
                if text and is_valid_label(text):
                    col_labels.append(text)
            elif is_year_like(cell_value):
                col_labels.append(str(cell_value))

    return dedupe_labels(row_labels), dedupe_labels(col_labels)


def get_row_labels(ws: Worksheet, row: int, col: int) -> list[str]:
    """
    Scan leftward from a cell to find text labels.
    """
    labels: list[str] = []

    # Start from the column immediately left of the cell
    current_col = col - 1

    while current_col >= 1:
        cell_value = ws.cell(row=row, column=current_col).value

        # Stop if we hit a blank cell
        if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
            break

        # Collect text values
        if isinstance(cell_value, str):
            text = cell_value.strip()
            # Only append if it passes validation
            if is_valid_label(text):
                labels.append(text)
            # Continue scanning even if invalid (skip over placeholders/errors)
        elif isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
            # Check if it's a year-like integer
            if is_year_like(cell_value):
                labels.append(str(cell_value))
            else:
                # Hit a non-year numeric value - stop scanning
                break
        elif not isinstance(cell_value, bool):
            # Convert other types to string if not numeric/bool
            text = str(cell_value)
            if is_valid_label(text):
                labels.append(text)

        current_col -= 1

    return dedupe_labels(labels)


def get_column_labels(ws: Worksheet, row: int, col: int) -> list[str]:
    """
    Scan upward from a cell to find text labels.
    """
    labels: list[str] = []

    # Start from the row immediately above the cell
    current_row = row - 1

    while current_row >= 1:
        cell_value = ws.cell(row=current_row, column=col).value

        # Stop if we hit a blank cell
        if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
            break

        # Collect text values
        if isinstance(cell_value, str):
            text = cell_value.strip()
            # Only append if it passes validation
            if is_valid_label(text):
                labels.append(text)
            # Continue scanning even if invalid (skip over placeholders/errors)
        elif isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
            # Check if it's a year-like integer
            if is_year_like(cell_value):
                labels.append(str(cell_value))
            else:
                # Hit a non-year numeric value - stop scanning
                break
        elif not isinstance(cell_value, bool):
            # Convert other types to string if not numeric/bool
            text = str(cell_value)
            if is_valid_label(text):
                labels.append(text)

        current_row -= 1

    return dedupe_labels(labels)


def enrich_graph_with_labels(
    graph: DependencyGraph,
    wb_path: Path,
) -> dict[str, dict[str, Any]]:
    """
    Enrich all nodes in the graph with row and column labels.
    """
    wb_values = openpyxl.load_workbook(wb_path, data_only=True, keep_vba=True)
    wb_formulas = openpyxl.load_workbook(wb_path, data_only=False, keep_vba=True)

    try:
        # Cache worksheets by name for efficiency
        worksheets: dict[str, Worksheet] = {}

        # Precompute year-offset maps: sheet → header_row → {col → offset}
        _offset_cache: dict[str, dict[int, dict[int, int]]] = {}

        def _get_offset_maps(sheet: str, config: RegionConfig) -> dict[int, dict[int, int]]:
            if sheet not in _offset_cache:
                _offset_cache[sheet] = {}
            sheet_cache = _offset_cache[sheet]
            for hr in config.get("header_rows", []):
                if hr not in sheet_cache:
                    ws_f = wb_formulas[sheet] if sheet in wb_formulas.sheetnames else None
                    ws_v = wb_values[sheet] if sheet in wb_values.sheetnames else None
                    if ws_f is not None and ws_v is not None:
                        sheet_cache[hr] = detect_year_offset_headers(ws_f, ws_v, sheet, hr)
                    else:
                        sheet_cache[hr] = {}
            return sheet_cache

        enrichment_results: dict[str, dict[str, Any]] = {}

        for key in graph:
            node = graph.get_node(key)
            if node is None:
                continue

            # Get or cache the worksheet
            if node.sheet not in worksheets:
                if node.sheet in wb_values.sheetnames:
                    worksheets[node.sheet] = wb_values[node.sheet]
                else:
                    continue

            ws = worksheets[node.sheet]

            # Get column index from column letter
            col_idx = openpyxl.utils.cell.column_index_from_string(node.column)

            # Check for region-based configuration first
            region_config = find_region_config(node.sheet, node.row, col_idx)

            if region_config is not None:
                offset_maps = _get_offset_maps(node.sheet, region_config)
                row_labels, col_labels = get_labels_from_region_config(
                    ws, node.row, col_idx, region_config, offset_maps
                )
            else:
                # Fall back to heuristic scanning
                row_labels = get_row_labels(ws, node.row, col_idx)
                col_labels = get_column_labels(ws, node.row, col_idx)

            # Store in node metadata
            node.metadata["row_labels"] = row_labels
            node.metadata["column_labels"] = col_labels

            # Track for reporting (all nodes, not just those with labels)
            enrichment_results[key] = {
                "sheet": node.sheet,
                "address": node.address,
                "row_labels": row_labels,
                "column_labels": col_labels,
                "is_leaf": node.is_leaf,
            }

        return enrichment_results

    finally:
        wb_values.close()
        wb_formulas.close()


def export_enrichment_audit(
    graph: DependencyGraph,
    enrichment_results: dict[str, dict[str, Any]],
    output_path: Path,
) -> None:
    """
    Export enrichment results to a JSON file for auditing.
    """
    # Compute per-sheet statistics
    sheet_stats: dict[str, dict[str, int]] = defaultdict(
        lambda: {"total": 0, "with_row": 0, "with_col": 0, "with_any": 0, "none": 0}
    )

    # Group cells by sheet
    cells_by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for key, data in enrichment_results.items():
        sheet = data["sheet"]
        sheet_stats[sheet]["total"] += 1

        has_row = bool(data["row_labels"])
        has_col = bool(data["column_labels"])

        if has_row:
            sheet_stats[sheet]["with_row"] += 1
        if has_col:
            sheet_stats[sheet]["with_col"] += 1
        if has_row or has_col:
            sheet_stats[sheet]["with_any"] += 1
        else:
            sheet_stats[sheet]["none"] += 1

        cells_by_sheet[sheet].append(
            {
                "key": key,
                "address": data["address"],
                "row_labels": data["row_labels"],
                "column_labels": data["column_labels"],
                "is_leaf": data["is_leaf"],
            }
        )

    # Sort cells within each sheet by address
    for sheet in cells_by_sheet:
        cells_by_sheet[sheet].sort(key=lambda x: x["address"])

    # Build output structure
    total_nodes = len(graph)
    nodes_with_any = sum(s["with_any"] for s in sheet_stats.values())

    output = {
        "summary": {
            "total_nodes": total_nodes,
            "nodes_with_any_label": nodes_with_any,
            "nodes_without_labels": total_nodes - nodes_with_any,
            "nodes_with_row_labels": sum(s["with_row"] for s in sheet_stats.values()),
            "nodes_with_column_labels": sum(s["with_col"] for s in sheet_stats.values()),
        },
        "by_sheet": {
            sheet: {
                "statistics": dict(sheet_stats[sheet]),
                "cells": cells_by_sheet[sheet],
            }
            for sheet in sorted(sheet_stats.keys())
        },
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)


def discover_formula_cells_in_rows(
    wb_path: Path,
    sheet_name: str,
    rows: list[int],
) -> list[str]:
    """
    Scan specified rows and return sheet-qualified addresses for formula cells.

    Includes every cell that contains a formula (value starts with '='). Cached
    values are not required to be numeric, so discovery works when the template
    has no pre-filled data (e.g. base year empty, formulas evaluate to errors).
    """
    wb_formulas = openpyxl.load_workbook(wb_path, data_only=False, keep_vba=True)
    try:
        if sheet_name not in wb_formulas.sheetnames:
            print(f"  Warning: Sheet '{sheet_name}' not found")
            return []

        ws_formulas = wb_formulas[sheet_name]
        targets: list[str] = []

        for row in rows:
            max_col = ws_formulas.max_column or 1
            for col_idx in range(1, max_col + 1):
                cell_formula = ws_formulas.cell(row=row, column=col_idx)
                if isinstance(cell_formula.value, str) and cell_formula.value.startswith("="):
                    col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
                    targets.append(f"{sheet_name}!{col_letter}{row}")

        return targets
    finally:
        wb_formulas.close()
