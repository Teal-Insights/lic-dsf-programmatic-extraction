#!/usr/bin/env python3
"""
Shared pipeline utilities for LIC-DSF export and input grouping.
"""

from __future__ import annotations

from pathlib import Path
import re

import openpyxl
import openpyxl.utils.cell
from excel_grapher import DependencyGraph, create_dependency_graph
from excel_formula_expander.codegen import CodeGenerator
from openpyxl.worksheet.worksheet import Worksheet

STRING_CONSTANT_EXCLUDES = {
    "START!K10",
    "'BLEND floating calculations WB'!C6",
    "'Input 6(optional)-Standard Test'!C4",
    "'Input 6(optional)-Standard Test'!C5",
    "'Input 6(optional)-Standard Test'!C7",
    "'Input 6(optional)-Standard Test'!C8",
    "'Input 6(optional)-Standard Test'!D18",
    "'Input 6(optional)-Standard Test'!D26",
    "'Input 6(optional)-Standard Test'!D30",
    "'Input 6(optional)-Standard Test'!D33",
    "'Input 6(optional)-Standard Test'!D8",
    "'Input 6(optional)-Standard Test'!D9",
}
BLANK_CONSTANT_EXCLUDES = {
    "'Input 6(optional)-Standard Test'!D8",
    "'Input 6(optional)-Standard Test'!D9",
}
_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z_][0-9A-Za-z_]*$")


def _format_sheet_name(sheet: str) -> str:
    """
    Format a sheet name for sheet-qualified Excel addresses.

    Only quote sheet names when needed (e.g., spaces or punctuation), and escape
    embedded single quotes by doubling them.
    """
    if _SAFE_SHEET_NAME_RE.match(sheet):
        return sheet
    escaped = sheet.replace("'", "''")
    return f"'{escaped}'"


def _format_address(sheet: str, a1: str) -> str:
    return f"{_format_sheet_name(sheet)}!{a1}"


def _is_blank_value(value: object) -> bool:
    return value is None or value == ""


def discover_targets(workbook: Path) -> list[str]:
    from lic_dsf_labels import INDICATOR_CONFIG, discover_formula_cells_in_rows

    all_targets: list[str] = []
    for config in INDICATOR_CONFIG:
        sheet = config["sheet"]
        rows = config["indicator_rows"]
        all_targets.extend(discover_formula_cells_in_rows(workbook, sheet, rows))
    # Preserve order while de-duping.
    return list(dict.fromkeys(all_targets))


def build_graph(workbook: Path, targets: list[str], max_depth: int) -> DependencyGraph:
    return create_dependency_graph(workbook, targets, load_values=False, max_depth=max_depth)


def populate_leaf_values(graph: DependencyGraph, workbook: Path) -> None:
    """
    Populate values for leaf (non-formula) nodes from cached workbook values.

    Code generation only needs `node.value` for leaf cells; formulas are emitted
    from `node.formula`.
    """
    wb = openpyxl.load_workbook(workbook, data_only=True, keep_vba=True)
    try:
        worksheets: dict[str, Worksheet] = {}
        for key in graph:
            node = graph.get_node(key)
            if node is None or node.formula is not None:
                continue
            if node.sheet not in worksheets:
                if node.sheet not in wb.sheetnames:
                    continue
                worksheets[node.sheet] = wb[node.sheet]
            ws = worksheets[node.sheet]
            col_idx = openpyxl.utils.cell.column_index_from_string(node.column)
            node.value = ws.cell(row=node.row, column=col_idx).value
    finally:
        wb.close()


def iter_string_constant_addresses(
    graph: DependencyGraph, exclude_addresses: set[str] | None = None
) -> list[str]:
    exclude = exclude_addresses or set()
    ranges: list[str] = []
    for key in graph:
        node = graph.get_node(key)
        if node is None or node.formula is not None or not node.is_leaf:
            continue
        if not isinstance(node.value, str):
            continue
        address = _format_address(node.sheet, node.address)
        if address in exclude:
            continue
        ranges.append(address)
    return list(dict.fromkeys(ranges))


def classify_input_addresses(
    graph: DependencyGraph,
    targets: list[str],
    *,
    constant_types: set[str] | None = None,
    constant_ranges: list[str] | None = None,
    constant_blanks: bool = False,
    blank_excludes: set[str] | None = None,
    attach_to_graph: bool = True,
) -> set[str]:
    inputs, _constants = CodeGenerator(graph).classify_leaf_nodes(
        targets,
        constant_types=constant_types,
        constant_ranges=constant_ranges,
        constant_blanks=constant_blanks,
        attach_to_graph=attach_to_graph,
    )
    input_addresses = {str(address) for address in inputs}
    if not blank_excludes:
        return input_addresses

    address_to_node: dict[str, object] = {}
    for key in graph:
        node = graph.get_node(key)
        if node is None:
            continue
        address_to_node[_format_address(node.sheet, node.address)] = node

    for address in blank_excludes:
        if address in input_addresses:
            continue
        node = address_to_node.get(address)
        if node is None:
            continue
        if node.formula is not None or not node.is_leaf:
            continue
        if _is_blank_value(node.value):
            input_addresses.add(address)
    return input_addresses


def enrich_graph(graph: DependencyGraph, workbook: Path) -> dict[str, dict[str, object]]:
    from lic_dsf_labels import enrich_graph_with_labels

    return enrich_graph_with_labels(graph, workbook)


def export_enrichment_audit(
    graph: DependencyGraph, enrichment_results: dict[str, dict[str, object]], path: Path
) -> None:
    from lic_dsf_labels import export_enrichment_audit

    export_enrichment_audit(graph, enrichment_results, path)
