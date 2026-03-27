"""
Template-specific configuration for LIC-DSF template 2026-01-31.

This module contains all configuration that is specific to this template version:
workbook path, export ranges, region config, constraints, and constant excludes.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Literal, TypedDict, Annotated, cast
from openpyxl.utils.cell import range_boundaries, get_column_letter

from excel_grapher import constrain
from excel_grapher.grapher import DynamicRefConfig
from excel_grapher.grapher.dynamic_refs import format_key
from excel_grapher.core.cell_types import Between

from ...lic_dsf_config import ExportRangeConfig, WorkbookMetadata
from ...lic_dsf_labels import RegionConfig


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

WORKBOOK_PATH = Path("workbooks/lic-dsf-template-2025-08-12.xlsm")
WORKBOOK_TEMPLATE_URL = (
    "https://thedocs.worldbank.org/en/doc/f0ade6bcf85b6f98dbeb2c39a2b7770c-0360012025/original/LIC-DSF-IDA21-Template-08-12-2025-vf.xlsm"
)
WORKBOOK_METADATA: WorkbookMetadata = {
    "creator": "spalazzo",
    "created": "2002-02-01",
    "modified": "2025-09-18T22:03:17",
}

# ---------------------------------------------------------------------------
# Export package
# ---------------------------------------------------------------------------

PACKAGE_NAME = "lic_dsf_2025_08_12"
EXPORT_DIR = Path("dist/lic-dsf-2025-08-12")

# ---------------------------------------------------------------------------
# Export ranges
# ---------------------------------------------------------------------------

STRESS_TEST_ROW_LABELS: list[str] = [
    "Baseline",
    "A1. Key variables at their historical averages in 2024-2034 2/",
    "B1. Real GDP growth",
    "B2. Primary balance",
    "B3. Exports",
    "B4. Other flows 3/",
    "B5. Depreciation",
    "B6. Combination of B1-B5",
    "",
    "C1. Combined contingent liabilities",
    "C2. Natural disaster",
    "C3. Commodity price",
    "C4. Market Financing",
    "A2. Alternative Scenario :[Customize, enter title]",
]


STRESS_TEST_BLOCKS: list[tuple[str, int]] = [
    ("PV of Debt-to-GDP Ratio", 239),
    ("PV of Debt-to-Revenue Ratio", 281),
    ("Debt Service-to-Revenue Ratio", 318),
    ("Debt Service-to-GDP Ratio", 351),
]


FIGURE_DATA_ROWS: list[int] = [
    # Figure 1 (Output 2-1 Stress_Charts_Ex)
    51,
    61,
    62,
    63,
    64,
    66,
    93,
    103,
    104,
    105,
    106,
    108,
    135,
    145,
    146,
    147,
    148,
    150,
    177,
    187,
    188,
    189,
    190,
    192,
    # Figure 2 extras (Output 2-2 Stress_Charts_Pub)
    263,
    264,
    265,
    267,
    306,
    341,
    342,
    343,
]


EXPORT_FIXED_RANGES: list[ExportRangeConfig] = [
    {
        "label": "External DSA risk rating signals",
        "range_spec": "'Chart Data'!D10:D17",
        "entrypoint_mode": "per_cell",
    },
    {
        "label": "Fiscal (Total Public Debt) risk rating signals",
        "range_spec": "'Chart Data'!I10:I14",
        "entrypoint_mode": "per_cell",
    },
    {
        "label": "Applicable tailored stress test signals",
        "range_spec": "'Chart Data'!I17:I19",
        "entrypoint_mode": "row_group",
    },
    {
        "label": "Fiscal space for moderate risk category",
        "range_spec": "'Chart Data'!E25:E27",
        "entrypoint_mode": "row_group",
    },
    {
        "label": "Overall rating",
        "range_spec": "'Chart Data'!L10:L11",
        "entrypoint_mode": "row_group",
    },
]


def _export_chart_data_ranges() -> list[ExportRangeConfig]:
    out: list[ExportRangeConfig] = list(EXPORT_FIXED_RANGES)
    seen_row_specs = {entry["range_spec"] for entry in out}

    def add_chart_data_row(row: int, label: str) -> None:
        range_spec = f"'Chart Data'!D{row}:X{row}"
        if range_spec in seen_row_specs:
            return
        out.append(
            {
                "label": label,
                "range_spec": range_spec,
                "entrypoint_mode": "row_group",
            }
        )
        seen_row_specs.add(range_spec)

    for metric_label, start_row in STRESS_TEST_BLOCKS:
        for i, row_label in enumerate(STRESS_TEST_ROW_LABELS):
            if not row_label:
                continue
            row = start_row + i
            add_chart_data_row(row, f"{metric_label} - {row_label}")

    for row in FIGURE_DATA_ROWS:
        add_chart_data_row(row, f"Figure data row {row}")

    return out


EXPORT_RANGES: list[ExportRangeConfig] = _export_chart_data_ranges()

# ---------------------------------------------------------------------------
# Region config (label extraction)
# ---------------------------------------------------------------------------

REGION_CONFIG: list[RegionConfig] = [
    {
        "sheet": "Input 5 - Local-debt Financing",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [5],
        "label_columns": ["A", "B"],
    },
    {
        "sheet": "Ext_Debt_Data",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [1, 9],
        "label_columns": ["A"],
    },
    {
        "sheet": "PV_Base",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [7],
        "label_columns": ["A", "C"],
    },
    {
        "sheet": "PV_LC_NR1",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [3],
        "label_columns": ["A", "C"],
    },
    {
        "sheet": "PV_LC_NR2",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [3],
        "label_columns": ["A", "C"],
    },
    {
        "sheet": "PV_LC_NR3",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [3],
        "label_columns": ["A", "C"],
    },
    {
        "sheet": "Input 3 - Macro-Debt data(DMX)",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [7],
        "label_columns": ["A", "B", "C"],
    },
    {
        "sheet": "Input 4 - External Financing",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [6],
        "label_columns": ["B"],
    },
    {
        "sheet": "Baseline - external",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [8],
        "label_columns": ["B"],
    },
    {
        "sheet": "Baseline - public",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [7],
        "label_columns": ["B"],
    },
    {
        "sheet": "Input 8 - SDR",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [9],
        "label_columns": ["A"],
    },
    {
        "sheet": "B1_GDP_ext",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [8],
        "label_columns": ["B", "Z"],
    },
    {
        "sheet": "B3_Exports_ext",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [8],
        "label_columns": ["B", "Z"],
    },
    {
        "sheet": "B4_other flows_ext",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [8],
        "label_columns": ["B", "Z"],
    },
    {
        "sheet": "Macro-Debt_Data",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [1, 5],
        "label_columns": ["B", "E"],
    },
    {
        "sheet": "PV Stress",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [3],
        "label_columns": ["A", "C"],
    },
    {
        "sheet": "Input 6(optional)-Standard Test",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["A", "B", "C"],
    },
    {
        "sheet": "Input 7 - Residual Financing",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["B"],
    },
    {
        "sheet": "START",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["A", "B"],
    },
    {
        "sheet": "lookup",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["A", "B"],
    },
    {
        "sheet": "translation",
        "min_row": None,
        "max_row": None,
        "min_col": None,
        "max_col": None,
        "header_rows": [],
        "label_columns": ["A", "B"],
    },
]

# ---------------------------------------------------------------------------
# Constraints (OFFSET / INDIRECT resolution)
# ---------------------------------------------------------------------------

"""
Dynamic refs (OFFSET/INDIRECT/INDEX) are resolved via a constraint-based config. Iterative workflow: run the script; if DynamicRefError is raised, the message includes the formula cell whose inputs need constraints. Inspect that cell and the row/column headers in the workbook to decide plausible input domains, use `constrain` to set a constraint (e.g., `constrain(LicDsfConstraints, "'Sheet Name'!A965", Literal["value"])` for a constant or something like `Annotated[float, Between(min=0)]` in lieu of `Literal` for a numeric constraint).

Then re-run until the graph builds. Note that if row/column labels or intentionally blank cells show up in error output, they have been referenced by a dynamic ref and must be constrained for the graph to resolve. Blank cells can be set to `Literal[None]`.

The goal is to set sensible constraints that reflect the range of sane values we will allow for the cells. To determine the plausible range of input values, investigate the cells by using enrichment_audit.json (or the heuristic label-scanning tools in src/lic_dsf_labels.py) to see their labels, and openpyxl to check their current values. In addition to the empty template workbook, workbooks/lic-dsf-template-2025-08-12.xlsm, we also have one filled out with illustrative data: workbooks/dsf-uga.xlsm.
"""

LiteralType = cast(Any, Literal)

# Constraint types for cells that feed OFFSET/INDIRECT. Keys must be address-style (e.g. "Sheet1!B1").
# Add entries when the script raises DynamicRefError: the message lists leaf cells that need
# constraints. Add each to __annotations__ (with Annotated[int, Between(lo, hi)],
# Annotated[..., FromWorkbook()], or Literal[...]) then re-run. Repeat until the graph builds.
class LicDsfConstraints(TypedDict, total=False):
    pass

# Lookup switches; treat as constants
constrain(LicDsfConstraints, "lookup!AF4", Literal["New"])
constrain(LicDsfConstraints, "lookup!AF5", Literal["Old"])

# Marker to use for applicable tailored stress test; we can treat as a constant
constrain(LicDsfConstraints, "'Chart Data'!I21", Literal[1])

# PV_Base!B9xx = CONCAT("$", A9xx, "$", $A$<row>) → INDIRECT($B9xx). Row-index cells A917, A941, A965 (fixed).
# Treat these as constants derived from the current workbook values.
constrain(LicDsfConstraints, "PV_Base!A917", Literal[64])
constrain(LicDsfConstraints, "PV_Base!A941", Literal[90])
constrain(LicDsfConstraints, "PV_Base!A965", Literal[115])

constrain(LicDsfConstraints, "PV_Base!A965", Annotated[float, Between(min=0)])

# A918:A938, A942:A962, A966:A986 each has a single cached letter D, E, …, X.
for _start, _end in [(918, 939), (942, 963), (966, 987)]:
    for _row in range(_start, _end):
        _letter = chr(ord("D") + _row - _start)
        LicDsfConstraints.__annotations__[f"PV_Base!A{_row}"] = LiteralType[_letter]

# Language selector and lookup table (feed INDIRECT/VLOOKUP for language-dependent refs).
# START!L10 = VLOOKUP(K10, lookup!BB4:BC7, 2); evaluator does not support VLOOKUP, so L10 is constrained too.
_LANG = Literal["English", "French", "Portuguese", "Spanish"]
_LANG_LOOKUP = Literal[
    "English", "French", "Portuguese", "Spanish", "Français", "Portugues", "Español"
]
constrain(LicDsfConstraints, "START!L10", _LANG)
constrain(LicDsfConstraints, "START!K10", _LANG)
constrain(LicDsfConstraints, "lookup!BB4:BC7", _LANG_LOOKUP)


# ---------------------------------------------------------------------------
# Market financing constraints
# ---------------------------------------------------------------------------

# Tailored stress test parameters (from Input 6 - Tailored Tests)
constrain(LicDsfConstraints, "C4_Market_financing!AB20", Literal[0, 1])  # New commercial debt projected
constrain(LicDsfConstraints, "C4_Market_financing!AB22", Annotated[float, Between(min=0, max=100)])  # FX depreciation shock (%)
constrain(LicDsfConstraints, "C4_Market_financing!AB23", Annotated[float, Between(min=0, max=1)])  # ER pass-through to inflation
constrain(LicDsfConstraints, "C4_Market_financing!AB25", Annotated[int, Between(min=0, max=2000)])  # Increase in cost, bps
constrain(LicDsfConstraints, "C4_Market_financing!AB28", Annotated[int, Between(min=1, max=50)])  # New maturity if original > 5y
constrain(LicDsfConstraints, "C4_Market_financing!AB29", Annotated[float, Between(min=0, max=1)])  # Maturity shortening factor if < 5y
constrain(LicDsfConstraints, "C4_Market_financing!AB30", Annotated[float, Between(min=0, max=1)])  # Grace period shortening factor

# New lending terms for the stress test (C4_Market_financing rows 35-39)
constrain(LicDsfConstraints, "C4_Market_financing!C35:C39", Annotated[int, Between(min=0, max=50)])  # Grace period
constrain(LicDsfConstraints, "C4_Market_financing!D35:D39", Annotated[int, Between(min=1, max=100)])  # Loan Maturity
constrain(LicDsfConstraints, "C4_Market_financing!I35:I39", Annotated[float, Between(min=0, max=1)])  # Interest rate

# Structural dependencies for INDEX/MATCH resolution
# 1. Set the default (None) for the bulk ranges
constrain(LicDsfConstraints, "C4_Market_financing!C4:C53", Literal[None])
constrain(LicDsfConstraints, "C4_Market_financing!D4:D77", Literal[None])
constrain(LicDsfConstraints, "C4_Market_financing!E4:G53", Literal[None])

# 2. Overlay the specific strings (overriding the None where needed)
constrain(LicDsfConstraints, "C4_Market_financing!D20:F20", Literal["Historical "])
constrain(LicDsfConstraints, "C4_Market_financing!D21:F21", Literal["Average "])
constrain(LicDsfConstraints, "C4_Market_financing!E33", Literal["Maturity - Grace (to determine bullet / amortization)"])
constrain(LicDsfConstraints, "C4_Market_financing!E34", Literal["Bullet (1) or Amort. (>1)"])
constrain(LicDsfConstraints, "C4_Market_financing!F33", Literal["Stress test"])
constrain(LicDsfConstraints, "C4_Market_financing!F34", Literal["Maturity"])
constrain(LicDsfConstraints, "C4_Market_financing!G34", Literal["Grace"])


_countries: list[tuple[int, str]] = [
    (4, 'Afghanistan'),
    (5, 'Bangladesh'),
    (6, 'Benin'),
    (7, 'Bhutan'),
    (8, 'Burkina Faso'),
    (9, 'Burundi'),
    (10, 'Cambodia'),
    (11, 'Cameroon'),
    (12, 'Cabo Verde'),
    (13, 'Central African Republic'),
    (14, 'Chad'),
    (15, 'Comoros'),
    (16, 'Congo, DR'),
    (17, 'Congo, Republic of'),
    (18, "Cote d'Ivoire"),
    (19, 'Djibouti'),
    (20, 'Dominica'),
    (21, 'Eritrea'),
    (22, 'Ethiopia'),
    (23, 'Gambia, The'),
    (24, 'Ghana'),
    (25, 'Grenada'),
    (26, 'Guinea'),
    (27, 'Guinea-Bissau'),
    (28, 'Guyana'),
    (29, 'Haiti'),
    (30, 'Honduras'),
    (31, 'Kenya'),
    (32, 'Kiribati'),
    (33, 'Kyrgyz Republic'),
    (34, 'Lao PDR'),
    (35, 'Lesotho'),
    (36, 'Liberia'),
    (37, 'Madagascar'),
    (38, 'Malawi'),
    (39, 'Maldives'),
    (40, 'Mali'),
    (41, 'Marshall Islands'),
    (42, 'Mauritania'),
    (43, 'Micronesia'),
    (44, 'Moldova'),
    (45, 'Mozambique'),
    (46, 'Myanmar'),
    (47, 'Nepal'),
    (48, 'Nicaragua'),
    (49, 'Niger'),
    (50, 'Papua New Guinea'),
    (51, 'Rwanda'),
    (52, 'Samoa'),
    (53, 'Sao Tome & Principe'),
    (54, 'Senegal'),
    (55, 'Sierra Leone'),
    (56, 'Solomon Islands'),
    (57, 'Somalia'),
    (58, 'South Sudan'),
    (59, 'St. Lucia'),
    (60, 'St. Vincent & the Grenadines'),
    (61, 'Sudan'),
    (62, 'Tajikistan'),
    (63, 'Tanzania'),
    (64, 'Timor-Leste'),
    (65, 'Togo'),
    (66, 'Tonga'),
    (67, 'Tuvalu'),
    (68, 'Uganda'),
    (69, 'Uzbekistan'),
    (70, 'Vanuatu'),
    (71, 'Yemen, Republic of'),
    (72, 'Zambia'),
    (73, 'Zimbabwe'),
]
for _row, _name in _countries:
    constrain(LicDsfConstraints, f"lookup!C{_row}", LiteralType[_name])


def _constrain_pv_stress_com(constraints: type[Any]) -> None:
    # Ranges from the user prompt:
    # AA36:AA140, AB36:AB140, AC36:AC140, AD36:AD140, AE36:AE140, AF37:AF141, BD27:BD131,
    # D9:D140, H36:H140, I36:I140, J36:J140, K36:K140, L36:L140, M36:M140, N36:N140,
    # O36:O140, P36:P140, Q36:Q140, R36:R140, S36:S140, T36:T140, U36:U140, V36:V140,
    # W36:W140, X36:X140, Y36:Y140, Z36:Z140

    # Non-negative financial flows / values
    financial_type = Annotated[float | None, Between(0, 1e15)]

    # D9:D140 has some specific constants
    for r in range(9, 141):
        addr = f"PV_stress_com!D{r}"
        if r in (10, 22, 35):
            constrain(constraints, addr, Literal[2024])
        elif r in (23, 24, 28):
            constrain(constraints, addr, Literal[100])
        else:
            constrain(constraints, addr, financial_type)

    # Standard year-based columns (H-AE, rows 36-140)
    # H: 2028, I: 2029, ..., Z: 2046, AA: 2047, ..., AE: 2051
    cols = (
        "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",
        "AA", "AB", "AC", "AD", "AE"
    )
    for col in cols:
        for r in range(36, 141):
            constrain(constraints, f"PV_stress_com!{col}{r}", financial_type)

    # Offset ranges
    for r in range(37, 142):
        constrain(constraints, f"PV_stress_com!AF{r}", financial_type)

    for r in range(27, 132):
        constrain(constraints, f"PV_stress_com!BD{r}", financial_type)


_constrain_pv_stress_com(LicDsfConstraints)


def _constrain_pv_baseline_com(constraints: type[Any]) -> None:
    # Non-negative financial flows / values (or None for empty cells)
    financial_type = Annotated[float | None, Between(0, 1e15)]

    # D column: mixed constants and financial values
    # D7: total commercial (financial)
    constrain(constraints, "PV_baseline_com!D7", financial_type)
    # D19, D45, D71, D97, D123: "Base" normalization = 100
    for r in (19, 45, 71, 97, 123):
        constrain(constraints, f"PV_baseline_com!D{r}", Literal[100])
    # D32, D58, D84, D110, D136: "New forex borrowing (gross, USD)"
    for r in (32, 58, 84, 110, 136):
        constrain(constraints, f"PV_baseline_com!D{r}", financial_type)

    # AF33, AF59, AF85, AF111, AF137: "cumulative" in Output sections (Eurobond thru COM5)
    for r in (33, 59, 85, 111, 137):
        constrain(constraints, f"PV_baseline_com!AF{r}", financial_type)

    # BD23, BD49, BD75, BD101, BD127: "Total debt service"
    for r in (23, 49, 75, 101, 127):
        constrain(constraints, f"PV_baseline_com!BD{r}", financial_type)

    # H:AE ranges for "New forex borrowing (gross, USD)" rows
    cols = (
        "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
        "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"
    )
    for r in (32, 58, 84, 110, 136):
        for col in cols:
            constrain(constraints, f"PV_baseline_com!{col}{r}", financial_type)


_constrain_pv_baseline_com(LicDsfConstraints)

# ---------------------------------------------------------------------------
# Translation table constraints
# ---------------------------------------------------------------------------

# Translation labels referenced by dynamic formulas (OFFSET/INDIRECT).
# Column C = English, D–F = other languages (Spanish, Portuguese, French per workbook layout).
constrain(LicDsfConstraints, "translation!C90", Literal["Residency-based"])
constrain(LicDsfConstraints, "translation!C451", Literal["Grace period"])
constrain(LicDsfConstraints, "translation!C452", Literal["Loan Maturity"])
constrain(LicDsfConstraints, "translation!C898", Literal["Projections"])

constrain(LicDsfConstraints, "translation!D451", Literal["Período de gracia"])
constrain(LicDsfConstraints, "translation!E451", Literal["Prazo de carência"])
constrain(LicDsfConstraints, "translation!F451", Literal["Différé d'amortissement "])

constrain(LicDsfConstraints, "translation!D452", Literal["Vencimiento del préstamo"])
constrain(LicDsfConstraints, "translation!E452", Literal["Vencimento do empr."])
constrain(LicDsfConstraints, "translation!F452", Literal["Échéance  crédit "])

constrain(LicDsfConstraints, "translation!D898", Literal["Projections"])
constrain(LicDsfConstraints, "translation!E898", Literal["Projecções"])
constrain(LicDsfConstraints, "translation!F898", Literal["Proyecciones"])


def _get_missing_constraints(specs: list[str], constraints: type[TypedDict]) -> list[str]:
    def _normalize_sheet(sheet: str) -> str:
        """Strip surrounding single-quotes so format_key can re-add them consistently."""
        if sheet.startswith("'") and sheet.endswith("'"):
            return sheet[1:-1]
        return sheet

    def expand_spec(spec: str) -> list[str]:
        if "!" not in spec:
            return [spec]
        sheet, range_part = spec.split("!", 1)
        sheet = _normalize_sheet(sheet)
        if ":" not in range_part:
            return [format_key(sheet, range_part)]

        # Handle ranges like Sheet!A1:Sheet!B2
        if ":" in range_part and "!" in range_part.split(":")[1]:
            parts = range_part.split(":")
            start_a1 = parts[0]
            end_a1 = parts[1].split("!", 1)[1]
        else:
            start_a1, end_a1 = range_part.split(":", 1)

        min_col, min_row, max_col, max_row = range_boundaries(f"{start_a1}:{end_a1}")
        cells = []
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cells.append(format_key(sheet, f"{get_column_letter(c)}{r}"))
        return cells

    missing = []
    for spec in specs:
        expanded = expand_spec(spec)
        for cell in expanded:
            if cell not in constraints.__annotations__:
                missing.append(spec)
                break
    return missing


def _check_constraints(constraints: type[TypedDict]) -> None:
    cells_to_constrain = ['C4_Market_financing!C13', 'C4_Market_financing!C19:C4_Market_financing!C32', 'C4_Market_financing!C34', 'C4_Market_financing!C40:C4_Market_financing!C53', 'C4_Market_financing!C4:C4_Market_financing!C7', 'C4_Market_financing!C9:C4_Market_financing!C11', 'C4_Market_financing!D19:C4_Market_financing!D22', 'C4_Market_financing!D29:C4_Market_financing!D32', 'C4_Market_financing!D34', 'C4_Market_financing!D40:C4_Market_financing!D53', 'C4_Market_financing!D4:C4_Market_financing!D6', 'C4_Market_financing!D77', 'C4_Market_financing!D9:C4_Market_financing!D11', 'C4_Market_financing!E19:C4_Market_financing!E34', 'C4_Market_financing!E40:C4_Market_financing!E47', 'C4_Market_financing!E4:C4_Market_financing!E7', 'C4_Market_financing!E9:C4_Market_financing!E11', 'C4_Market_financing!F19:C4_Market_financing!F22', 'C4_Market_financing!F29:C4_Market_financing!F34', 'C4_Market_financing!F40', 'C4_Market_financing!F42:C4_Market_financing!F47', 'C4_Market_financing!F4:C4_Market_financing!F7', 'C4_Market_financing!F9:C4_Market_financing!F11', 'C4_Market_financing!G19:C4_Market_financing!G34', 'C4_Market_financing!G40', 'C4_Market_financing!G42:C4_Market_financing!G47', 'C4_Market_financing!G4:C4_Market_financing!G7', 'C4_Market_financing!G9:C4_Market_financing!G11', "'Chart Data'!I21", 'Ext_Debt_Data!AA403:Ext_Debt_Data!AG403', 'Ext_Debt_Data!F383:Ext_Debt_Data!F384', 'Input 1 - Basics!C18', 'Input 1 - Basics!C25', 'Input 1 - Basics!C33', 'Input 3 - Macro-Debt data(DMX)!AB100:Input 3 - Macro-Debt data(DMX)!AQ100', 'Input 3 - Macro-Debt data(DMX)!AB109:Input 3 - Macro-Debt data(DMX)!AQ109', 'Input 3 - Macro-Debt data(DMX)!AB111:Input 3 - Macro-Debt data(DMX)!AQ111', 'Input 3 - Macro-Debt data(DMX)!AB113:Input 3 - Macro-Debt data(DMX)!AQ113', 'Input 3 - Macro-Debt data(DMX)!AB116:Input 3 - Macro-Debt data(DMX)!AQ116', 'Input 3 - Macro-Debt data(DMX)!AB120:Input 3 - Macro-Debt data(DMX)!AQ120', 'Input 3 - Macro-Debt data(DMX)!AB122:Input 3 - Macro-Debt data(DMX)!AQ122', 'Input 3 - Macro-Debt data(DMX)!AB126:Input 3 - Macro-Debt data(DMX)!AQ126', 'Input 3 - Macro-Debt data(DMX)!AB128:Input 3 - Macro-Debt data(DMX)!AQ128', 'Input 3 - Macro-Debt data(DMX)!AB12:Input 3 - Macro-Debt data(DMX)!AQ13', 'Input 3 - Macro-Debt data(DMX)!AB132:Input 3 - Macro-Debt data(DMX)!AQ132', 'Input 3 - Macro-Debt data(DMX)!AB141:Input 3 - Macro-Debt data(DMX)!AQ144', 'Input 3 - Macro-Debt data(DMX)!AB155:Input 3 - Macro-Debt data(DMX)!AQ155', 'Input 3 - Macro-Debt data(DMX)!AB157:Input 3 - Macro-Debt data(DMX)!AQ157', 'Input 3 - Macro-Debt data(DMX)!AB166:Input 3 - Macro-Debt data(DMX)!AQ169', 'Input 3 - Macro-Debt data(DMX)!AB175:Input 3 - Macro-Debt data(DMX)!AQ175', 'Input 3 - Macro-Debt data(DMX)!AB177:Input 3 - Macro-Debt data(DMX)!AQ178', 'Input 3 - Macro-Debt data(DMX)!AB180:Input 3 - Macro-Debt data(DMX)!AQ180', 'Input 3 - Macro-Debt data(DMX)!AB19:Input 3 - Macro-Debt data(DMX)!AQ20', 'Input 3 - Macro-Debt data(DMX)!AB22:Input 3 - Macro-Debt data(DMX)!AQ22', 'Input 3 - Macro-Debt data(DMX)!AB24:Input 3 - Macro-Debt data(DMX)!AQ24', 'Input 3 - Macro-Debt data(DMX)!AB26:Input 3 - Macro-Debt data(DMX)!AQ27', 'Input 3 - Macro-Debt data(DMX)!AB30:Input 3 - Macro-Debt data(DMX)!AQ30', 'Input 3 - Macro-Debt data(DMX)!AB34:Input 3 - Macro-Debt data(DMX)!AQ35', 'Input 3 - Macro-Debt data(DMX)!AB38:Input 3 - Macro-Debt data(DMX)!AQ38', 'Input 3 - Macro-Debt data(DMX)!AB41:Input 3 - Macro-Debt data(DMX)!AQ41', 'Input 3 - Macro-Debt data(DMX)!AB43:Input 3 - Macro-Debt data(DMX)!AQ43', 'Input 3 - Macro-Debt data(DMX)!AB52:Input 3 - Macro-Debt data(DMX)!AQ52', 'Input 3 - Macro-Debt data(DMX)!AB55:Input 3 - Macro-Debt data(DMX)!AQ55', 'Input 3 - Macro-Debt data(DMX)!AB57:Input 3 - Macro-Debt data(DMX)!AQ59', 'Input 3 - Macro-Debt data(DMX)!AB65:Input 3 - Macro-Debt data(DMX)!AQ65', 'Input 3 - Macro-Debt data(DMX)!AB70:Input 3 - Macro-Debt data(DMX)!AQ70', 'Input 3 - Macro-Debt data(DMX)!AB72:Input 3 - Macro-Debt data(DMX)!AQ72', 'Input 3 - Macro-Debt data(DMX)!AB74:Input 3 - Macro-Debt data(DMX)!AQ74', 'Input 3 - Macro-Debt data(DMX)!AB77:Input 3 - Macro-Debt data(DMX)!AQ77', 'Input 3 - Macro-Debt data(DMX)!AB81:Input 3 - Macro-Debt data(DMX)!AQ81', 'Input 3 - Macro-Debt data(DMX)!AB83:Input 3 - Macro-Debt data(DMX)!AQ83', 'Input 3 - Macro-Debt data(DMX)!AB87:Input 3 - Macro-Debt data(DMX)!AQ87', 'Input 3 - Macro-Debt data(DMX)!AB89:Input 3 - Macro-Debt data(DMX)!AQ89', 'Input 3 - Macro-Debt data(DMX)!AB93:Input 3 - Macro-Debt data(DMX)!AQ93', 'Input 3 - Macro-Debt data(DMX)!AB95:Input 3 - Macro-Debt data(DMX)!AQ95', 'Input 3 - Macro-Debt data(DMX)!AR100', 'Input 3 - Macro-Debt data(DMX)!AR109', 'Input 3 - Macro-Debt data(DMX)!AR111', 'Input 3 - Macro-Debt data(DMX)!AR113', 'Input 3 - Macro-Debt data(DMX)!AR116', 'Input 3 - Macro-Debt data(DMX)!AR120', 'Input 3 - Macro-Debt data(DMX)!AR122', 'Input 3 - Macro-Debt data(DMX)!AR126', 'Input 3 - Macro-Debt data(DMX)!AR128', 'Input 3 - Macro-Debt data(DMX)!AR12:Input 3 - Macro-Debt data(DMX)!AR13', 'Input 3 - Macro-Debt data(DMX)!AR132', 'Input 3 - Macro-Debt data(DMX)!AR141:Input 3 - Macro-Debt data(DMX)!AR144', 'Input 3 - Macro-Debt data(DMX)!AR147', 'Input 3 - Macro-Debt data(DMX)!AR155', 'Input 3 - Macro-Debt data(DMX)!AR157', 'Input 3 - Macro-Debt data(DMX)!AR166:Input 3 - Macro-Debt data(DMX)!AR169', 'Input 3 - Macro-Debt data(DMX)!AR175', 'Input 3 - Macro-Debt data(DMX)!AR177:Input 3 - Macro-Debt data(DMX)!AR178', 'Input 3 - Macro-Debt data(DMX)!AR180', 'Input 3 - Macro-Debt data(DMX)!AR19:Input 3 - Macro-Debt data(DMX)!AR20', 'Input 3 - Macro-Debt data(DMX)!AR22', 'Input 3 - Macro-Debt data(DMX)!AR24', 'Input 3 - Macro-Debt data(DMX)!AR26:Input 3 - Macro-Debt data(DMX)!AR27', 'Input 3 - Macro-Debt data(DMX)!AR30', 'Input 3 - Macro-Debt data(DMX)!AR34:Input 3 - Macro-Debt data(DMX)!AR35', 'Input 3 - Macro-Debt data(DMX)!AR38', 'Input 3 - Macro-Debt data(DMX)!AR41', 'Input 3 - Macro-Debt data(DMX)!AR43', 'Input 3 - Macro-Debt data(DMX)!AR52', 'Input 3 - Macro-Debt data(DMX)!AR55', 'Input 3 - Macro-Debt data(DMX)!AR57:Input 3 - Macro-Debt data(DMX)!AR59', 'Input 3 - Macro-Debt data(DMX)!AR65', 'Input 3 - Macro-Debt data(DMX)!AR70', 'Input 3 - Macro-Debt data(DMX)!AR72', 'Input 3 - Macro-Debt data(DMX)!AR74', 'Input 3 - Macro-Debt data(DMX)!AR77', 'Input 3 - Macro-Debt data(DMX)!AR81', 'Input 3 - Macro-Debt data(DMX)!AR83', 'Input 3 - Macro-Debt data(DMX)!AR87', 'Input 3 - Macro-Debt data(DMX)!AR89', 'Input 3 - Macro-Debt data(DMX)!AR93', 'Input 3 - Macro-Debt data(DMX)!AR95', 'Input 3 - Macro-Debt data(DMX)!BP65', 'Input 3 - Macro-Debt data(DMX)!BP70', 'Input 3 - Macro-Debt data(DMX)!BP72', 'Input 3 - Macro-Debt data(DMX)!BP74', 'Input 3 - Macro-Debt data(DMX)!BP77', 'Input 3 - Macro-Debt data(DMX)!BP81', 'Input 3 - Macro-Debt data(DMX)!BP83', 'Input 3 - Macro-Debt data(DMX)!BP87', 'Input 3 - Macro-Debt data(DMX)!BP89', 'Input 3 - Macro-Debt data(DMX)!BP93', 'Input 3 - Macro-Debt data(DMX)!M12:Input 3 - Macro-Debt data(DMX)!M13', 'Input 3 - Macro-Debt data(DMX)!M35', 'Input 3 - Macro-Debt data(DMX)!N12:Input 3 - Macro-Debt data(DMX)!N13', 'Input 3 - Macro-Debt data(DMX)!N142', 'Input 3 - Macro-Debt data(DMX)!N166:Input 3 - Macro-Debt data(DMX)!N167', 'Input 3 - Macro-Debt data(DMX)!N20', 'Input 3 - Macro-Debt data(DMX)!N34:Input 3 - Macro-Debt data(DMX)!N35', 'Input 3 - Macro-Debt data(DMX)!N41', 'Input 3 - Macro-Debt data(DMX)!N43', 'Input 3 - Macro-Debt data(DMX)!N53', 'Input 3 - Macro-Debt data(DMX)!N59', 'Input 3 - Macro-Debt data(DMX)!V12:Input 3 - Macro-Debt data(DMX)!V13', 'Input 3 - Macro-Debt data(DMX)!V20', 'Input 3 - Macro-Debt data(DMX)!V35', 'Input 3 - Macro-Debt data(DMX)!W12:Input 3 - Macro-Debt data(DMX)!W13', 'Input 3 - Macro-Debt data(DMX)!W138:Input 3 - Macro-Debt data(DMX)!W139', 'Input 3 - Macro-Debt data(DMX)!W142', 'Input 3 - Macro-Debt data(DMX)!W161:Input 3 - Macro-Debt data(DMX)!W164', 'Input 3 - Macro-Debt data(DMX)!W166:Input 3 - Macro-Debt data(DMX)!W167', 'Input 3 - Macro-Debt data(DMX)!W19:Input 3 - Macro-Debt data(DMX)!W20', 'Input 3 - Macro-Debt data(DMX)!W34:Input 3 - Macro-Debt data(DMX)!W35', 'Input 3 - Macro-Debt data(DMX)!W41', 'Input 3 - Macro-Debt data(DMX)!W43', 'Input 3 - Macro-Debt data(DMX)!W51:Input 3 - Macro-Debt data(DMX)!W53', 'Input 3 - Macro-Debt data(DMX)!W55', 'Input 3 - Macro-Debt data(DMX)!W57:Input 3 - Macro-Debt data(DMX)!W59', 'Input 3 - Macro-Debt data(DMX)!X100', 'Input 3 - Macro-Debt data(DMX)!X109', 'Input 3 - Macro-Debt data(DMX)!X111', 'Input 3 - Macro-Debt data(DMX)!X113', 'Input 3 - Macro-Debt data(DMX)!X116', 'Input 3 - Macro-Debt data(DMX)!X120', 'Input 3 - Macro-Debt data(DMX)!X122', 'Input 3 - Macro-Debt data(DMX)!X126', 'Input 3 - Macro-Debt data(DMX)!X128', 'Input 3 - Macro-Debt data(DMX)!X12:Input 3 - Macro-Debt data(DMX)!X13', 'Input 3 - Macro-Debt data(DMX)!X132', 'Input 3 - Macro-Debt data(DMX)!X141:Input 3 - Macro-Debt data(DMX)!X144', 'Input 3 - Macro-Debt data(DMX)!X147', 'Input 3 - Macro-Debt data(DMX)!X149:Input 3 - Macro-Debt data(DMX)!X150', 'Input 3 - Macro-Debt data(DMX)!X152', 'Input 3 - Macro-Debt data(DMX)!X154:Input 3 - Macro-Debt data(DMX)!X155', 'Input 3 - Macro-Debt data(DMX)!X157', 'Input 3 - Macro-Debt data(DMX)!X166:Input 3 - Macro-Debt data(DMX)!X169', 'Input 3 - Macro-Debt data(DMX)!X172:Input 3 - Macro-Debt data(DMX)!X173', 'Input 3 - Macro-Debt data(DMX)!X175', 'Input 3 - Macro-Debt data(DMX)!X177:Input 3 - Macro-Debt data(DMX)!X178', 'Input 3 - Macro-Debt data(DMX)!X180', 'Input 3 - Macro-Debt data(DMX)!X19:Input 3 - Macro-Debt data(DMX)!X20', 'Input 3 - Macro-Debt data(DMX)!X22', 'Input 3 - Macro-Debt data(DMX)!X24', 'Input 3 - Macro-Debt data(DMX)!X26:Input 3 - Macro-Debt data(DMX)!X27', 'Input 3 - Macro-Debt data(DMX)!X30', 'Input 3 - Macro-Debt data(DMX)!X35', 'Input 3 - Macro-Debt data(DMX)!X41', 'Input 3 - Macro-Debt data(DMX)!X52', 'Input 3 - Macro-Debt data(DMX)!X55', 'Input 3 - Macro-Debt data(DMX)!X57:Input 3 - Macro-Debt data(DMX)!X58', 'Input 3 - Macro-Debt data(DMX)!X65', 'Input 3 - Macro-Debt data(DMX)!X70', 'Input 3 - Macro-Debt data(DMX)!X72', 'Input 3 - Macro-Debt data(DMX)!X74', 'Input 3 - Macro-Debt data(DMX)!X77', 'Input 3 - Macro-Debt data(DMX)!X81', 'Input 3 - Macro-Debt data(DMX)!X83', 'Input 3 - Macro-Debt data(DMX)!X87', 'Input 3 - Macro-Debt data(DMX)!X89', 'Input 3 - Macro-Debt data(DMX)!X93', 'Input 3 - Macro-Debt data(DMX)!X95', 'Input 3 - Macro-Debt data(DMX)!Y100:Input 3 - Macro-Debt data(DMX)!AA100', 'Input 3 - Macro-Debt data(DMX)!Y109:Input 3 - Macro-Debt data(DMX)!AA109', 'Input 3 - Macro-Debt data(DMX)!Y111:Input 3 - Macro-Debt data(DMX)!AA111', 'Input 3 - Macro-Debt data(DMX)!Y113:Input 3 - Macro-Debt data(DMX)!AA113', 'Input 3 - Macro-Debt data(DMX)!Y116:Input 3 - Macro-Debt data(DMX)!AA116', 'Input 3 - Macro-Debt data(DMX)!Y120:Input 3 - Macro-Debt data(DMX)!AA120', 'Input 3 - Macro-Debt data(DMX)!Y122:Input 3 - Macro-Debt data(DMX)!AA122', 'Input 3 - Macro-Debt data(DMX)!Y126:Input 3 - Macro-Debt data(DMX)!AA126', 'Input 3 - Macro-Debt data(DMX)!Y128:Input 3 - Macro-Debt data(DMX)!AA132', 'Input 3 - Macro-Debt data(DMX)!Y12:Input 3 - Macro-Debt data(DMX)!AA13', 'Input 3 - Macro-Debt data(DMX)!Y141:Input 3 - Macro-Debt data(DMX)!AA144', 'Input 3 - Macro-Debt data(DMX)!Y155:Input 3 - Macro-Debt data(DMX)!AA155', 'Input 3 - Macro-Debt data(DMX)!Y157:Input 3 - Macro-Debt data(DMX)!AA157', 'Input 3 - Macro-Debt data(DMX)!Y166:Input 3 - Macro-Debt data(DMX)!AA169', 'Input 3 - Macro-Debt data(DMX)!Y175:Input 3 - Macro-Debt data(DMX)!AA175', 'Input 3 - Macro-Debt data(DMX)!Y177:Input 3 - Macro-Debt data(DMX)!AA178', 'Input 3 - Macro-Debt data(DMX)!Y180:Input 3 - Macro-Debt data(DMX)!AA180', 'Input 3 - Macro-Debt data(DMX)!Y19:Input 3 - Macro-Debt data(DMX)!AA20', 'Input 3 - Macro-Debt data(DMX)!Y22:Input 3 - Macro-Debt data(DMX)!AA22', 'Input 3 - Macro-Debt data(DMX)!Y24:Input 3 - Macro-Debt data(DMX)!AA24', 'Input 3 - Macro-Debt data(DMX)!Y26:Input 3 - Macro-Debt data(DMX)!AA27', 'Input 3 - Macro-Debt data(DMX)!Y30:Input 3 - Macro-Debt data(DMX)!AA30', 'Input 3 - Macro-Debt data(DMX)!Y34:Input 3 - Macro-Debt data(DMX)!AA35', 'Input 3 - Macro-Debt data(DMX)!Y38:Input 3 - Macro-Debt data(DMX)!AA38', 'Input 3 - Macro-Debt data(DMX)!Y41:Input 3 - Macro-Debt data(DMX)!AA41', 'Input 3 - Macro-Debt data(DMX)!Y43:Input 3 - Macro-Debt data(DMX)!AA43', 'Input 3 - Macro-Debt data(DMX)!Y52:Input 3 - Macro-Debt data(DMX)!AA52', 'Input 3 - Macro-Debt data(DMX)!Y55:Input 3 - Macro-Debt data(DMX)!AA55', 'Input 3 - Macro-Debt data(DMX)!Y57:Input 3 - Macro-Debt data(DMX)!AA59', 'Input 3 - Macro-Debt data(DMX)!Y65:Input 3 - Macro-Debt data(DMX)!AA65', 'Input 3 - Macro-Debt data(DMX)!Y70:Input 3 - Macro-Debt data(DMX)!AA70', 'Input 3 - Macro-Debt data(DMX)!Y72:Input 3 - Macro-Debt data(DMX)!AA72', 'Input 3 - Macro-Debt data(DMX)!Y74:Input 3 - Macro-Debt data(DMX)!AA74', 'Input 3 - Macro-Debt data(DMX)!Y77:Input 3 - Macro-Debt data(DMX)!AA77', 'Input 3 - Macro-Debt data(DMX)!Y81:Input 3 - Macro-Debt data(DMX)!AA81', 'Input 3 - Macro-Debt data(DMX)!Y83:Input 3 - Macro-Debt data(DMX)!AA83', 'Input 3 - Macro-Debt data(DMX)!Y87:Input 3 - Macro-Debt data(DMX)!AA87', 'Input 3 - Macro-Debt data(DMX)!Y89:Input 3 - Macro-Debt data(DMX)!AA89', 'Input 3 - Macro-Debt data(DMX)!Y93:Input 3 - Macro-Debt data(DMX)!AA93', 'Input 3 - Macro-Debt data(DMX)!Y95:Input 3 - Macro-Debt data(DMX)!AA95', 'Input 4 - External Financing!AG10:Input 4 - External Financing!AM10', 'Input 4 - External Financing!AG19:Input 4 - External Financing!AM19', 'Input 4 - External Financing!AG21:Input 4 - External Financing!AM21', 'Input 4 - External Financing!AG23:Input 4 - External Financing!AM23', 'Input 4 - External Financing!AG26:Input 4 - External Financing!AM26', 'Input 4 - External Financing!AG30:Input 4 - External Financing!AM30', 'Input 4 - External Financing!AG32:Input 4 - External Financing!AM32', 'Input 4 - External Financing!AG36:Input 4 - External Financing!AM36', 'Input 4 - External Financing!AG38:Input 4 - External Financing!AM38', 'Input 4 - External Financing!AG42:Input 4 - External Financing!AM42', 'Input 4 - External Financing!F10', 'Input 4 - External Financing!F19', 'Input 4 - External Financing!F21', 'Input 4 - External Financing!F23', 'Input 4 - External Financing!F26', 'Input 4 - External Financing!F30', 'Input 4 - External Financing!F32', 'Input 4 - External Financing!F36', 'Input 4 - External Financing!F38:Input 4 - External Financing!F42', 'Input 4 - External Financing!F45', 'Input 4 - External Financing!G10:Input 4 - External Financing!H10', 'Input 4 - External Financing!G19:Input 4 - External Financing!H19', 'Input 4 - External Financing!G21:Input 4 - External Financing!H21', 'Input 4 - External Financing!G23:Input 4 - External Financing!H23', 'Input 4 - External Financing!G26:Input 4 - External Financing!H26', 'Input 4 - External Financing!G30:Input 4 - External Financing!H30', 'Input 4 - External Financing!G32:Input 4 - External Financing!H32', 'Input 4 - External Financing!G36:Input 4 - External Financing!H36', 'Input 4 - External Financing!G38:Input 4 - External Financing!H42', 'Input 5 - Local-debt Financing!AB250', 'Input 5 - Local-debt Financing!AB274', 'Input 5 - Local-debt Financing!AB298', 'Input 5 - Local-debt Financing!AB322', 'Input 5 - Local-debt Financing!AB488', 'Input 5 - Local-debt Financing!AB512', 'Input 5 - Local-debt Financing!AB581', 'Input 5 - Local-debt Financing!AB63', 'Input 5 - Local-debt Financing!AC63', 'Input 5 - Local-debt Financing!AD108', 'Input 5 - Local-debt Financing!AD110', 'Input 5 - Local-debt Financing!AD188', 'Input 5 - Local-debt Financing!AD191', 'Input 5 - Local-debt Financing!AD193', 'Input 5 - Local-debt Financing!AD93', 'Input 5 - Local-debt Financing!AD95', 'Input 5 - Local-debt Financing!AE108', 'Input 5 - Local-debt Financing!AE110', 'Input 5 - Local-debt Financing!AE250', 'Input 5 - Local-debt Financing!AE254', 'Input 5 - Local-debt Financing!AE274', 'Input 5 - Local-debt Financing!AE278', 'Input 5 - Local-debt Financing!AE298', 'Input 5 - Local-debt Financing!AE302', 'Input 5 - Local-debt Financing!AE322', 'Input 5 - Local-debt Financing!AE392', 'Input 5 - Local-debt Financing!AE461', 'Input 5 - Local-debt Financing!AE93', 'Input 5 - Local-debt Financing!AE95', 'Input 5 - Local-debt Financing!AF108', 'Input 5 - Local-debt Financing!AF110', 'Input 5 - Local-debt Financing!AF250', 'Input 5 - Local-debt Financing!AF274', 'Input 5 - Local-debt Financing!AF298', 'Input 5 - Local-debt Financing!AF322', 'Input 5 - Local-debt Financing!AF392', 'Input 5 - Local-debt Financing!AF461', 'Input 5 - Local-debt Financing!AF488', 'Input 5 - Local-debt Financing!AF512', 'Input 5 - Local-debt Financing!AF93', 'Input 5 - Local-debt Financing!AF95', 'Input 5 - Local-debt Financing!AG108:Input 5 - Local-debt Financing!AJ108', 'Input 5 - Local-debt Financing!AG110:Input 5 - Local-debt Financing!AJ110', 'Input 5 - Local-debt Financing!AG250:Input 5 - Local-debt Financing!AJ250', 'Input 5 - Local-debt Financing!AG254:Input 5 - Local-debt Financing!AJ254', 'Input 5 - Local-debt Financing!AG274:Input 5 - Local-debt Financing!AJ274', 'Input 5 - Local-debt Financing!AG278:Input 5 - Local-debt Financing!AJ278', 'Input 5 - Local-debt Financing!AG298:Input 5 - Local-debt Financing!AJ298', 'Input 5 - Local-debt Financing!AG302:Input 5 - Local-debt Financing!AJ302', 'Input 5 - Local-debt Financing!AG322:Input 5 - Local-debt Financing!AJ322', 'Input 5 - Local-debt Financing!AG392:Input 5 - Local-debt Financing!AJ392', 'Input 5 - Local-debt Financing!AG461:Input 5 - Local-debt Financing!AJ461', 'Input 5 - Local-debt Financing!AG468:Input 5 - Local-debt Financing!AJ468', 'Input 5 - Local-debt Financing!AG488:Input 5 - Local-debt Financing!AJ488', 'Input 5 - Local-debt Financing!AG492:Input 5 - Local-debt Financing!AJ492', 'Input 5 - Local-debt Financing!AG512:Input 5 - Local-debt Financing!AJ512', 'Input 5 - Local-debt Financing!AG93:Input 5 - Local-debt Financing!AJ93', 'Input 5 - Local-debt Financing!AG95:Input 5 - Local-debt Financing!AJ95', 'Input 5 - Local-debt Financing!AK250:Input 5 - Local-debt Financing!AX250', 'Input 5 - Local-debt Financing!AK254:Input 5 - Local-debt Financing!AX254', 'Input 5 - Local-debt Financing!AK274:Input 5 - Local-debt Financing!AX274', 'Input 5 - Local-debt Financing!AK278:Input 5 - Local-debt Financing!AX278', 'Input 5 - Local-debt Financing!AK298:Input 5 - Local-debt Financing!AX298', 'Input 5 - Local-debt Financing!AK302:Input 5 - Local-debt Financing!AX302', 'Input 5 - Local-debt Financing!AK322:Input 5 - Local-debt Financing!AX322', 'Input 5 - Local-debt Financing!AK392:Input 5 - Local-debt Financing!AX392', 'Input 5 - Local-debt Financing!AK461:Input 5 - Local-debt Financing!AX461', 'Input 5 - Local-debt Financing!AK468:Input 5 - Local-debt Financing!AX468', 'Input 5 - Local-debt Financing!AK488:Input 5 - Local-debt Financing!AX488', 'Input 5 - Local-debt Financing!AK492:Input 5 - Local-debt Financing!AX492', 'Input 5 - Local-debt Financing!AK512:Input 5 - Local-debt Financing!AX512', 'Input 5 - Local-debt Financing!AY254', 'Input 5 - Local-debt Financing!AY278', 'Input 5 - Local-debt Financing!AY302', 'Input 5 - Local-debt Financing!AY392', 'Input 5 - Local-debt Financing!AY468', 'Input 5 - Local-debt Financing!AY492', 'Input 5 - Local-debt Financing!BA250', 'Input 5 - Local-debt Financing!BA274', 'Input 5 - Local-debt Financing!BA298', 'Input 5 - Local-debt Financing!BA322', 'Input 5 - Local-debt Financing!BA392', 'Input 5 - Local-debt Financing!BA463', 'Input 5 - Local-debt Financing!BB250:Input 5 - Local-debt Financing!BT250', 'Input 5 - Local-debt Financing!BB274:Input 5 - Local-debt Financing!BT274', 'Input 5 - Local-debt Financing!BB298:Input 5 - Local-debt Financing!BT298', 'Input 5 - Local-debt Financing!BB322:Input 5 - Local-debt Financing!BT322', 'Input 5 - Local-debt Financing!BB392:Input 5 - Local-debt Financing!BT392', 'Input 5 - Local-debt Financing!BB463:Input 5 - Local-debt Financing!BT463', 'Input 5 - Local-debt Financing!BB488:Input 5 - Local-debt Financing!BT488', 'Input 5 - Local-debt Financing!BB512:Input 5 - Local-debt Financing!BT512', 'Input 5 - Local-debt Financing!BU392', 'Input 5 - Local-debt Financing!BU463', 'Input 5 - Local-debt Financing!C10', 'Input 5 - Local-debt Financing!C100:Input 5 - Local-debt Financing!C101', 'Input 5 - Local-debt Financing!C104:Input 5 - Local-debt Financing!C106', 'Input 5 - Local-debt Financing!C108:Input 5 - Local-debt Financing!C110', 'Input 5 - Local-debt Financing!C16', 'Input 5 - Local-debt Financing!C18', 'Input 5 - Local-debt Financing!C20', 'Input 5 - Local-debt Financing!C22', 'Input 5 - Local-debt Financing!C78', 'Input 5 - Local-debt Financing!C83', 'Input 5 - Local-debt Financing!C86', 'Input 5 - Local-debt Financing!C89:Input 5 - Local-debt Financing!C91', 'Input 5 - Local-debt Financing!C93:Input 5 - Local-debt Financing!C95', 'Input 5 - Local-debt Financing!D10', 'Input 5 - Local-debt Financing!D104', 'Input 5 - Local-debt Financing!D106', 'Input 5 - Local-debt Financing!D108', 'Input 5 - Local-debt Financing!D110', 'Input 5 - Local-debt Financing!D16', 'Input 5 - Local-debt Financing!D18', 'Input 5 - Local-debt Financing!D20', 'Input 5 - Local-debt Financing!D22', 'Input 5 - Local-debt Financing!D93', 'Input 5 - Local-debt Financing!D95', 'Input 5 - Local-debt Financing!E104', 'Input 5 - Local-debt Financing!E106', 'Input 5 - Local-debt Financing!E108', 'Input 5 - Local-debt Financing!E110', 'Input 5 - Local-debt Financing!E93', 'Input 5 - Local-debt Financing!E95', 'Input 5 - Local-debt Financing!F104', 'Input 5 - Local-debt Financing!F106', 'Input 5 - Local-debt Financing!F108', 'Input 5 - Local-debt Financing!F110', 'Input 5 - Local-debt Financing!F83', 'Input 5 - Local-debt Financing!F93', 'Input 5 - Local-debt Financing!F95', 'Input 5 - Local-debt Financing!H230', 'Input 5 - Local-debt Financing!H254', 'Input 5 - Local-debt Financing!H278', 'Input 5 - Local-debt Financing!H302', 'Input 5 - Local-debt Financing!H327', 'Input 5 - Local-debt Financing!H397', 'Input 5 - Local-debt Financing!I16', 'Input 5 - Local-debt Financing!I18', 'Input 5 - Local-debt Financing!I20', 'Input 5 - Local-debt Financing!I22', 'Input 5 - Local-debt Financing!I461', 'Input 5 - Local-debt Financing!I488', 'Input 5 - Local-debt Financing!I581', 'Input 5 - Local-debt Financing!I63', 'Input 5 - Local-debt Financing!J488:Input 5 - Local-debt Financing!M488', 'Input 5 - Local-debt Financing!J581:Input 5 - Local-debt Financing!M581', 'Input 5 - Local-debt Financing!J63:Input 5 - Local-debt Financing!M63', 'Input 5 - Local-debt Financing!N16', 'Input 5 - Local-debt Financing!N18', 'Input 5 - Local-debt Financing!N20', 'Input 5 - Local-debt Financing!N22', 'Input 5 - Local-debt Financing!N488', 'Input 5 - Local-debt Financing!N581', 'Input 5 - Local-debt Financing!N63', 'Input 5 - Local-debt Financing!O488:Input 5 - Local-debt Financing!AA488', 'Input 5 - Local-debt Financing!O581:Input 5 - Local-debt Financing!AA581', 'Input 5 - Local-debt Financing!O63:Input 5 - Local-debt Financing!AA63', 'Input 6 - Tailored Tests!C6', 'Input 6(optional)-Standard Test!C17', 'Input 6(optional)-Standard Test!C4:Input 6(optional)-Standard Test!C5', 'Input 6(optional)-Standard Test!C7:Input 6(optional)-Standard Test!C8', 'Input 6(optional)-Standard Test!D18', 'Input 6(optional)-Standard Test!D8:Input 6(optional)-Standard Test!D9', 'Input 8 - SDR!AG37', 'Input 8 - SDR!B6:Input 8 - SDR!B7', 'Input 8 - SDR!C11:Input 8 - SDR!C12', 'Input 8 - SDR!D11:Input 8 - SDR!V12', 'Input 8 - SDR!D14:Input 8 - SDR!V14', 'Input 8 - SDR!W14', 'Input 8 - SDR!X27', 'Input 8 - SDR!Y28', 'PV Stress!D147', 'PV Stress!D161', 'PV Stress!D4', 'PV Stress!E161:PV Stress!G161', 'PV Stress!H147:PV Stress!X147', 'PV Stress!Y148:PV Stress!AF148', 'PV Stress!Y162:PV Stress!AF162', 'PV Stress!Y30:PV Stress!AF30', 'PV_Base!AF23', 'PV_Base!AF272', 'PV_Base!AF298', 'PV_Base!AF350', 'PV_Base!AF376', 'PV_Base!AF480', 'PV_Base!AF506', 'PV_Base!AF610', 'PV_Base!AF636', 'PV_Base!AF740', 'PV_Base!AF766', 'PV_Base!AF818', 'PV_Base!AF844', 'PV_Base!AF896', 'PV_Base!BD366', 'PV_Base!BD470', 'PV_Base!BD496', 'PV_Base!BD600', 'PV_Base!BD626', 'PV_Base!BD730', 'PV_Base!BD756', 'PV_Base!BD808', 'PV_Base!BD834', 'PV_Base!BD886', 'PV_Base!D258', 'PV_Base!D27', 'PV_Base!D276', 'PV_Base!D284', 'PV_Base!D302', 'PV_Base!D336', 'PV_Base!D354', 'PV_Base!D362', 'PV_Base!D380', 'PV_Base!D466', 'PV_Base!D484', 'PV_Base!D49', 'PV_Base!D492', 'PV_Base!D510', 'PV_Base!D596', 'PV_Base!D614', 'PV_Base!D622', 'PV_Base!D640', 'PV_Base!D726', 'PV_Base!D744', 'PV_Base!D752', 'PV_Base!D770', 'PV_Base!D804', 'PV_Base!D822', 'PV_Base!D830', 'PV_Base!D848', 'PV_Base!D882', 'PV_Base!D9', 'PV_Base!D900', 'PV_LC_NR1!AF102', 'PV_LC_NR1!AF121', 'PV_LC_NR1!AF140', 'PV_LC_NR1!AF159', 'PV_LC_NR1!AF178', 'PV_LC_NR1!AF197', 'PV_LC_NR1!AF216', 'PV_LC_NR1!AF235', 'PV_LC_NR1!AF254', 'PV_LC_NR1!AF26', 'PV_LC_NR1!AF273', 'PV_LC_NR1!AF292', 'PV_LC_NR1!AF311', 'PV_LC_NR1!AF330', 'PV_LC_NR1!AF349', 'PV_LC_NR1!AF368', 'PV_LC_NR1!AF387', 'PV_LC_NR1!AF406', 'PV_LC_NR1!AF45', 'PV_LC_NR1!AF64', 'PV_LC_NR1!AF83', 'PV_LC_NR1!BB106', 'PV_LC_NR1!BB125', 'PV_LC_NR1!BB144', 'PV_LC_NR1!BB163', 'PV_LC_NR1!BB182', 'PV_LC_NR1!BB201', 'PV_LC_NR1!BB220', 'PV_LC_NR1!BB239', 'PV_LC_NR1!BB258', 'PV_LC_NR1!BB277', 'PV_LC_NR1!BB296', 'PV_LC_NR1!BB30', 'PV_LC_NR1!BB315', 'PV_LC_NR1!BB334', 'PV_LC_NR1!BB353', 'PV_LC_NR1!BB372', 'PV_LC_NR1!BB391', 'PV_LC_NR1!BB410', 'PV_LC_NR1!BB49', 'PV_LC_NR1!BB68', 'PV_LC_NR1!BB87', 'PV_LC_NR1!BD7', 'PV_LC_NR1!C28', 'PV_LC_NR1!D104', 'PV_LC_NR1!D107', 'PV_LC_NR1!D118', 'PV_LC_NR1!D123', 'PV_LC_NR1!D126', 'PV_LC_NR1!D137', 'PV_LC_NR1!D142', 'PV_LC_NR1!D145', 'PV_LC_NR1!D156', 'PV_LC_NR1!D161', 'PV_LC_NR1!D164', 'PV_LC_NR1!D175', 'PV_LC_NR1!D180', 'PV_LC_NR1!D183', 'PV_LC_NR1!D194', 'PV_LC_NR1!D199', 'PV_LC_NR1!D202', 'PV_LC_NR1!D213', 'PV_LC_NR1!D218', 'PV_LC_NR1!D221', 'PV_LC_NR1!D23', 'PV_LC_NR1!D232', 'PV_LC_NR1!D237', 'PV_LC_NR1!D240', 'PV_LC_NR1!D251', 'PV_LC_NR1!D256', 'PV_LC_NR1!D259', 'PV_LC_NR1!D270', 'PV_LC_NR1!D275', 'PV_LC_NR1!D278', 'PV_LC_NR1!D289', 'PV_LC_NR1!D294', 'PV_LC_NR1!D297', 'PV_LC_NR1!D308', 'PV_LC_NR1!D31', 'PV_LC_NR1!D313', 'PV_LC_NR1!D316', 'PV_LC_NR1!D327', 'PV_LC_NR1!D332', 'PV_LC_NR1!D335', 'PV_LC_NR1!D346', 'PV_LC_NR1!D351', 'PV_LC_NR1!D354', 'PV_LC_NR1!D365', 'PV_LC_NR1!D370', 'PV_LC_NR1!D373', 'PV_LC_NR1!D384', 'PV_LC_NR1!D389', 'PV_LC_NR1!D392', 'PV_LC_NR1!D403', 'PV_LC_NR1!D408', 'PV_LC_NR1!D411', 'PV_LC_NR1!D42', 'PV_LC_NR1!D47', 'PV_LC_NR1!D50', 'PV_LC_NR1!D61', 'PV_LC_NR1!D66', 'PV_LC_NR1!D69', 'PV_LC_NR1!D80', 'PV_LC_NR1!D85', 'PV_LC_NR1!D88', 'PV_LC_NR1!D99', 'PV_LC_NR1!Y6:PV_LC_NR1!AE6', 'PV_LC_NR3!AF102', 'PV_LC_NR3!AF121', 'PV_LC_NR3!AF140', 'PV_LC_NR3!AF159', 'PV_LC_NR3!AF178', 'PV_LC_NR3!AF197', 'PV_LC_NR3!AF216', 'PV_LC_NR3!AF235', 'PV_LC_NR3!AF254', 'PV_LC_NR3!AF26', 'PV_LC_NR3!AF273', 'PV_LC_NR3!AF292', 'PV_LC_NR3!AF311', 'PV_LC_NR3!AF330', 'PV_LC_NR3!AF349', 'PV_LC_NR3!AF368', 'PV_LC_NR3!AF387', 'PV_LC_NR3!AF406', 'PV_LC_NR3!AF45', 'PV_LC_NR3!AF64', 'PV_LC_NR3!AF83', 'PV_LC_NR3!BB106', 'PV_LC_NR3!BB125', 'PV_LC_NR3!BB144', 'PV_LC_NR3!BB163', 'PV_LC_NR3!BB182', 'PV_LC_NR3!BB201', 'PV_LC_NR3!BB220', 'PV_LC_NR3!BB239', 'PV_LC_NR3!BB258', 'PV_LC_NR3!BB277', 'PV_LC_NR3!BB296', 'PV_LC_NR3!BB30', 'PV_LC_NR3!BB315', 'PV_LC_NR3!BB334', 'PV_LC_NR3!BB353', 'PV_LC_NR3!BB372', 'PV_LC_NR3!BB391', 'PV_LC_NR3!BB410', 'PV_LC_NR3!BB49', 'PV_LC_NR3!BB68', 'PV_LC_NR3!BB87', 'PV_LC_NR3!BD7', 'PV_LC_NR3!C28', 'PV_LC_NR3!D104', 'PV_LC_NR3!D107', 'PV_LC_NR3!D118', 'PV_LC_NR3!D123', 'PV_LC_NR3!D126', 'PV_LC_NR3!D137', 'PV_LC_NR3!D142', 'PV_LC_NR3!D145', 'PV_LC_NR3!D156', 'PV_LC_NR3!D161', 'PV_LC_NR3!D164', 'PV_LC_NR3!D175', 'PV_LC_NR3!D180', 'PV_LC_NR3!D183', 'PV_LC_NR3!D194', 'PV_LC_NR3!D199', 'PV_LC_NR3!D202', 'PV_LC_NR3!D213', 'PV_LC_NR3!D218', 'PV_LC_NR3!D221', 'PV_LC_NR3!D23', 'PV_LC_NR3!D232', 'PV_LC_NR3!D237', 'PV_LC_NR3!D240', 'PV_LC_NR3!D251', 'PV_LC_NR3!D256', 'PV_LC_NR3!D259', 'PV_LC_NR3!D270', 'PV_LC_NR3!D275', 'PV_LC_NR3!D278', 'PV_LC_NR3!D289', 'PV_LC_NR3!D294', 'PV_LC_NR3!D297', 'PV_LC_NR3!D308', 'PV_LC_NR3!D31', 'PV_LC_NR3!D313', 'PV_LC_NR3!D316', 'PV_LC_NR3!D327', 'PV_LC_NR3!D332', 'PV_LC_NR3!D335', 'PV_LC_NR3!D346', 'PV_LC_NR3!D351', 'PV_LC_NR3!D354', 'PV_LC_NR3!D365', 'PV_LC_NR3!D370', 'PV_LC_NR3!D373', 'PV_LC_NR3!D384', 'PV_LC_NR3!D389', 'PV_LC_NR3!D392', 'PV_LC_NR3!D403', 'PV_LC_NR3!D408', 'PV_LC_NR3!D411', 'PV_LC_NR3!D42', 'PV_LC_NR3!D47', 'PV_LC_NR3!D50', 'PV_LC_NR3!D61', 'PV_LC_NR3!D66', 'PV_LC_NR3!D69', 'PV_LC_NR3!D80', 'PV_LC_NR3!D85', 'PV_LC_NR3!D88', 'PV_LC_NR3!D99', 'PV_LC_NR3!Y6:PV_LC_NR3!AE6', 'PV_baseline_com!AF111', 'PV_baseline_com!AF137', 'PV_baseline_com!AF33', 'PV_baseline_com!AF59', 'PV_baseline_com!AF85', 'PV_baseline_com!BD101', 'PV_baseline_com!BD127', 'PV_baseline_com!BD23', 'PV_baseline_com!BD49', 'PV_baseline_com!BD75', 'PV_baseline_com!D110', 'PV_baseline_com!D123', 'PV_baseline_com!D136', 'PV_baseline_com!D19', 'PV_baseline_com!D32', 'PV_baseline_com!D45', 'PV_baseline_com!D58', 'PV_baseline_com!D7', 'PV_baseline_com!D71', 'PV_baseline_com!D84', 'PV_baseline_com!D97', 'PV_baseline_com!H110:PV_baseline_com!AE110', 'PV_baseline_com!H136:PV_baseline_com!AE136', 'PV_baseline_com!H32:PV_baseline_com!AE32', 'PV_baseline_com!H58:PV_baseline_com!AE58', 'PV_baseline_com!H84:PV_baseline_com!AE84', 'PV_stress_com!AF115', 'PV_stress_com!AF141', 'PV_stress_com!AF37', 'PV_stress_com!AF63', 'PV_stress_com!AF89', 'PV_stress_com!BD105', 'PV_stress_com!BD131', 'PV_stress_com!BD27', 'PV_stress_com!BD53', 'PV_stress_com!BD79', 'PV_stress_com!D101', 'PV_stress_com!D114', 'PV_stress_com!D127', 'PV_stress_com!D140', 'PV_stress_com!D23', 'PV_stress_com!D36', 'PV_stress_com!D49', 'PV_stress_com!D62', 'PV_stress_com!D75', 'PV_stress_com!D88', 'PV_stress_com!D9', 'PV_stress_com!H114:PV_stress_com!AE114', 'PV_stress_com!H140:PV_stress_com!AE140', 'PV_stress_com!H36:PV_stress_com!AE36', 'PV_stress_com!H62:PV_stress_com!AE62', 'PV_stress_com!H88:PV_stress_com!AE88', 'lookup!AF4', 'translation!C451:translation!C452', 'translation!C898', 'translation!C90', 'translation!D451:translation!F452', 'translation!D898:translation!F898']
    missing = _get_missing_constraints(cells_to_constrain, constraints)
    if missing:
        raise ValueError(f"Missing constraints for: {missing}")


_check_constraints(LicDsfConstraints)


def get_dynamic_ref_config() -> DynamicRefConfig:
    """Return a DynamicRefConfig for constraint-based resolution of OFFSET/INDIRECT."""
    return DynamicRefConfig.from_constraints_and_workbook(
        LicDsfConstraints, WORKBOOK_PATH
    )


# ---------------------------------------------------------------------------
# Constant excludes (for input classification)
# ---------------------------------------------------------------------------

STRING_CONSTANT_EXCLUDES = {
    "START!K10",
    "'BLEND floating calculations WB'!C6",
    "'Input 6(optional)-Standard Test'!C4",
    "'Input 6(optional)-Standard Test'!C5",
    "'Input 6(optional)-Standard Test'!C7",
    "'Input 6(optional)-Standard Test'!C8",
    "'Input 6(optional)-Standard Test'!D18",
    "'Input 6(optional)-Standard Test'!D26",
    "'Input 6(optional)-Standard Test'!D30",
    "'Input 6(optional)-Standard Test'!D33",
    "'Input 6(optional)-Standard Test'!D8",
    "'Input 6(optional)-Standard Test'!D9",
}
BLANK_CONSTANT_EXCLUDES = {
    "'Input 6(optional)-Standard Test'!D8",
    "'Input 6(optional)-Standard Test'!D9",
}
