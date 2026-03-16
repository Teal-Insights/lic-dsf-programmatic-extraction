#!/usr/bin/env python3
"""
Export LIC-DSF workbook formulas as standalone Python code.

This script discovers formula targets from `lic_dsf_labels.INDICATOR_CONFIG`,
builds a dependency graph, and uses excel-grapher's CodeGenerator (exporter) to
emit a small Python package under `dist/`.
"""

from __future__ import annotations

from pathlib import Path
import argparse
import ast
import json
import re
from typing import Mapping

import openpyxl
import openpyxl.utils.cell
from openpyxl.worksheet.worksheet import Worksheet
from excel_grapher.exporter import CodeGenerator
from excel_grapher.grapher import get_calc_settings

from .lic_dsf_config import (
    ensure_workbook_available,
    validate_workbook_metadata,
    parse_range_spec,
    cells_in_range,
)
from .lic_dsf_labels import (
    detect_year_offset_headers,
    find_region_config,
    get_column_labels,
    get_row_labels,
    get_labels_from_region_config,
    is_offset_label,
    parse_offset_label,
)
from .lic_dsf_pipeline import (
    build_graph,
    classify_input_addresses,
    enrich_graph,
    export_enrichment_audit,
    iter_string_constant_addresses,
    populate_leaf_values,
)
from .lic_dsf_group_inputs import build_input_groups_payload, iter_input_cells
from .configs import available_templates, load_template_config


MAX_DEPTH = 50


def discover_targets_by_indicator_row(workbook: Path) -> dict[tuple[str, int], list[str]]:
    """
    Legacy helper retained for backwards compatibility; no longer used.

    Row-based discovery has been superseded by explicit range-based targets.
    """
    return {}


def normalize_entrypoint_name(name: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z]+", "_", name.strip()).strip("_").lower()
    if not cleaned:
        return "sheet"
    if not cleaned[0].isalpha():
        return f"sheet_{cleaned}"
    return cleaned


def load_enrichment_audit_labels(path: Path) -> dict[tuple[str, int], list[str]]:
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}

    labels_by_row: dict[tuple[str, int], list[str]] = {}
    by_sheet = payload.get("by_sheet", {})
    for sheet, sheet_payload in by_sheet.items():
        for cell in sheet_payload.get("cells", []):
            address = cell.get("address")
            if not isinstance(address, str):
                continue
            match = re.match(r"^[A-Z]+(\d+)$", address)
            if not match:
                continue
            row = int(match.group(1))
            row_labels = cell.get("row_labels") or []
            if not isinstance(row_labels, list):
                continue
            if row_labels:
                labels_by_row.setdefault((sheet, row), []).extend(
                    label for label in row_labels if isinstance(label, str) and label.strip()
                )
    return labels_by_row


def build_entrypoints(
    targets: list[str],
    audit_path: Path,
    export_ranges: list | None = None,
) -> dict[str, list[str]]:
    labels_by_row = load_enrichment_audit_labels(audit_path)
    entrypoints: dict[str, list[str]] = {}

    # Precompute per-cell targets from configured ranges.
    per_cell_targets: set[str] = set()
    for cfg in (export_ranges or []):
        if cfg.get("entrypoint_mode") != "per_cell":
            continue
        sheet_name, range_a1 = parse_range_spec(cfg["range_spec"])
        per_cell_targets.update(cells_in_range(sheet_name, range_a1))

    # Partition targets into row-grouped vs per-cell.
    targets_by_row: dict[tuple[str, int], list[str]] = {}
    per_cell_list: list[str] = []
    for addr in targets:
        if addr in per_cell_targets:
            per_cell_list.append(addr)
            continue
        m = re.match(r"^(.+)!([A-Z]+)(\d+)$", addr)
        if not m:
            continue
        sheet, _col, row_str = m.group(1), m.group(2), m.group(3)
        row = int(row_str)
        targets_by_row.setdefault((sheet, row), []).append(addr)

    # Row-grouped entrypoints (one per row).
    for (sheet, row), row_targets in targets_by_row.items():
        prefix_match = re.match(r"^([A-Za-z]\d+)", sheet.strip())
        sheet_prefix = normalize_entrypoint_name(prefix_match.group(1) if prefix_match else sheet)
        label = next(iter(labels_by_row.get((sheet, row), [])), "")
        base = normalize_entrypoint_name(label or f"{sheet} {row}")
        if base.startswith(f"{sheet_prefix}_") or base == sheet_prefix:
            name = base
        else:
            name = f"{sheet_prefix}_{base}"
        suffix = 2
        while name in entrypoints:
            name = f"{base}_{suffix}"
            suffix += 1
        entrypoints[name] = row_targets

    # Per-cell entrypoints (one per cell; avoid grouping D/I cells on same row).
    for addr in per_cell_list:
        m = re.match(r"^(.+)!([A-Z]+)(\d+)$", addr)
        if not m:
            continue
        sheet, col_letter, row_str = m.group(1), m.group(2), m.group(3)
        row = int(row_str)
        prefix_match = re.match(r"^([A-Za-z]\d+)", sheet.strip())
        sheet_prefix = normalize_entrypoint_name(prefix_match.group(1) if prefix_match else sheet)
        label = next(iter(labels_by_row.get((sheet, row), [])), "")
        base_row = normalize_entrypoint_name(label or f"{sheet} {row}")
        base = f"{base_row}_{col_letter.lower()}"
        if not base.startswith(f"{sheet_prefix}_") and base != sheet_prefix:
            name = f"{sheet_prefix}_{base}"
        else:
            name = base
        suffix = 2
        while name in entrypoints:
            name = f"{base}_{suffix}"
            suffix += 1
        entrypoints[name] = [addr]
    return entrypoints


_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z_][0-9A-Za-z_]*$")


def format_sheet_name(sheet: str) -> str:
    """
    Format a sheet name for sheet-qualified Excel addresses.

    Only quote sheet names when needed (e.g., spaces or punctuation), and escape
    embedded single quotes by doubling them.
    """
    if _SAFE_SHEET_NAME_RE.match(sheet):
        return sheet
    escaped = sheet.replace("'", "''")
    return f"'{escaped}'"


def format_address(sheet: str, a1: str) -> str:
    return f"{format_sheet_name(sheet)}!{a1}"


def parse_year_label(label: str) -> int | None:
    try:
        y = int(str(label).strip())
    except ValueError:
        return None
    return y if 1900 <= y <= 2100 else None


def _parse_year_or_offset(label: str) -> int | None:
    """Parse a label as either an offset label or an absolute year."""
    if is_offset_label(label):
        return parse_offset_label(label)
    return parse_year_label(label)


def year_for_column(
    ws: Worksheet,
    row: int,
    col: int,
    offset_maps: dict[int, dict[int, int]] | None = None,
) -> int | None:
    """
    Determine the year/offset label for a cell's column header context.

    When ``offset_maps`` are provided, offset labels (``"offset:N"``) are
    recognised in addition to absolute year labels.  Returns a single integer
    if exactly one year-like or offset label is found; otherwise ``None``.
    """
    cfg = find_region_config(ws.title, row, col)
    if cfg is not None:
        _row_labels, col_labels = get_labels_from_region_config(
            ws, row, col, cfg, offset_maps
        )
    else:
        col_labels = get_column_labels(ws, row, col)

    values: list[int] = []
    for lab in col_labels:
        v = _parse_year_or_offset(lab)
        if v is not None:
            values.append(v)

    values = list(dict.fromkeys(values))
    if len(values) != 1:
        return None
    return values[0]


def year_for_row(
    ws: Worksheet,
    row: int,
    col: int,
    offset_maps: dict[int, dict[int, int]] | None = None,
) -> int | None:
    """
    Determine the year/offset label for a cell's row label context.

    When ``offset_maps`` are provided, offset labels (``"offset:N"``) are
    recognised in addition to absolute year labels.  Returns a single integer
    if exactly one year-like or offset label is found; otherwise ``None``.
    """
    cfg = find_region_config(ws.title, row, col)
    if cfg is not None:
        row_labels, _col_labels = get_labels_from_region_config(
            ws, row, col, cfg, offset_maps
        )
    else:
        row_labels = get_row_labels(ws, row, col)

    values: list[int] = []
    for lab in row_labels:
        v = _parse_year_or_offset(lab)
        if v is not None:
            values.append(v)

    values = list(dict.fromkeys(values))
    if len(values) != 1:
        return None
    return values[0]


def load_input_groups(path: Path) -> list[dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    groups = payload.get("groups", [])
    if not isinstance(groups, list):
        return []
    return [g for g in groups if isinstance(g, dict)]


def iter_wide_year_series_groups(groups: list[dict]) -> list[dict]:
    out: list[dict] = []
    for g in groups:
        if g.get("mode") != "ignore_column_axis_years":
            continue
        shape = g.get("shape") or {}
        if not isinstance(shape, dict):
            continue
        if shape.get("rows") != 1:
            continue
        if not g.get("bounding_box"):
            continue
        if not g.get("range_a1"):
            continue
        out.append(g)
    return out


def generate_setter_method_name(sheet: str, row_labels_key: list[str], group_id: str) -> str:
    prefix = normalize_entrypoint_name(sheet)
    label = next((s for s in row_labels_key if isinstance(s, str) and s.strip()), "")
    base = normalize_entrypoint_name(label) if label else group_id.lower()
    return f"set_{prefix}_{base}"


def split_columns_by_year_presence(
    *,
    start_col: int,
    end_col: int,
    year_by_col: Mapping[int, int | None],
) -> list[tuple[str, int, int]]:
    """
    Split a contiguous column span into segments of year vs non-year columns.

    Returns segments as tuples: (kind, seg_start_col, seg_end_col), where kind is
    either "year" or "meta".
    """
    segments: list[tuple[str, int, int]] = []
    current_kind: str | None = None
    seg_start: int | None = None

    for col in range(start_col, end_col + 1):
        kind = "year" if year_by_col.get(col) is not None else "meta"
        if current_kind is None:
            current_kind = kind
            seg_start = col
            continue
        if kind == current_kind:
            continue
        # close previous segment
        if seg_start is None:
            raise RuntimeError("Missing segment start while splitting columns")
        segments.append((current_kind, seg_start, col - 1))
        current_kind = kind
        seg_start = col

    if current_kind is not None and seg_start is not None:
        segments.append((current_kind, seg_start, end_col))

    return segments


def split_rows_by_year_presence(
    *,
    start_row: int,
    end_row: int,
    year_by_row: Mapping[int, int | None],
) -> list[tuple[str, int, int]]:
    """
    Split a contiguous row span into segments of year vs non-year rows.

    Returns segments as tuples: (kind, seg_start_row, seg_end_row), where kind is
    either "year" or "meta".
    """
    segments: list[tuple[str, int, int]] = []
    current_kind: str | None = None
    seg_start: int | None = None

    for row in range(start_row, end_row + 1):
        kind = "year" if year_by_row.get(row) is not None else "meta"
        if current_kind is None:
            current_kind = kind
            seg_start = row
            continue
        if kind == current_kind:
            continue
        if seg_start is None:
            raise RuntimeError("Missing segment start while splitting rows")
        segments.append((current_kind, seg_start, row - 1))
        current_kind = kind
        seg_start = row

    if current_kind is not None and seg_start is not None:
        segments.append((current_kind, seg_start, end_row))

    return segments


def generate_setters_module(
    *,
    workbook: Path,
    groups: list[dict],
    wb_values: openpyxl.Workbook | None = None,
    wb_formulas: openpyxl.Workbook | None = None,
) -> str:
    """
    Generate a `setters.py` module for the exported package.

    Generates:
    - Wide-format year-series setters for eligible groups (mode=ignore_column_axis_years, 1-row range)
    - Tall-format year-series setters for eligible groups (mode=ignore_row_axis_years, sparse cell sets)
    - Scalar / 1D / 2D setters for non-year rectangular groups (mode=no_year_axis)

    For year-series groups that contain a mix of year and non-year columns,
    split the group at codegen time into year-series setters and metadata setters.
    """
    if not groups:
        return "from __future__ import annotations\n\n# No setters were generated.\n"

    wb = wb_values or openpyxl.load_workbook(workbook, data_only=True)
    wb_formulas = wb_formulas or openpyxl.load_workbook(workbook)
    try:
        # Precompute offset maps per (sheet, header_row)
        _offset_cache: dict[str, dict[int, dict[int, int]]] = {}

        def _get_offset_maps(sheet_name: str) -> dict[int, dict[int, int]]:
            if sheet_name in _offset_cache:
                return _offset_cache[sheet_name]
            cfg = find_region_config(sheet_name, 1, 1)
            maps: dict[int, dict[int, int]] = {}
            if cfg is not None:
                for hr in cfg.get("header_rows", []):
                    ws_f = wb_formulas[sheet_name] if sheet_name in wb_formulas.sheetnames else None
                    ws_v = wb[sheet_name] if sheet_name in wb.sheetnames else None
                    if ws_f and ws_v:
                        maps[hr] = detect_year_offset_headers(ws_f, ws_v, sheet_name, hr)
            _offset_cache[sheet_name] = maps
            return maps

        year_specs: list[dict] = []
        tall_year_specs: list[dict] = []
        range_specs: list[dict] = []
        used_names: set[str] = set()

        for g in groups:
            sheet = g.get("sheet")
            group_id = str(g.get("group_id", "group"))
            row_labels_key = g.get("row_labels_key") or []
            mode = g.get("mode")

            if not isinstance(sheet, str) or sheet not in wb.sheetnames:
                continue

            base_name = generate_setter_method_name(sheet, row_labels_key, group_id)

            def unique(name: str) -> str:
                n = name
                suffix = 2
                while n in used_names:
                    n = f"{name}_{suffix}"
                    suffix += 1
                used_names.add(n)
                return n

            ws = wb[sheet]
            sheet_offset_maps = _get_offset_maps(sheet)

            # Tall-format year series (year axis is the row axis).
            if mode == "ignore_row_axis_years":
                cells = g.get("cells") or []
                if not isinstance(cells, list) or not cells:
                    continue

                offset_to_addresses: dict[int, list[str]] = {}
                meta_addresses: list[str] = []

                for addr in cells:
                    if not isinstance(addr, str) or "!" not in addr:
                        continue
                    _sheet, a1 = addr.split("!", 1)
                    m = re.match(r"^([A-Z]+)(\d+)$", a1)
                    if not m:
                        continue
                    col_letter, row_str = m.group(1), m.group(2)
                    r = int(row_str)
                    c = openpyxl.utils.cell.column_index_from_string(col_letter)

                    y = year_for_row(ws, r, c, sheet_offset_maps)
                    if y is None:
                        meta_addresses.append(addr)
                        continue
                    offset_to_addresses.setdefault(int(y), []).append(addr)

                if offset_to_addresses:
                    offsets_sorted = sorted(offset_to_addresses.keys())
                    offset_to_addresses_sorted = {
                        o: tuple(sorted(offset_to_addresses[o])) for o in offsets_sorted
                    }
                    tall_year_specs.append(
                        {
                            "name": unique(f"{base_name}_by_year"),
                            "offsets": offsets_sorted,
                            "offset_to_addresses": offset_to_addresses_sorted,
                        }
                    )

                if meta_addresses:
                    range_specs.append(
                        {
                            "name": unique(f"{base_name}_meta"),
                            "shape": (1, len(meta_addresses)),
                            "addresses": sorted(meta_addresses),
                        }
                    )

                continue

            # Everything below requires rectangular metadata from the grouping step.
            bbox = g.get("bounding_box") or {}
            shape = g.get("shape") or {}
            if not isinstance(bbox, dict) or not isinstance(shape, dict):
                continue
            if not g.get("range_a1"):
                continue

            row = bbox.get("start_row")
            start_col = bbox.get("start_col")
            end_col = bbox.get("end_col")
            end_row = bbox.get("end_row")
            if not isinstance(row, int):
                continue
            if not isinstance(start_col, int):
                continue
            if not isinstance(end_col, int):
                continue
            if not isinstance(end_row, int):
                continue

            rows = shape.get("rows")
            cols = shape.get("cols")
            if not (isinstance(rows, int) and isinstance(cols, int)):
                continue
            rows = int(rows)
            cols = int(cols)

            # Wide-format year series
            if mode == "ignore_column_axis_years" and rows == 1 and cols >= 1 and row == end_row:
                offset_by_col: dict[int, int | None] = {}
                for col in range(start_col, end_col + 1):
                    offset_by_col[col] = year_for_column(ws, row, col, sheet_offset_maps)

                segments = split_columns_by_year_presence(
                    start_col=start_col, end_col=end_col, year_by_col=offset_by_col
                )

                year_seg_index = 0
                meta_seg_index = 0
                for kind, seg_start, seg_end in segments:
                    if kind == "year":
                        year_seg_index += 1
                        offset_to_address: dict[int, str] = {}
                        offsets: list[int] = []
                        ok = True
                        for col in range(seg_start, seg_end + 1):
                            o = offset_by_col.get(col)
                            if o is None:
                                ok = False
                                break
                            if o in offset_to_address:
                                ok = False
                                break
                            col_letter = openpyxl.utils.cell.get_column_letter(col)
                            addr = format_address(sheet, f"{col_letter}{row}")
                            offset_to_address[int(o)] = addr
                            offsets.append(int(o))
                        if not ok or not offset_to_address:
                            continue

                        name = base_name if len(segments) == 1 else f"{base_name}_years_{year_seg_index}"
                        year_specs.append(
                            {
                                "name": unique(name),
                                "offsets": offsets,
                                "offset_to_address": offset_to_address,
                            }
                        )
                    else:
                        meta_seg_index += 1
                        addresses: list[str] = []
                        for col in range(seg_start, seg_end + 1):
                            col_letter = openpyxl.utils.cell.get_column_letter(col)
                            addresses.append(format_address(sheet, f"{col_letter}{row}"))
                        if not addresses:
                            continue
                        meta_name = f"{base_name}_meta_{meta_seg_index}"
                        range_specs.append(
                            {
                                "name": unique(meta_name),
                                "shape": (1, len(addresses)),
                                "addresses": addresses,
                            }
                        )

                continue

            # Non-year rectangular setters (scalar/1D/2D)
            if mode == "no_year_axis":
                addresses: list[str] = []
                for r in range(row, row + rows):
                    for c in range(start_col, start_col + cols):
                        col_letter = openpyxl.utils.cell.get_column_letter(c)
                        addresses.append(format_address(sheet, f"{col_letter}{r}"))

                if not addresses:
                    continue

                range_specs.append(
                    {
                        "name": unique(base_name),
                        "shape": (rows, cols),
                        "addresses": addresses,
                    }
                )

        lines: list[str] = []
        lines.append("from __future__ import annotations")
        lines.append("")
        lines.append("from dataclasses import dataclass")
        lines.append("from collections.abc import Mapping as MappingABC, Sequence as SequenceABC")
        lines.append("from typing import Mapping, Sequence")
        lines.append("")
        lines.append("from .internals import CellValue, EvalContext")
        lines.append("from .inputs import DEFAULT_INPUTS")
        lines.append("")
        lines.append("")
        lines.append("BASE_YEAR_ADDRESS = \"'Input 1 - Basics'!C18\"")
        lines.append("")
        lines.append("")
        lines.append("RangeValue = CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]]")
        lines.append("")
        lines.append("")
        lines.append("@dataclass(frozen=True, slots=True)")
        lines.append("class YearSeriesAssignment:")
        lines.append("    offsets: tuple[int, ...]")
        lines.append("    applied: dict[int, str]")
        lines.append("    ignored: dict[int, CellValue]")
        lines.append("")
        lines.append("")
        lines.append("@dataclass(frozen=True, slots=True)")
        lines.append("class RangeAssignment:")
        lines.append("    shape: tuple[int, int]")
        lines.append("    addresses: tuple[str, ...]")
        lines.append("")
        lines.append("")
        lines.append("@dataclass(frozen=True, slots=True)")
        lines.append("class YearRowAssignment:")
        lines.append("    offsets: tuple[int, ...]")
        lines.append("    applied: dict[int, tuple[str, ...]]")
        lines.append("    ignored: dict[int, CellValue]")
        lines.append("")
        lines.append("")
        lines.append("def _split_sheet_address(address: str) -> tuple[str, str]:")
        lines.append("    if '!' not in address:")
        lines.append("        raise ValueError(f\"Invalid address: {address}\")")
        lines.append("    if address.startswith(\"'\"):")
        lines.append("        i = 1")
        lines.append("        sheet_chars: list[str] = []")
        lines.append("        while i < len(address):")
        lines.append("            ch = address[i]")
        lines.append("            if ch == \"'\":")
        lines.append("                if i + 1 < len(address) and address[i + 1] == \"'\":")
        lines.append("                    sheet_chars.append(\"'\")")
        lines.append("                    i += 2")
        lines.append("                    continue")
        lines.append("                if i + 1 < len(address) and address[i + 1] == \"!\":")
        lines.append("                    sheet = \"\".join(sheet_chars)")
        lines.append("                    a1 = address[i + 2 :]")
        lines.append("                    if not a1:")
        lines.append("                        raise ValueError(f\"Invalid address: {address}\")")
        lines.append("                    return sheet, a1")
        lines.append("                raise ValueError(f\"Invalid address: {address}\")")
        lines.append("            sheet_chars.append(ch)")
        lines.append("            i += 1")
        lines.append("        raise ValueError(f\"Invalid address: {address}\")")
        lines.append("    sheet, a1 = address.split(\"!\", 1)")
        lines.append("    if not sheet or not a1:")
        lines.append("        raise ValueError(f\"Invalid address: {address}\")")
        lines.append("    return sheet, a1")
        lines.append("")
        lines.append("")
        lines.append("def _get_base_year(ctx: EvalContext) -> int | None:")
        lines.append("    v = ctx.inputs.get(BASE_YEAR_ADDRESS)")
        lines.append("    if isinstance(v, (int, float)) and not isinstance(v, bool) and 1900 <= int(v) <= 2100:")
        lines.append("        return int(v)")
        lines.append("    return None")
        lines.append("")
        lines.append("")
        lines.append("def _resolve_key(key: int, base_year: int | None) -> int:")
        lines.append("    if 1900 <= key <= 2100:")
        lines.append("        if base_year is None:")
        lines.append("            raise ValueError(")
        lines.append("                f\"Cannot resolve year {key} to an offset without a base year. \"")
        lines.append("                \"Set the base year first via set_input_1_basics_first_year_of_projections().\"")
        lines.append("            )")
        lines.append("        return key - base_year")
        lines.append("    return key")
        lines.append("")
        lines.append("")
        lines.append("def _read_inputs_from_workbook(workbook_path: str) -> dict[str, CellValue]:")
        lines.append("    try:")
        lines.append("        import openpyxl")
        lines.append("    except ImportError as exc:")
        lines.append(
            "        raise ImportError(\"openpyxl is required to read inputs from a workbook\") from exc"
        )
        lines.append("    wb = openpyxl.load_workbook(workbook_path, data_only=True)")
        lines.append("    try:")
        lines.append("        updates: dict[str, CellValue] = {}")
        lines.append("        ws_cache: dict[str, object] = {}")
        lines.append("        for addr in DEFAULT_INPUTS.keys():")
        lines.append("            sheet_name, a1 = _split_sheet_address(str(addr))")
        lines.append("            if sheet_name not in wb.sheetnames:")
        lines.append(
            "                raise KeyError(f\"Workbook is missing sheet {sheet_name!r} for address {addr}\")"
        )
        lines.append("            ws = ws_cache.get(sheet_name)")
        lines.append("            if ws is None:")
        lines.append("                ws = wb[sheet_name]")
        lines.append("                ws_cache[sheet_name] = ws")
        lines.append("            value = ws[a1].value")
        lines.append("            updates[str(addr)] = 0 if value is None else value")
        lines.append("        return updates")
        lines.append("    finally:")
        lines.append("        wb.close()")
        lines.append("")
        lines.append("")
        lines.append("def _apply_range(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    shape: tuple[int, int],")
        lines.append("    addresses: Sequence[str],")
        lines.append("    values: RangeValue,")
        lines.append(") -> RangeAssignment:")
        lines.append("    rows, cols = shape")
        lines.append("    updates: dict[str, CellValue] = {}")
        lines.append("    flat: list[CellValue] = []")
        lines.append("    if rows == 1 and cols == 1:")
        lines.append("        flat = [values]  # scalar")
        lines.append("    elif rows == 1 or cols == 1:")
        lines.append("        if not isinstance(values, Sequence):")
        lines.append("            raise TypeError('Expected a sequence for 1D range')") 
        lines.append("        flat = list(values)")
        lines.append("        if len(flat) != len(addresses):")
        lines.append("            raise ValueError(f'Expected {len(addresses)} values, got {len(flat)}')")
        lines.append("    else:")
        lines.append("        if not isinstance(values, Sequence):")
        lines.append("            raise TypeError('Expected a sequence of sequences for 2D range')")
        lines.append("        rows_values = list(values)")
        lines.append("        if len(rows_values) != rows:")
        lines.append("            raise ValueError(f'Expected {rows} rows, got {len(rows_values)}')")
        lines.append("        for rv in rows_values:")
        lines.append("            if not isinstance(rv, Sequence):")
        lines.append("                raise TypeError('Expected a sequence of sequences for 2D range')")
        lines.append("            row_list = list(rv)")
        lines.append("            if len(row_list) != cols:")
        lines.append("                raise ValueError(f'Expected {cols} columns, got {len(row_list)}')")
        lines.append("            flat.extend(row_list)")
        lines.append("    if len(flat) != len(addresses):")
        lines.append("        raise ValueError(f'Expected {len(addresses)} values, got {len(flat)}')")
        lines.append("    for addr, value in zip(addresses, flat):")
        lines.append("        v = 0 if value is None else value")
        lines.append("        updates[str(addr)] = v")
        lines.append("    if updates:")
        lines.append("        ctx.set_inputs(updates)")
        lines.append("    return RangeAssignment(shape=shape, addresses=tuple(addresses))")
        lines.append("")
        lines.append("")
        lines.append("def _apply_year_row_mapping(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    offsets: tuple[int, ...],")
        lines.append("    offset_to_addresses: dict[int, tuple[str, ...]],")
        lines.append("    values_by_key: Mapping[int, CellValue],")
        lines.append("    strict: bool = True,")
        lines.append(") -> YearRowAssignment:")
        lines.append("    base_year = _get_base_year(ctx)")
        lines.append("    applied: dict[int, tuple[str, ...]] = {}")
        lines.append("    ignored: dict[int, CellValue] = {}")
        lines.append("    updates: dict[str, CellValue] = {}")
        lines.append("    for key, value in values_by_key.items():")
        lines.append("        offset = _resolve_key(int(key), base_year)")
        lines.append("        addrs = offset_to_addresses.get(offset)")
        lines.append("        if addrs is None:")
        lines.append("            if strict:")
        lines.append("                raise KeyError(f\"Key {key} (offset {offset}) is not in this table: {offsets}\")")
        lines.append("            ignored[int(key)] = value")
        lines.append("            continue")
        lines.append("        v = 0 if value is None else value")
        lines.append("        for addr in addrs:")
        lines.append("            updates[str(addr)] = v")
        lines.append("        applied[int(key)] = tuple(addrs)")
        lines.append("    if updates:")
        lines.append("        ctx.set_inputs(updates)")
        lines.append("    return YearRowAssignment(offsets=offsets, applied=applied, ignored=ignored)")
        lines.append("")
        lines.append("")
        lines.append("def _apply_year_row_array(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    offsets: tuple[int, ...],")
        lines.append("    offset_to_addresses: dict[int, tuple[str, ...]],")
        lines.append("    values: Sequence[CellValue],")
        lines.append("    start_key: int,")
        lines.append("    strict: bool = True,")
        lines.append(") -> YearRowAssignment:")
        lines.append("    base_year = _get_base_year(ctx)")
        lines.append("    start_offset = _resolve_key(start_key, base_year)")
        lines.append("    if start_offset not in offset_to_addresses:")
        lines.append("        raise KeyError(f\"start_year {start_key} (offset {start_offset}) is not in this table: {offsets}\")")
        lines.append("    offsets_list = list(offsets)")
        lines.append("    start_idx = offsets_list.index(start_offset)")
        lines.append("    remaining = offsets_list[start_idx:]")
        lines.append("    if len(values) > len(remaining):")
        lines.append("        raise ValueError(")
        lines.append("            f\"Too many values ({len(values)}) for table from offset {start_offset}; \"")
        lines.append("            f\"only {len(remaining)} offsets available\"")
        lines.append("        )")
        lines.append("    expected = list(range(start_offset, start_offset + len(remaining)))")
        lines.append("    if remaining != expected:")
        lines.append("        raise ValueError(")
        lines.append("            \"Non-contiguous offsets; array mapping is disallowed for this table. \"")
        lines.append("            \"Use dict-based mapping instead.\"")
        lines.append("        )")
        lines.append("    values_by_key = {start_offset + i: values[i] for i in range(len(values))}")
        lines.append("    return _apply_year_row_mapping(")
        lines.append("        ctx, offsets=offsets, offset_to_addresses=offset_to_addresses, values_by_key=values_by_key, strict=strict")
        lines.append("    )")
        lines.append("")
        lines.append("")
        lines.append("def _apply_year_series_mapping(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    offsets: tuple[int, ...],")
        lines.append("    offset_to_address: dict[int, str],")
        lines.append("    values_by_key: Mapping[int, CellValue],")
        lines.append("    strict: bool = True,")
        lines.append(") -> YearSeriesAssignment:")
        lines.append("    base_year = _get_base_year(ctx)")
        lines.append("    applied: dict[int, str] = {}")
        lines.append("    ignored: dict[int, CellValue] = {}")
        lines.append("    updates: dict[str, CellValue] = {}")
        lines.append("    for key, value in values_by_key.items():")
        lines.append("        offset = _resolve_key(int(key), base_year)")
        lines.append("        addr = offset_to_address.get(offset)")
        lines.append("        if addr is None:")
        lines.append("            if strict:")
        lines.append("                raise KeyError(f\"Key {key} (offset {offset}) is not in this series: {offsets}\")")
        lines.append("            ignored[int(key)] = value")
        lines.append("            continue")
        lines.append("        v = 0 if value is None else value")
        lines.append("        updates[addr] = v")
        lines.append("        applied[int(key)] = addr")
        lines.append("    if updates:")
        lines.append("        ctx.set_inputs(updates)")
        lines.append("    return YearSeriesAssignment(offsets=offsets, applied=applied, ignored=ignored)")
        lines.append("")
        lines.append("")
        lines.append("def _apply_year_series_array(")
        lines.append("    ctx: EvalContext,")
        lines.append("    *,")
        lines.append("    offsets: tuple[int, ...],")
        lines.append("    offset_to_address: dict[int, str],")
        lines.append("    values: Sequence[CellValue],")
        lines.append("    start_key: int,")
        lines.append("    strict: bool = True,")
        lines.append(") -> YearSeriesAssignment:")
        lines.append("    base_year = _get_base_year(ctx)")
        lines.append("    start_offset = _resolve_key(start_key, base_year)")
        lines.append("    if start_offset not in offset_to_address:")
        lines.append("        raise KeyError(f\"start_year {start_key} (offset {start_offset}) is not in this series: {offsets}\")")
        lines.append("    offsets_list = list(offsets)")
        lines.append("    start_idx = offsets_list.index(start_offset)")
        lines.append("    remaining = offsets_list[start_idx:]")
        lines.append("    if len(values) > len(remaining):")
        lines.append("        raise ValueError(")
        lines.append("            f\"Too many values ({len(values)}) for series from offset {start_offset}; \"")
        lines.append("            f\"only {len(remaining)} offsets available\"")
        lines.append("        )")
        lines.append("    expected = list(range(start_offset, start_offset + len(remaining)))")
        lines.append("    if remaining != expected:")
        lines.append("        raise ValueError(")
        lines.append("            \"Non-contiguous offsets; array mapping is disallowed for this series. \"")
        lines.append("            \"Use dict-based mapping instead.\"")
        lines.append("        )")
        lines.append("    values_by_key = {start_offset + i: values[i] for i in range(len(values))}")
        lines.append("    return _apply_year_series_mapping(")
        lines.append("        ctx, offsets=offsets, offset_to_address=offset_to_address, values_by_key=values_by_key, strict=strict")
        lines.append("    )")
        lines.append("")
        lines.append("")
        lines.append("class LicDsfContext(EvalContext):")
        lines.append("    __slots__ = ()")
        lines.append("")
        lines.append("    def load_inputs_from_workbook(self, workbook_path: str) -> dict[str, CellValue]:")
        lines.append("        updates = _read_inputs_from_workbook(workbook_path)")
        lines.append("        if updates:")
        lines.append("            self.set_inputs(updates)")
        lines.append("        return updates")
        lines.append("")

        for spec in year_specs:
            name = spec["name"]
            offsets = spec["offsets"]
            offset_to_address = spec["offset_to_address"]

            lines.append(f"    def {name}(")
            lines.append("        self,")
            lines.append("        values: Mapping[int, CellValue] | Sequence[CellValue],")
            lines.append("        *,")
            lines.append("        start_year: int | None = None,")
            lines.append("        strict: bool = True,")
            lines.append("    ) -> YearSeriesAssignment:")
            lines.append("        if isinstance(values, MappingABC):")
            lines.append("            return _apply_year_series_mapping(")
            lines.append(f"                self, offsets={tuple(offsets)!r}, offset_to_address={offset_to_address!r},")
            lines.append("                values_by_key=values, strict=strict,")
            lines.append("            )")
            lines.append("        if not isinstance(values, SequenceABC):")
            lines.append("            raise TypeError(\"Expected a mapping or sequence for year-series inputs\")")
            lines.append("        if start_year is None:")
            lines.append("            raise TypeError(\"start_year is required for sequence inputs\")")
            lines.append("        return _apply_year_series_array(")
            lines.append(f"            self, offsets={tuple(offsets)!r}, offset_to_address={offset_to_address!r},")
            lines.append("            values=values, start_key=start_year, strict=strict,")
            lines.append("        )")
            lines.append("")

        for spec in range_specs:
            name = spec["name"]
            shape = spec["shape"]
            addresses = spec["addresses"]
            lines.append(f"    def {name}(")
            lines.append("        self,")
            lines.append("        values: RangeValue,")
            lines.append("    ) -> RangeAssignment:")
            lines.append("        return _apply_range(")
            lines.append(f"            self, shape={tuple(shape)!r}, addresses={list(addresses)!r}, values=values")
            lines.append("        )")
            lines.append("")

        for spec in tall_year_specs:
            name = spec["name"]
            offsets = spec["offsets"]
            offset_to_addresses = spec["offset_to_addresses"]
            lines.append(f"    def {name}(")
            lines.append("        self,")
            lines.append("        values: Mapping[int, CellValue] | Sequence[CellValue],")
            lines.append("        *,")
            lines.append("        start_year: int | None = None,")
            lines.append("        strict: bool = True,")
            lines.append("    ) -> YearRowAssignment:")
            lines.append("        if isinstance(values, MappingABC):")
            lines.append("            return _apply_year_row_mapping(")
            lines.append(f"                self, offsets={tuple(offsets)!r}, offset_to_addresses={offset_to_addresses!r},")
            lines.append("                values_by_key=values, strict=strict,")
            lines.append("            )")
            lines.append("        if not isinstance(values, SequenceABC):")
            lines.append("            raise TypeError(\"Expected a mapping or sequence for year-row inputs\")")
            lines.append("        if start_year is None:")
            lines.append("            raise TypeError(\"start_year is required for sequence inputs\")")
            lines.append("        return _apply_year_row_array(")
            lines.append(f"            self, offsets={tuple(offsets)!r}, offset_to_addresses={offset_to_addresses!r},")
            lines.append("            values=values, start_key=start_year, strict=strict,")
            lines.append("        )")
            lines.append("")

        return "\n".join(lines).rstrip() + "\n"
    finally:
        if wb_values is None:
            wb.close()
        if wb_formulas is None:
            wb_formulas.close()


def patch_entrypoint_for_setters(entrypoint_py: str) -> str:
    if "from .setters import LicDsfContext" not in entrypoint_py:
        lines = entrypoint_py.splitlines()
        insert_at = None
        for idx, line in enumerate(lines):
            if line.startswith("from .internals import "):
                insert_at = idx + 1
                break
        if insert_at is not None:
            lines.insert(insert_at, "from .setters import LicDsfContext")
            entrypoint_py = "\n".join(lines) + ("\n" if entrypoint_py.endswith("\n") else "")
    entrypoint_py = entrypoint_py.replace(
        "return EvalContext(inputs=merged, resolver=_resolve_formula)",
        "return LicDsfContext(inputs=merged, resolver=_resolve_formula)",
    )
    return entrypoint_py


def patch_init_for_setters(init_py: str) -> str:
    # Ensure the public symbols are imported.
    if "from .setters import LicDsfContext" not in init_py:
        init_py = init_py.replace(
            "from .inputs import DEFAULT_INPUTS  # noqa: F401\n",
            "from .inputs import DEFAULT_INPUTS  # noqa: F401\n"
            "from .setters import LicDsfContext, YearSeriesAssignment, RangeAssignment, YearRowAssignment  # noqa: F401\n",
        )
    else:
        # Expand existing import line to include missing symbols.
        if "YearRowAssignment" not in init_py:
            init_py = init_py.replace(
                "from .setters import LicDsfContext, YearSeriesAssignment, RangeAssignment  # noqa: F401\n",
                "from .setters import LicDsfContext, YearSeriesAssignment, RangeAssignment, YearRowAssignment  # noqa: F401\n",
            )
        if "RangeAssignment" not in init_py:
            init_py = init_py.replace(
                "from .setters import LicDsfContext, YearSeriesAssignment  # noqa: F401\n",
                "from .setters import LicDsfContext, YearSeriesAssignment, RangeAssignment, YearRowAssignment  # noqa: F401\n",
            )

    # Ensure __all__ includes these names even if they are already imported.
    names_to_add = ["LicDsfContext", "YearSeriesAssignment", "RangeAssignment", "YearRowAssignment"]
    for line in init_py.splitlines():
        if line.strip().startswith("__all__"):
            try:
                _, rhs = line.split("=", 1)
                current = ast.literal_eval(rhs.strip())
            except Exception:
                current = None
            if isinstance(current, list):
                updated = list(current)
                for n in names_to_add:
                    if n not in updated:
                        updated.append(n)
                new_line = f"__all__ = {updated!r}"
                init_py = init_py.replace(line, new_line)
            break

    return init_py


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Export LIC-DSF workbook formulas (with optional audit-only mode)."
    )
    templates = available_templates()
    parser.add_argument(
        "--template",
        type=str,
        required=True,
        choices=templates,
        help=f"Template version to use (available: {', '.join(templates)})",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help="Override path to workbook (default: from template config)",
    )
    parser.add_argument(
        "--workbook-url",
        type=str,
        default=None,
        help="Override download URL for the workbook",
    )
    parser.add_argument("--max-depth", type=int, default=MAX_DEPTH, help="Dependency graph max depth")
    parser.add_argument(
        "--audit-path",
        type=Path,
        default=None,
        help="Path for enrichment audit JSON output (default: in config dir)",
    )
    parser.add_argument(
        "--audit-only",
        action="store_true",
        help="Generate enrichment audit only; skip export generation",
    )
    parser.add_argument(
        "--input-groups-path",
        type=Path,
        default=None,
        help="Path to input groups JSON for setter generation (default: in config dir)",
    )
    parser.add_argument(
        "--input-groups-audit-path",
        type=Path,
        default=None,
        help="Path for input groups JSON output in audit-only mode (default: in config dir)",
    )
    parser.add_argument(
        "--export-dir",
        type=Path,
        default=None,
        help="Directory to write exported package (default: from template config)",
    )
    parser.add_argument(
        "--package-name",
        type=str,
        default=None,
        help="Package name for generated modules (default: from template config)",
    )
    return parser


def main(argv: list[str] | None = None) -> None:
    args = build_arg_parser().parse_args(argv)

    # Load template-specific configuration
    cfg = load_template_config(args.template)
    config_dir = Path(__file__).parent / "configs" / args.template

    workbook = args.workbook or cfg.WORKBOOK_PATH
    workbook_url = args.workbook_url or getattr(cfg, "WORKBOOK_TEMPLATE_URL", None)
    audit_path = args.audit_path or (config_dir / "enrichment_audit.json")
    input_groups_path = args.input_groups_path or (config_dir / "input_groups.json")
    input_groups_audit_path = args.input_groups_audit_path or (config_dir / "input_groups.json")
    export_dir = args.export_dir or cfg.EXPORT_DIR
    package_name = args.package_name or cfg.PACKAGE_NAME

    if not ensure_workbook_available(workbook, workbook_url):
        if not workbook.exists():
            print(f"Error: Workbook not available at {workbook}")
            return

    expected_meta = getattr(cfg, "WORKBOOK_METADATA", None)
    if expected_meta:
        validate_workbook_metadata(workbook, expected_meta)

    import time as _time

    from .lic_dsf_pipeline import discover_targets

    _t0 = _time.monotonic()
    targets = discover_targets(cfg.EXPORT_RANGES)
    print(f"[TIMING] discover_targets: {_time.monotonic() - _t0:.2f}s")
    if not targets:
        print("No targets found. Nothing to export.")
        return

    print("=" * 70)
    print(f"LIC-DSF Workbook Export (template: {args.template})")
    print("=" * 70)
    print(f"Workbook: {workbook}")
    print(f"Target cells: {len(targets)}")

    dynamic_refs = cfg.get_dynamic_ref_config()
    region_config = cfg.REGION_CONFIG
    string_constant_excludes = cfg.STRING_CONSTANT_EXCLUDES
    blank_constant_excludes = cfg.BLANK_CONSTANT_EXCLUDES

    wb_values: openpyxl.Workbook | None = None
    wb_formulas: openpyxl.Workbook | None = None
    keep_vba = workbook.suffix.lower() == ".xlsm"
    try:
        wb_formulas = openpyxl.load_workbook(workbook, data_only=False, keep_vba=keep_vba)
        wb_values = openpyxl.load_workbook(workbook, data_only=True, keep_vba=keep_vba)

        print("\nBuilding dependency graph...")
        _t0 = _time.monotonic()
        graph = build_graph(
            workbook,
            targets,
            max_depth=args.max_depth,
            wb_formulas=wb_formulas,
            dynamic_refs=dynamic_refs,
        )
        print(f"[TIMING] build_graph: {_time.monotonic() - _t0:.2f}s")

        print(f"   Nodes in graph: {len(graph)}")
        print(f"   Leaf nodes: {sum(1 for _ in graph.leaves())}")
        print(f"   Formula nodes: {len(graph) - sum(1 for _ in graph.leaves())}")

        sheets: dict[str, int] = {}
        for key in graph:
            node = graph.get_node(key)
            if node:
                sheets[node.sheet] = sheets.get(node.sheet, 0) + 1
        print("\n   Nodes by sheet:")
        for sheet_name in sorted(sheets.keys()):
            print(f"      {sheet_name}: {sheets[sheet_name]}")

        print("\nPopulating leaf values from cached workbook values...")
        _t0 = _time.monotonic()
        populate_leaf_values(graph, workbook, wb_values=wb_values)
        print(f"[TIMING] populate_leaf_values: {_time.monotonic() - _t0:.2f}s")

        constant_ranges = iter_string_constant_addresses(graph, string_constant_excludes)
        input_addresses = classify_input_addresses(
            graph,
            targets,
            constant_ranges=constant_ranges,
            constant_blanks=True,
            blank_excludes=blank_constant_excludes,
            attach_to_graph=True,
        )

        print("\nEnriching nodes with row/column labels...")
        _t0 = _time.monotonic()
        enrichment_results = enrich_graph(
            graph,
            workbook,
            wb_values=wb_values,
            wb_formulas=wb_formulas,
            region_config=region_config,
        )
        print(f"[TIMING] enrich_graph: {_time.monotonic() - _t0:.2f}s")
        total_nodes = len(enrichment_results)
        nodes_with_row_labels = sum(1 for r in enrichment_results.values() if r["row_labels"])
        nodes_with_col_labels = sum(1 for r in enrichment_results.values() if r["column_labels"])
        nodes_with_any_label = sum(
            1 for r in enrichment_results.values() if r["row_labels"] or r["column_labels"]
        )
        nodes_without_labels = total_nodes - nodes_with_any_label
        print(f"   Nodes with any label: {nodes_with_any_label}")
        print(f"   Nodes with row labels: {nodes_with_row_labels}")
        print(f"   Nodes with column labels: {nodes_with_col_labels}")
        print(f"   Nodes without labels: {nodes_without_labels}")

        input_with_row = 0
        input_with_col = 0
        input_with_any = 0
        input_without = 0
        for data in enrichment_results.values():
            sheet = data.get("sheet")
            a1 = data.get("address")
            if not isinstance(sheet, str) or not isinstance(a1, str):
                continue
            address = format_address(sheet, a1)
            if address not in input_addresses:
                continue
            has_row = bool(data["row_labels"])
            has_col = bool(data["column_labels"])
            if has_row:
                input_with_row += 1
            if has_col:
                input_with_col += 1
            if has_row or has_col:
                input_with_any += 1
            else:
                input_without += 1

        print("\n   Input nodes (constants excluded):")
        print(f"      Inputs with any label: {input_with_any}")
        print(f"      Inputs with row labels: {input_with_row}")
        print(f"      Inputs with column labels: {input_with_col}")
        print(f"      Inputs without labels: {input_without}")

        print("\n   Sample enriched nodes:")
        sample_count = 0
        for key, data in enrichment_results.items():
            if sample_count >= 5:
                break
            row_labels = data.get("row_labels")
            col_labels = data.get("column_labels")
            if not isinstance(row_labels, list) or not isinstance(col_labels, list):
                continue
            if not row_labels and not col_labels:
                continue
            row_str = ", ".join(str(label) for label in row_labels) if row_labels else "(none)"
            col_str = ", ".join(str(label) for label in col_labels) if col_labels else "(none)"
            print(f"      {key}:")
            print(f"         Row labels: {row_str}")
            print(f"         Col labels: {col_str}")
            sample_count += 1

        print(f"\nExporting audit file to {audit_path}...")
        export_enrichment_audit(graph, enrichment_results, audit_path)
        print(f"   Done. Review {audit_path} for sheet-by-sheet details.")

        print(f"\nExporting input groups to {input_groups_audit_path}...")
        input_cells = iter_input_cells(graph, enrichment_results)
        input_cells = [c for c in input_cells if c.address in input_addresses]
        input_groups_payload = build_input_groups_payload(
            targets=targets,
            graph=graph,
            input_cells=input_cells,
            restricted_to_export_default_inputs=False,
            export_default_inputs_count=None,
            workbook_path=str(workbook),
        )
        input_groups_audit_path.write_text(
            json.dumps(input_groups_payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        print(f"   Done. Review {input_groups_audit_path} for group details.")

        print("\nWorkbook calculation settings...")
        settings = get_calc_settings(workbook)
        print(f"   Iterate enabled: {settings.iterate_enabled}")
        print(f"   Iterate count:   {settings.iterate_count}")
        print(f"   Iterate delta:   {settings.iterate_delta}")

        if args.audit_only:
            return

        entrypoints = build_entrypoints(targets, audit_path=audit_path, export_ranges=cfg.EXPORT_RANGES)
        print(f"Entrypoints: {len(entrypoints)}")

        print("Generating Python package...")
        _t0 = _time.monotonic()
        generator = CodeGenerator(graph)
        modules = generator.generate_modules(
            targets,
            package_name=package_name,
            entrypoints=entrypoints if entrypoints else None,
        )
        print(f"[TIMING] generate_modules: {_time.monotonic() - _t0:.2f}s")

        # Generate year-aware input setters from canonical input groups.
        if input_groups_path.exists():
            _t0 = _time.monotonic()
            groups = load_input_groups(input_groups_path)
            setters_py = generate_setters_module(
                workbook=workbook,
                groups=groups,
                wb_values=wb_values,
                wb_formulas=wb_formulas,
            )
            print(f"[TIMING] generate_setters_module: {_time.monotonic() - _t0:.2f}s")
            modules[f"{package_name}/setters.py"] = setters_py
            entrypoint_path = f"{package_name}/entrypoint.py"
            init_path = f"{package_name}/__init__.py"
            if entrypoint_path in modules:
                modules[entrypoint_path] = patch_entrypoint_for_setters(modules[entrypoint_path])
            if init_path in modules:
                modules[init_path] = patch_init_for_setters(modules[init_path])
        else:
            print(
                f"Warning: {input_groups_path} not found; skipping input setter generation."
            )

        export_dir.mkdir(parents=True, exist_ok=True)
        for rel, content in modules.items():
            dst = export_dir / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            dst.write_text(content, encoding="utf-8")

        pkg_prefix = Path(next(iter(modules.keys()))).parts[0] if modules else "(unknown)"
        print(f"\nWrote {len(modules)} files under {export_dir / pkg_prefix}")
    finally:
        if wb_values is not None:
            wb_values.close()
        if wb_formulas is not None:
            wb_formulas.close()


if __name__ == "__main__":
    main()
