from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Mapping, Protocol, Sequence, TypeAlias

import openpyxl
import openpyxl.utils.cell
from openpyxl.worksheet.worksheet import Worksheet

from .lic_dsf_labels import (
    RegionConfig,
    detect_year_offset_headers,
    find_region_config,
    get_column_labels,
    get_labels_from_region_config,
    is_offset_label,
    parse_offset_label,
)

BASE_YEAR_ADDRESS = "'Input 1 - Basics'!C18"

CellValue: TypeAlias = Any


class InputsContext(Protocol):
    @property
    def inputs(self) -> dict[str, CellValue]: ...
    def set_inputs(self, inputs: dict[str, CellValue]) -> None: ...


@dataclass(frozen=True, slots=True)
class YearSeriesAssignment:
    """
    Transparent record of how a year-series setter mapped values to addresses.

    - `offsets` is the canonical offset ordering (left-to-right for wide series).
    - `applied` maps each user-provided key to the address that was updated.
    - `ignored` records user-provided keys that were dropped under strict=False.
    """

    offsets: tuple[int, ...]
    applied: dict[int, str]
    ignored: dict[int, CellValue]


def _parse_year(label: str) -> int | None:
    try:
        y = int(str(label).strip())
    except ValueError:
        return None
    return y if 1900 <= y <= 2100 else None


def _parse_year_or_offset(label: str) -> int | None:
    if is_offset_label(label):
        return parse_offset_label(label)
    return _parse_year(label)


def _dedupe_preserve_order(values: Sequence[int]) -> list[int]:
    seen: set[int] = set()
    out: list[int] = []
    for v in values:
        if v in seen:
            continue
        seen.add(v)
        out.append(v)
    return out


def _year_labels_for_cell(
    ws: Worksheet,
    row: int,
    col: int,
    offset_maps: dict[int, dict[int, int]] | None = None,
    region_config: list[RegionConfig] | None = None,
) -> list[int]:
    cfg = find_region_config(ws.title, row, col, region_config)
    if cfg is not None:
        _row_labels, col_labels = get_labels_from_region_config(ws, row, col, cfg, offset_maps)
    else:
        col_labels = get_column_labels(ws, row, col)

    values: list[int] = []
    for lab in col_labels:
        v = _parse_year_or_offset(lab)
        if v is not None:
            values.append(v)

    return _dedupe_preserve_order(values)


def _get_base_year(ctx: InputsContext) -> int | None:
    v = ctx.inputs.get(BASE_YEAR_ADDRESS)
    if isinstance(v, (int, float)) and not isinstance(v, bool) and 1900 <= int(v) <= 2100:
        return int(v)
    return None


def _resolve_key(key: int, base_year: int | None) -> int:
    if 1900 <= key <= 2100:
        if base_year is None:
            raise ValueError(
                f"Cannot resolve year {key} to an offset without a base year. "
                "Set the base year first via set_input_1_basics_first_year_of_projections()."
            )
        return key - base_year
    return key


@dataclass(frozen=True, slots=True)
class WideYearSeriesSpec:
    """
    Specification for a 1-row, multi-column year series (wide format).

    `offset_to_address` maps year-offset integers to cell addresses.
    """

    offsets: tuple[int, ...]
    offset_to_address: dict[int, str]


def build_wide_year_series_spec(
    *,
    workbook_path: str,
    sheet: str,
    row: int,
    start_col: int,
    end_col: int,
    region_config: list[RegionConfig] | None = None,
) -> WideYearSeriesSpec:
    """
    Build an offset->address mapping for a 1-row wide range.

    Uses formula-based offset detection and falls back to cached-value
    label extraction when offset detection doesn't apply.

    Pass the template's ``REGION_CONFIG`` so column/row headers resolve the same
    way as in export-time setter generation.
    """
    from pathlib import Path

    wb_path = Path(workbook_path)
    wb_values = openpyxl.load_workbook(wb_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(wb_path)
    try:
        ws_v = wb_values[sheet]

        offset_maps: dict[int, dict[int, int]] = {}
        cfg = find_region_config(sheet, row, start_col, region_config)
        if cfg is not None:
            ws_f = wb_formulas[sheet]
            for hr in cfg.get("header_rows", []):
                offset_maps[hr] = detect_year_offset_headers(ws_f, ws_v, sheet, hr)

        offset_to_address: dict[int, str] = {}
        ordered_offsets: list[int] = []

        for col in range(start_col, end_col + 1):
            values = _year_labels_for_cell(
                ws_v, row, col, offset_maps, region_config=region_config
            )
            if len(values) != 1:
                col_letter = openpyxl.utils.cell.get_column_letter(col)
                raise ValueError(
                    f"Ambiguous year label for {sheet}!{col_letter}{row}: {values}"
                )

            o = values[0]
            if o in offset_to_address:
                col_letter = openpyxl.utils.cell.get_column_letter(col)
                raise ValueError(
                    f"Duplicate offset {o} in series (also mapped earlier); at {sheet}!{col_letter}{row}"
                )

            addr = f"{sheet}!{openpyxl.utils.cell.get_column_letter(col)}{row}"
            offset_to_address[o] = addr
            ordered_offsets.append(o)

        return WideYearSeriesSpec(
            offsets=tuple(ordered_offsets),
            offset_to_address=offset_to_address,
        )
    finally:
        wb_values.close()
        wb_formulas.close()


def apply_year_series_mapping(
    ctx: InputsContext,
    spec: WideYearSeriesSpec,
    values_by_key: Mapping[int, CellValue],
    *,
    strict: bool = True,
) -> YearSeriesAssignment:
    """
    Apply a key->value mapping to the provided context.

    Keys may be year-like integers (auto-resolved to offsets via the base year
    in the context) or raw offset integers.
    """
    base_year = _get_base_year(ctx)
    applied: dict[int, str] = {}
    ignored: dict[int, CellValue] = {}
    updates: dict[str, CellValue] = {}

    for key, value in values_by_key.items():
        offset = _resolve_key(int(key), base_year)
        addr = spec.offset_to_address.get(offset)
        if addr is None:
            if strict:
                raise KeyError(f"Key {key} (offset {offset}) is not in this series: {spec.offsets}")
            ignored[int(key)] = value
            continue

        v = 0 if value is None else value
        updates[addr] = v
        applied[int(key)] = addr

    if updates:
        ctx.set_inputs(updates)

    return YearSeriesAssignment(offsets=spec.offsets, applied=applied, ignored=ignored)


def apply_year_series_array(
    ctx: InputsContext,
    spec: WideYearSeriesSpec,
    values: Sequence[CellValue],
    *,
    start_year: int,
    strict: bool = True,
) -> YearSeriesAssignment:
    """
    Apply an array to the series, aligned by `start_year`.

    `start_year` may be a year-like integer or a raw offset.
    """
    base_year = _get_base_year(ctx)
    start_offset = _resolve_key(start_year, base_year)

    if start_offset not in spec.offset_to_address:
        raise KeyError(f"start_year {start_year} (offset {start_offset}) is not in this series: {spec.offsets}")

    offsets_list = list(spec.offsets)
    start_idx = offsets_list.index(start_offset)
    remaining = offsets_list[start_idx:]

    if len(values) > len(remaining):
        raise ValueError(
            f"Too many values ({len(values)}) for series from offset {start_offset}; "
            f"only {len(remaining)} offsets available"
        )

    expected = list(range(start_offset, start_offset + len(remaining)))
    if remaining != expected:
        raise ValueError(
            "Non-contiguous offsets; array mapping is disallowed for this series. "
            "Use dict-based mapping instead."
        )

    values_by_key = {start_offset + i: values[i] for i in range(len(values))}
    return apply_year_series_mapping(ctx, spec, values_by_key, strict=strict)
