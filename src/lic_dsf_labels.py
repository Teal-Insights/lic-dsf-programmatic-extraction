#!/usr/bin/env python3
"""
Label extraction and workbook configuration helpers for LIC-DSF workflows.
"""

from __future__ import annotations

import re as _re
import json
from collections import defaultdict
from pathlib import Path
from typing import Any, Literal, TypedDict

import openpyxl
import openpyxl.utils.cell
from excel_grapher.grapher import DependencyGraph
from openpyxl.worksheet.worksheet import Worksheet


# Region-based label configuration
class RegionConfig(TypedDict, total=False):
    """
    Configuration for label extraction in a specific region of a sheet.

    Attributes:
        sheet: The sheet name this config applies to
        min_row: Minimum row of the region (inclusive, 1-indexed). None = no min.
        max_row: Maximum row of the region (inclusive, 1-indexed). None = no max.
        min_col: Minimum column of the region (inclusive, e.g., "A"). None = no min.
        max_col: Maximum column of the region (inclusive, e.g., "Z"). None = no max.
        header_rows: List of row numbers that contain column headers (1-indexed)
        label_columns: List of column letters that contain row labels
        annotation_axis: Axis for deduplicating annotations - "row" for wide-format
            time series (one annotation per row), "column" for columnar time series,
            or "cell" for individual cell annotations. Default: auto-detect.
    """

    sheet: str
    min_row: int | None
    max_row: int | None
    min_col: str | None
    max_col: str | None
    header_rows: list[int]
    label_columns: list[str]
    annotation_axis: Literal["row", "column", "cell"]



# Excel error values to filter out
EXCEL_ERRORS = frozenset(
    {
        "#DIV/0!",
        "#REF!",
        "#NAME?",
        "#VALUE!",
        "#N/A",
        "#NULL!",
        "#NUM!",
        "#GETTING_DATA",
        "#SPILL!",
        "#CALC!",
    }
)

# Placeholder patterns to filter out (exact matches after stripping)
PLACEHOLDER_PATTERNS = frozenset(
    {
        "...",
        "…",  # Unicode ellipsis
        "---",
        "--",
        "-",
        "n/a",
        "N/A",
        "n.a.",
        "N.A.",
        "TBD",
        "tbd",
    }
)


def get_effective_indent(cell) -> int:
    """
    Compute an effective indentation level for a cell.

    Combines ``cell.alignment.indent`` (Excel's formatting indent) with
    leading whitespace in the cell value (a secondary visual indent cue
    used in some LIC-DSF templates).
    """
    alignment_indent = (
        int(cell.alignment.indent)
        if cell.alignment and cell.alignment.indent
        else 0
    )
    text_indent = 0
    if isinstance(cell.value, str):
        stripped = cell.value.lstrip()
        leading_spaces = len(cell.value) - len(stripped)
        if leading_spaces > 0:
            text_indent = 1
    return alignment_indent + text_indent


def build_label_hierarchy(
    ws: Worksheet,
    label_col_idx: int,
    min_row: int | None = None,
    max_row: int | None = None,
) -> dict[int, list[str]]:
    """
    Build a mapping of ``row → ancestor labels`` for a label column.

    Scans the label column top-to-bottom, using effective indentation to
    determine parent–child relationships.  For each row that contains a
    valid label, the returned list contains its *ancestor* labels ordered
    from outermost to innermost (the row's own label is **not** included).

    Rows whose label column is empty are skipped (they don't break the
    hierarchy).
    """
    start = min_row or 1
    end = max_row or (ws.max_row or 1)

    # Stack entries: (effective_indent, label_text)
    stack: list[tuple[int, str]] = []
    hierarchy: dict[int, list[str]] = {}

    for row in range(start, end + 1):
        cell = ws.cell(row=row, column=label_col_idx)
        if cell.value is None:
            continue
        if not isinstance(cell.value, str):
            continue
        text = cell.value.strip()
        if not text or not is_valid_label(text):
            continue

        indent = get_effective_indent(cell)

        # Pop stack entries that are at the same indent level or deeper —
        # they are siblings or children of the previous context, not parents
        # of the current row.
        while stack and stack[-1][0] >= indent:
            stack.pop()

        # The remaining stack entries are the ancestors of this row.
        hierarchy[row] = [label for _, label in stack]

        # Push the current row onto the stack so it can be a parent of
        # subsequent deeper-indented rows.
        stack.append((indent, text))

    return hierarchy


def is_valid_label(text: str) -> bool:
    """
    Check if a text string is a valid label (not an error or placeholder).
    """
    stripped = text.strip()

    # Filter out empty strings
    if not stripped:
        return False

    # Filter out Excel errors
    if stripped in EXCEL_ERRORS:
        return False

    # Filter out placeholder patterns
    if stripped in PLACEHOLDER_PATTERNS:
        return False

    # Filter out strings that are just repeated punctuation (like "....." or "----")
    if stripped and all(c in ".-…" for c in stripped):
        return False

    return True


# Patterns that identify a header cell as the ProjectionYear anchor (offset 0).
_ANCHOR_PATTERNS: list[_re.Pattern[str]] = [
    _re.compile(r"^=\+?ProjectionYear$", _re.IGNORECASE),
    _re.compile(r"^=\+?'?Macro-Debt_Data'?!U[45]$"),
    _re.compile(r"^=\+?U4$"),
]

# =<COL><ROW>+1  or  =+<COL><ROW>+1
_PLUS_ONE_RE = _re.compile(r"^=\+?([A-Z]{1,3})(\d+)\+1$")
# =<COL><ROW>-1  or  =+<COL><ROW>-1
_MINUS_ONE_RE = _re.compile(r"^=\+?([A-Z]{1,3})(\d+)-1$")
# =+<COL><OTHER_ROW>  (same column, different row — row-copy pattern)
_ROW_COPY_RE = _re.compile(r"^=\+?([A-Z]{1,3})(\d+)$")


def _is_anchor_formula(formula: str) -> bool:
    return any(p.match(formula) for p in _ANCHOR_PATTERNS)


def detect_year_offset_headers(
    ws_formulas: Worksheet,
    ws_values: Worksheet,
    sheet: str,
    header_row: int,
) -> dict[int, int]:
    """
    Detect year-offset header cells by formula pattern analysis.

    Scans a header row for cells whose formulas reference ProjectionYear
    (directly or via +1/-1 chains) and returns a mapping of column index
    to integer offset (0 = projection year, positive = future, negative = past).
    """
    # Gather columns with populated cells for this row.
    # We use internal _cells if available for a massive performance win on wide sheets.
    # If the row is empty in _cells, fall back to scanning up to max_column to preserve
    # behavior for intentionally blank header rows.
    if hasattr(ws_formulas, "_cells") or hasattr(ws_values, "_cells"):
        relevant_cols: set[int] = set()
        if hasattr(ws_formulas, "_cells"):
            relevant_cols.update(c for r, c in ws_formulas._cells if r == header_row)
        if hasattr(ws_values, "_cells"):
            relevant_cols.update(c for r, c in ws_values._cells if r == header_row)
        if not relevant_cols:
            max_col = max(ws_formulas.max_column or 1, ws_values.max_column or 1)
            relevant_cols = set(range(1, max_col + 1))
    else:
        max_col = ws_formulas.max_column or 1
        relevant_cols = set(range(1, max_col + 1))

    formulas: dict[int, str] = {}
    values: dict[int, int] = {}
    for col in sorted(relevant_cols):
        f = ws_formulas.cell(row=header_row, column=col).value
        v = ws_values.cell(row=header_row, column=col).value
        if isinstance(f, str) and f.startswith("="):
            formulas[col] = f
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            values[col] = int(v)

    offsets: dict[int, int] = {}

    # Pass 1: find anchor cells (offset 0)
    for col, f in formulas.items():
        if _is_anchor_formula(f):
            offsets[col] = values.get(col, 0)

    # Pass 2: walk +1/-1 chains outward from known offsets.
    # Repeat until no new cells are resolved.
    changed = True
    while changed:
        changed = False
        for col, f in formulas.items():
            if col in offsets:
                continue
            m = _PLUS_ONE_RE.match(f)
            if m:
                ref_col_letter, ref_row_str = m.group(1), m.group(2)
                if int(ref_row_str) == header_row:
                    ref_col = openpyxl.utils.cell.column_index_from_string(
                        ref_col_letter
                    )
                    if ref_col in offsets:
                        offsets[col] = offsets[ref_col] + 1
                        changed = True
                        continue
            m = _MINUS_ONE_RE.match(f)
            if m:
                ref_col_letter, ref_row_str = m.group(1), m.group(2)
                if int(ref_row_str) == header_row:
                    ref_col = openpyxl.utils.cell.column_index_from_string(
                        ref_col_letter
                    )
                    if ref_col in offsets:
                        offsets[col] = offsets[ref_col] - 1
                        changed = True
                        continue

    # Pass 3: neighbor-consistency propagation. For any unresolved formula
    # cell with an integer cached value, if a resolved neighbor (col ± 1)
    # has an offset differing by exactly 1, adopt the cached value.
    # This handles cross-sheet references (e.g. ='Macro-Debt_Data'!T5)
    # and row-copy formulas (e.g. =+E50) without needing to pattern-match
    # every formula variant.
    if offsets:
        changed = True
        while changed:
            changed = False
            for col in sorted(formulas.keys()):
                if col in offsets or col not in values:
                    continue
                v = values[col]
                left = offsets.get(col - 1)
                right = offsets.get(col + 1)
                if left is not None and v == left + 1:
                    offsets[col] = v
                    changed = True
                elif right is not None and v == right - 1:
                    offsets[col] = v
                    changed = True
    else:
        # No anchors found — check if ALL formula cells are row-copy
        # references with integer cached values forming a contiguous sequence.
        row_copy_candidates: dict[int, int] = {}
        all_row_copy = True
        for col, f in formulas.items():
            m = _ROW_COPY_RE.match(f)
            if m:
                ref_col_letter, ref_row_str = m.group(1), m.group(2)
                ref_col = openpyxl.utils.cell.column_index_from_string(ref_col_letter)
                ref_row = int(ref_row_str)
                if ref_row != header_row and ref_col == col and col in values:
                    row_copy_candidates[col] = values[col]
                    continue
            all_row_copy = False

        if all_row_copy and row_copy_candidates:
            sorted_cols = sorted(row_copy_candidates.keys())
            vals = [row_copy_candidates[c] for c in sorted_cols]
            if all(vals[i + 1] - vals[i] == 1 for i in range(len(vals) - 1)):
                offsets = row_copy_candidates

    return offsets


_OFFSET_PREFIX = "offset:"


def is_offset_label(label: str) -> bool:
    """Return True if ``label`` is an offset label (e.g. ``"offset:0"``)."""
    if not label.startswith(_OFFSET_PREFIX):
        return False
    rest = label[len(_OFFSET_PREFIX) :]
    return rest.lstrip("-").isdigit()


def parse_offset_label(label: str) -> int:
    """Parse ``"offset:N"`` → ``N``.  Raises ValueError if not valid."""
    if not label.startswith(_OFFSET_PREFIX):
        raise ValueError(f"Not an offset label: {label!r}")
    return int(label[len(_OFFSET_PREFIX) :])


def is_year_like(value: int | float) -> bool:
    """
    Check if a numeric value looks like a year (1900-2100 range).
    """
    # Only integers can be years
    if not isinstance(value, int) or isinstance(value, bool):
        return False

    return 1900 <= value <= 2100


def dedupe_labels(labels: list[str]) -> list[str]:
    """
    De-duplicate labels while preserving their original order.
    """
    seen: set[str] = set()
    out: list[str] = []
    for label in labels:
        if label in seen:
            continue
        seen.add(label)
        out.append(label)
    return out


def find_region_config(
    sheet: str, row: int, col: int, region_config: list[RegionConfig] | None = None
) -> RegionConfig | None:
    """
    Find the first matching region configuration for a cell.
    """
    if region_config is None:
        return None
    for config in region_config:
        # Check sheet name
        if config.get("sheet") != sheet:
            continue

        # Check row bounds
        min_row = config.get("min_row")
        max_row = config.get("max_row")
        if min_row is not None and row < min_row:
            continue
        if max_row is not None and row > max_row:
            continue

        # Check column bounds
        min_col = config.get("min_col")
        max_col = config.get("max_col")
        if min_col is not None:
            min_col_idx = openpyxl.utils.cell.column_index_from_string(min_col)
            if col < min_col_idx:
                continue
        if max_col is not None:
            max_col_idx = openpyxl.utils.cell.column_index_from_string(max_col)
            if col > max_col_idx:
                continue

        return config

    return None


def get_labels_from_region_config(
    ws: Worksheet,
    row: int,
    col: int,
    config: RegionConfig,
    offset_maps: dict[int, dict[int, int]] | None = None,
    label_hierarchies: dict[str, dict[int, list[str]]] | None = None,
) -> tuple[list[str], list[str]]:
    """
    Extract row and column labels using explicit region configuration.

    ``offset_maps`` is an optional mapping of ``header_row → {col_idx → offset}``
    produced by :func:`detect_year_offset_headers`.  When present, header cells
    that are year-offset formulas emit ``"offset:<N>"`` labels.

    ``label_hierarchies`` is an optional mapping of
    ``col_letter → {row → [ancestor_labels]}`` produced by
    :func:`build_label_hierarchy`.  When present, ancestor labels are
    prepended to the row labels so that the hierarchy is captured.
    """
    row_labels: list[str] = []
    col_labels: list[str] = []
    offset_maps = offset_maps or {}
    label_hierarchies = label_hierarchies or {}

    # Get row labels from specified label columns
    label_columns = config.get("label_columns", [])
    for col_letter in label_columns:
        label_col_idx = openpyxl.utils.cell.column_index_from_string(col_letter)

        # Prepend ancestor labels from the hierarchy if available
        hierarchy = label_hierarchies.get(col_letter, {})
        ancestors = hierarchy.get(row, [])
        row_labels.extend(ancestors)

        cell_value = ws.cell(row=row, column=label_col_idx).value

        if cell_value is not None:
            if isinstance(cell_value, str):
                text = cell_value.strip()
                if text and is_valid_label(text):
                    row_labels.append(text)
            elif is_year_like(cell_value):
                row_labels.append(str(cell_value))

    # Get column labels from specified header rows
    header_rows = config.get("header_rows", [])
    for header_row in header_rows:
        # Check precomputed offset map first
        hr_offsets = offset_maps.get(header_row, {})
        if col in hr_offsets:
            col_labels.append(f"{_OFFSET_PREFIX}{hr_offsets[col]}")
            continue

        cell_value = ws.cell(row=header_row, column=col).value

        if cell_value is not None:
            if isinstance(cell_value, str):
                text = cell_value.strip()
                if text and is_valid_label(text):
                    col_labels.append(text)
            elif is_year_like(cell_value):
                col_labels.append(str(cell_value))

    return dedupe_labels(row_labels), dedupe_labels(col_labels)


def get_row_labels(ws: Worksheet, row: int, col: int) -> list[str]:
    """
    Scan leftward from a cell to find text labels.
    """
    labels: list[str] = []

    # Start from the column immediately left of the cell
    current_col = col - 1

    while current_col >= 1:
        cell_value = ws.cell(row=row, column=current_col).value

        # Stop if we hit a blank cell
        if cell_value is None or (
            isinstance(cell_value, str) and cell_value.strip() == ""
        ):
            break

        # Collect text values
        if isinstance(cell_value, str):
            text = cell_value.strip()
            # Only append if it passes validation
            if is_valid_label(text):
                labels.append(text)
            # Continue scanning even if invalid (skip over placeholders/errors)
        elif isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
            # Check if it's a year-like integer
            if is_year_like(cell_value):
                labels.append(str(cell_value))
            else:
                # Hit a non-year numeric value - stop scanning
                break
        elif not isinstance(cell_value, bool):
            # Convert other types to string if not numeric/bool
            text = str(cell_value)
            if is_valid_label(text):
                labels.append(text)

        current_col -= 1

    return dedupe_labels(labels)


def get_column_labels(ws: Worksheet, row: int, col: int) -> list[str]:
    """
    Scan upward from a cell to find text labels.
    """
    labels: list[str] = []

    # Start from the row immediately above the cell
    current_row = row - 1

    while current_row >= 1:
        cell_value = ws.cell(row=current_row, column=col).value

        # Stop if we hit a blank cell
        if cell_value is None or (
            isinstance(cell_value, str) and cell_value.strip() == ""
        ):
            break

        # Collect text values
        if isinstance(cell_value, str):
            text = cell_value.strip()
            # Only append if it passes validation
            if is_valid_label(text):
                labels.append(text)
            # Continue scanning even if invalid (skip over placeholders/errors)
        elif isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
            # Check if it's a year-like integer
            if is_year_like(cell_value):
                labels.append(str(cell_value))
            else:
                # Hit a non-year numeric value - stop scanning
                break
        elif not isinstance(cell_value, bool):
            # Convert other types to string if not numeric/bool
            text = str(cell_value)
            if is_valid_label(text):
                labels.append(text)

        current_row -= 1

    return dedupe_labels(labels)


def enrich_graph_with_labels(
    graph: DependencyGraph,
    wb_path: Path,
    *,
    wb_values: openpyxl.Workbook | None = None,
    wb_formulas: openpyxl.Workbook | None = None,
    region_config: list[RegionConfig] | None = None,
) -> dict[str, dict[str, Any]]:
    """
    Enrich nodes in the graph with row and column labels.
    """
    values_wb = wb_values or openpyxl.load_workbook(wb_path, data_only=True)
    formulas_wb = wb_formulas or openpyxl.load_workbook(wb_path)
    try:
        # Cache worksheets by name for efficiency
        worksheets: dict[str, Worksheet] = {}

        # Precompute year-offset maps: sheet → header_row → {col → offset}
        _offset_cache: dict[str, dict[int, dict[int, int]]] = {}

        def _get_offset_maps(
            sheet: str, config: RegionConfig
        ) -> dict[int, dict[int, int]]:
            if sheet not in _offset_cache:
                _offset_cache[sheet] = {}
            sheet_cache = _offset_cache[sheet]
            for hr in config.get("header_rows", []):
                if hr not in sheet_cache:
                    ws_f = (
                        formulas_wb[sheet] if sheet in formulas_wb.sheetnames else None
                    )
                    ws_v = values_wb[sheet] if sheet in values_wb.sheetnames else None
                    if ws_f is not None and ws_v is not None:
                        sheet_cache[hr] = detect_year_offset_headers(
                            ws_f, ws_v, sheet, hr
                        )
                    else:
                        sheet_cache[hr] = {}
            return sheet_cache

        # Precompute label hierarchies: sheet → col_letter → {row → [ancestors]}
        _hierarchy_cache: dict[str, dict[str, dict[int, list[str]]]] = {}

        def _get_label_hierarchies(
            sheet: str, config: RegionConfig
        ) -> dict[str, dict[int, list[str]]]:
            if sheet not in _hierarchy_cache:
                _hierarchy_cache[sheet] = {}
            sheet_cache = _hierarchy_cache[sheet]
            ws_h = worksheets.get(sheet)
            if ws_h is None:
                return sheet_cache
            for col_letter in config.get("label_columns", []):
                if col_letter not in sheet_cache:
                    col_idx = openpyxl.utils.cell.column_index_from_string(col_letter)
                    sheet_cache[col_letter] = build_label_hierarchy(
                        ws_h,
                        col_idx,
                        min_row=config.get("min_row"),
                        max_row=config.get("max_row"),
                    )
            return sheet_cache

        enrichment_results: dict[str, dict[str, Any]] = {}

        for key in graph:
            node = graph.get_node(key)
            if node is None:
                continue

            # Get or cache the worksheet
            if node.sheet not in worksheets:
                if node.sheet in values_wb.sheetnames:
                    worksheets[node.sheet] = values_wb[node.sheet]
                else:
                    continue

            ws = worksheets[node.sheet]

            # Get column index from column letter
            col_idx = openpyxl.utils.cell.column_index_from_string(node.column)

            # Check for region-based configuration first
            matched_region = find_region_config(node.sheet, node.row, col_idx, region_config)

            if matched_region is not None:
                offset_maps = _get_offset_maps(node.sheet, matched_region)
                label_hierarchies = _get_label_hierarchies(
                    node.sheet, matched_region
                )
                row_labels, col_labels = get_labels_from_region_config(
                    ws, node.row, col_idx, matched_region, offset_maps,
                    label_hierarchies=label_hierarchies,
                )
            else:
                # Fall back to heuristic scanning
                row_labels = get_row_labels(ws, node.row, col_idx)
                col_labels = get_column_labels(ws, node.row, col_idx)

            # Store in node metadata
            node.metadata["row_labels"] = row_labels
            node.metadata["column_labels"] = col_labels

            # Track for reporting (all nodes, not just those with labels)
            enrichment_results[key] = {
                "sheet": node.sheet,
                "address": node.address,
                "row_labels": row_labels,
                "column_labels": col_labels,
                "is_leaf": node.is_leaf,
            }

        return enrichment_results

    finally:
        if wb_values is None:
            values_wb.close()
        if wb_formulas is None:
            formulas_wb.close()


def export_enrichment_audit(
    graph: DependencyGraph,
    enrichment_results: dict[str, dict[str, Any]],
    output_path: Path,
) -> None:
    """
    Export enrichment results to a JSON file for auditing.
    """
    # Compute per-sheet statistics
    sheet_stats: dict[str, dict[str, int]] = defaultdict(
        lambda: {"total": 0, "with_row": 0, "with_col": 0, "with_any": 0, "none": 0}
    )

    # Group cells by sheet
    cells_by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for key, data in enrichment_results.items():
        sheet = data["sheet"]
        sheet_stats[sheet]["total"] += 1

        has_row = bool(data["row_labels"])
        has_col = bool(data["column_labels"])

        if has_row:
            sheet_stats[sheet]["with_row"] += 1
        if has_col:
            sheet_stats[sheet]["with_col"] += 1
        if has_row or has_col:
            sheet_stats[sheet]["with_any"] += 1
        else:
            sheet_stats[sheet]["none"] += 1

        cells_by_sheet[sheet].append(
            {
                "key": key,
                "address": data["address"],
                "row_labels": data["row_labels"],
                "column_labels": data["column_labels"],
                "is_leaf": data["is_leaf"],
            }
        )

    # Sort cells within each sheet by address
    for sheet in cells_by_sheet:
        cells_by_sheet[sheet].sort(key=lambda x: x["address"])

    # Build output structure
    total_nodes = len(graph)
    nodes_with_any = sum(s["with_any"] for s in sheet_stats.values())

    output = {
        "summary": {
            "total_nodes": total_nodes,
            "nodes_with_any_label": nodes_with_any,
            "nodes_without_labels": total_nodes - nodes_with_any,
            "nodes_with_row_labels": sum(s["with_row"] for s in sheet_stats.values()),
            "nodes_with_column_labels": sum(
                s["with_col"] for s in sheet_stats.values()
            ),
        },
        "by_sheet": {
            sheet: {
                "statistics": dict(sheet_stats[sheet]),
                "cells": cells_by_sheet[sheet],
            }
            for sheet in sorted(sheet_stats.keys())
        },
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)


def discover_formula_cells_in_rows(
    wb_path: Path,
    sheet_name: str,
    rows: list[int],
) -> list[str]:
    """
    Scan specified rows and return sheet-qualified addresses for formula cells.

    Includes every cell that contains a formula (value starts with '='). Cached
    values are not required to be numeric, so discovery works when the template
    has no pre-filled data (e.g. base year empty, formulas evaluate to errors).
    """
    wb_formulas = openpyxl.load_workbook(wb_path)
    try:
        if sheet_name not in wb_formulas.sheetnames:
            print(f"  Warning: Sheet '{sheet_name}' not found")
            return []

        ws_formulas = wb_formulas[sheet_name]
        targets: list[str] = []

        for row in rows:
            if hasattr(ws_formulas, "_cells"):
                relevant_cols = sorted(c for r, c in ws_formulas._cells if r == row)
            else:
                max_col = ws_formulas.max_column or 1
                relevant_cols = range(1, max_col + 1)

            for col_idx in relevant_cols:
                cell_formula = ws_formulas.cell(row=row, column=col_idx)
                if isinstance(
                    cell_formula.value, str
                ) and cell_formula.value.startswith("="):
                    col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
                    targets.append(f"{sheet_name}!{col_letter}{row}")

        return targets
    finally:
        wb_formulas.close()
