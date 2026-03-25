#!/usr/bin/env python3
"""
Shared pipeline utilities for LIC-DSF export and input grouping.
"""

from __future__ import annotations

import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import openpyxl.utils.cell
from excel_grapher.exporter import CodeGenerator
from excel_grapher.grapher import DependencyGraph, Node, create_dependency_graph
from openpyxl import Workbook

from openpyxl.worksheet.worksheet import Worksheet

from excel_grapher.grapher import DynamicRefConfig
_SAFE_SHEET_NAME_RE = re.compile(r"^[A-Za-z_][0-9A-Za-z_]*$")

GRAPH_JSON_VERSION = 1


def _json_safe(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return str(value)
        return value
    if isinstance(value, str):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(v) for v in value]
    return str(value)


def dependency_graph_to_dict(graph: DependencyGraph) -> dict[str, Any]:
    """
    Serialize a DependencyGraph to a JSON-compatible dict (nodes, edges, metadata).
    """
    nodes_out: dict[str, Any] = {}
    for key in graph:
        node = graph.get_node(key)
        if node is None:
            continue
        nodes_out[key] = {
            "sheet": node.sheet,
            "column": node.column,
            "row": node.row,
            "formula": node.formula,
            "normalized_formula": node.normalized_formula,
            "value": _json_safe(node.value),
            "is_leaf": node.is_leaf,
            "metadata": _json_safe(node.metadata),
        }
    edges_out: list[dict[str, Any]] = []
    for key in graph:
        for dep in sorted(graph.dependencies(key)):
            raw_attrs = dict(graph.edge_attrs(key, dep))
            attrs_out: dict[str, Any] = {}
            for ak, av in raw_attrs.items():
                if ak == "guard":
                    attrs_out[ak] = None if av is None else str(av)
                else:
                    attrs_out[ak] = _json_safe(av)
            edges_out.append({"from": key, "to": dep, "attrs": attrs_out})
    return {
        "version": GRAPH_JSON_VERSION,
        "nodes": nodes_out,
        "edges": edges_out,
    }


def export_graph_json(graph: DependencyGraph, path: Path) -> None:
    """Write `dependency_graph_to_dict(graph)` to `path` as UTF-8 JSON."""
    path.write_text(
        json.dumps(dependency_graph_to_dict(graph), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def _format_sheet_name(sheet: str) -> str:
    """
    Format a sheet name for sheet-qualified Excel addresses.

    Only quote sheet names when needed (e.g., spaces or punctuation), and escape
    embedded single quotes by doubling them.
    """
    if _SAFE_SHEET_NAME_RE.match(sheet):
        return sheet
    escaped = sheet.replace("'", "''")
    return f"'{escaped}'"


def _format_address(sheet: str, a1: str) -> str:
    return f"{_format_sheet_name(sheet)}!{a1}"


def _is_blank_value(value: object) -> bool:
    return value is None or value == ""


def discover_targets(export_ranges: list) -> list[str]:
    """
    Discover graph targets from explicit range specifications.

    Targets are derived from the template's EXPORT_RANGES and expanded via
    sheet-qualified A1 ranges.
    """
    from .lic_dsf_config import discover_targets_from_ranges

    return discover_targets_from_ranges(export_ranges)


def build_graph(
    workbook: Path,
    targets: list[str],
    max_depth: int,
    *,
    wb_formulas: Workbook | None = None,
    dynamic_refs: DynamicRefConfig | None = None,
) -> DependencyGraph:
    source = wb_formulas if wb_formulas is not None else workbook
    return create_dependency_graph(
        source,
        targets,
        load_values=False,
        max_depth=max_depth,
        dynamic_refs=dynamic_refs,
        use_cached_dynamic_refs=False,
    )


def populate_leaf_values(
    graph: DependencyGraph,
    workbook: Path,
    *,
    wb_values: Workbook | None = None,
) -> None:
    """
    Populate values for leaf (non-formula) nodes from cached workbook values.

    Code generation only needs `node.value` for leaf cells; formulas are emitted
    from `node.formula`.
    """
    wb = wb_values or openpyxl.load_workbook(workbook, data_only=True)
    try:
        worksheets: dict[str, Worksheet] = {}
        for key in graph:
            node = graph.get_node(key)
            if node is None or node.formula is not None:
                continue
            if node.sheet not in worksheets:
                if node.sheet not in wb.sheetnames:
                    continue
                worksheets[node.sheet] = wb[node.sheet]
            ws = worksheets[node.sheet]
            col_idx = openpyxl.utils.cell.column_index_from_string(node.column)
            node.value = ws.cell(row=node.row, column=col_idx).value
    finally:
        if wb_values is None:
            wb.close()


def iter_string_constant_addresses(
    graph: DependencyGraph, exclude_addresses: set[str] | None = None
) -> list[str]:
    exclude = exclude_addresses or set()
    ranges: list[str] = []
    for key in graph:
        node = graph.get_node(key)
        if node is None or node.formula is not None or not node.is_leaf:
            continue
        if not isinstance(node.value, str):
            continue
        address = _format_address(node.sheet, node.address)
        if address in exclude:
            continue
        ranges.append(address)
    return list(dict.fromkeys(ranges))


def classify_input_addresses(
    graph: DependencyGraph,
    targets: list[str],
    *,
    constant_types: set[str] | None = None,
    constant_ranges: list[str] | None = None,
    constant_blanks: bool = False,
    blank_excludes: set[str] | None = None,
    attach_to_graph: bool = True,
) -> set[str]:
    inputs, _constants = CodeGenerator(graph).classify_leaf_nodes(
        targets,
        constant_types=constant_types,
        constant_ranges=constant_ranges,
        constant_blanks=constant_blanks,
        attach_to_graph=attach_to_graph,
    )
    input_addresses = {str(address) for address in inputs}
    if not blank_excludes:
        return input_addresses

    address_to_node: dict[str, Node] = {}
    for key in graph:
        node = graph.get_node(key)
        if node is None:
            continue
        address_to_node[_format_address(node.sheet, node.address)] = node

    for address in blank_excludes:
        if address in input_addresses:
            continue
        node = address_to_node.get(address)
        if node is None:
            continue
        if node.formula is not None or not node.is_leaf:
            continue
        if _is_blank_value(node.value):
            input_addresses.add(address)
    return input_addresses


def enrich_graph(
    graph: DependencyGraph,
    workbook: Path,
    *,
    wb_values: Workbook | None = None,
    wb_formulas: Workbook | None = None,
    region_config: list | None = None,
) -> dict[str, dict[str, object]]:
    from .lic_dsf_labels import enrich_graph_with_labels

    return enrich_graph_with_labels(
        graph,
        workbook,
        wb_values=wb_values,
        wb_formulas=wb_formulas,
        region_config=region_config,
    )


def export_enrichment_audit(
    graph: DependencyGraph, enrichment_results: dict[str, dict[str, object]], path: Path
) -> None:
    from .lic_dsf_labels import export_enrichment_audit

    export_enrichment_audit(graph, enrichment_results, path)
